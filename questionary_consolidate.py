import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from deep_translator import GoogleTranslator

def get_cell_display_value(cell):
    """
    Best-effort text for detecting question cells or building plain strings.
    Note: This is only used for reading questions and simple text checks.
    """
    if cell.value is None:
        return None

    # Use the raw value; don't try to compute Excel's displayed string here
    return str(cell.value).strip()

def get_cell_value_and_format(cell):
    """
    Return the raw cell value and its number format so we can write the same
    value back and let Excel render formatting (percentages, dates, etc.).
    """
    if cell.value is None:
        return None, None

    return cell.value, (cell.number_format if cell.number_format else None)

def _format_value_for_text(value, number_format):
    """Render a value to string respecting common formats (esp. percentages).
    This is only used when we must concatenate into plain text.
    """
    if value is None:
        return ""

    # Percent handling: Excel displays value*100 followed by %
    if number_format and '%' in str(number_format) and isinstance(value, (int, float)):
        fmt = str(number_format)
        # Heuristic: detect decimals from format like 0%, 0.0%, 0.00%
        decimals = 0
        if '.' in fmt:
            try:
                decimals = len(fmt.split('%')[0].split('.')[-1])
            except Exception:
                decimals = 0
        scaled = value * 100
        formatted = f"{scaled:.{decimals}f}%"
        # Trim trailing .0 if decimals=0 (safety)
        if decimals == 0:
            formatted = formatted.replace('.0%', '%')
        return formatted

    # Fallback: plain string
    return str(value)

def format_regular_data(table_data):
    """
    Format non-tabular data - preserve tuple structure for format application
    """
    if not table_data or len(table_data) == 0:
        return ""
    
    # Handle new tuple structure (values, formats) - keep the structure for format preservation
    if isinstance(table_data[0], tuple):
        # Keep as-is so later writing can preserve cell formats precisely
        return table_data
    
    # Handle legacy list structure
    if isinstance(table_data[0], list):
        # Convert structured data to string format
        formatted_lines = []
        for row in table_data:
            non_empty = [str(cell).strip() for cell in row if str(cell).strip()]
            if non_empty:
                formatted_lines.append(" ".join(non_empty))
        return "\n".join(formatted_lines).strip()
    else:
        # Handle old string format
        if isinstance(table_data[0], str):
            return "\n".join(table_data).strip()
        else:
            # Fallback for any other format - convert to strings
            return "\n".join([str(item) for item in table_data]).strip()

def translate_to_english_if_needed(text):
    try:
        if not text or not isinstance(text, str):
            return text
        # Very basic heuristic: check for presence of Chinese characters
        if any('\u4e00' <= char <= '\u9fff' for char in text):
            translated = GoogleTranslator(source='auto', target='en').translate(text)
            return translated
        return text
    except Exception as e:
        print(f"Translation error: {e}")
        return text

def translate_table_data(answer_parts):
    """
    Translate tabular data while preserving structure
    """
    translated_parts = []
    for part in answer_parts:
        if isinstance(part, list):
            # This is already structured data (list of cells)
            translated_cells = [translate_to_english_if_needed(str(cell)) for cell in part]
            translated_cells = [str(cell) for cell in part]
            translated_parts.append(translated_cells)
        elif '|' in part:
            # Split by separator, translate each cell, then rejoin
            cells = [cell.strip() for cell in part.split('|')]
            translated_cells = [translate_to_english_if_needed(cell) for cell in cells]
            translated_cells = [cell for cell in cells]
            translated_parts.append(translated_cells)
        else:
            # Single cell or non-tabular data
            translated_parts.append([translate_to_english_if_needed(str(part))])
            translated_parts.append([str(part)])
    return translated_parts

def is_tabular_data(table_data):
    """
    Determine if the extracted data is tabular based on structure
    """
    if not table_data or len(table_data) <= 1:
        return False
    
    # Handle new tuple structure (values, formats)
    if isinstance(table_data[0], tuple):
        # Extract just the values for analysis
        rows = [row[0] for row in table_data]  # Get values from tuples
        col_counts = [len([cell for cell in row if cell and str(cell).strip()]) for row in rows]
        # If most rows have data in multiple columns, it's tabular
        return max(col_counts) > 1 and len([c for c in col_counts if c > 1]) > 1
    
    # Check if we have structured data (list of lists) - legacy format
    if isinstance(table_data[0], list):
        # Check if multiple rows have similar structure
        col_counts = [len([cell for cell in row if cell and str(cell).strip()]) for row in table_data]
        # If most rows have data in multiple columns, it's tabular
        return max(col_counts) > 1 and len([c for c in col_counts if c > 1]) > 1
    
    # Fallback for old string-based detection
    if isinstance(table_data[0], str):
        separator_counts = [part.count('|') for part in table_data]
        return len(set(separator_counts)) <= 2 and max(separator_counts) > 1
    
    return False

def parse_tabular_data(table_data):
    """
    Parse structured table data for proper Excel placement
    """
    if not table_data:
        return []
    
    # If it's already structured data (list of lists), just clean it up
    if isinstance(table_data[0], list):
        cleaned_data = []
        for row in table_data:
            cleaned_row = [str(cell).strip() if cell else "" for cell in row]
            cleaned_data.append(cleaned_row)
        return cleaned_data
    
    # Fallback for pipe-separated strings
    rows = []
    max_cols = 0
    
    for part in table_data:
        if isinstance(part, str):
            cols = [col.strip() for col in part.split('|')]
            rows.append(cols)
            max_cols = max(max_cols, len(cols))
    
    # Pad rows to same length
    for row in rows:
        while len(row) < max_cols:
            row.append("")
    
    return rows

def extract_complete_answer_from_merged_area(ws, start_row, start_col=4, end_col=12, max_look_ahead=30):
    """
    Enhanced function to extract complete table data from merged Excel area
    This preserves the original table structure by maintaining column positions
    Returns: (table_data, last_processed_row)
    """
    table_data = []
    current_row = start_row
    consecutive_empty_rows = 0
    max_consecutive_empty = 2  # Reduce from 3 to 2 for more conservative extraction  # Allow up to 3 consecutive empty rows before stopping
    last_data_row = start_row
    
    print(f"  Extracting from row {start_row}, cols {start_col}-{end_col}")
    
    while current_row <= min(ws.max_row, start_row + max_look_ahead):
        # Check if there's a new question in column C at this row
        question_cell_obj = ws.cell(row=current_row, column=3)
        question_cell = get_cell_display_value(question_cell_obj)
        
        if question_cell and current_row > start_row:
            # Found a new question, stop here
            print(f"    Found new question at row {current_row}, stopping extraction")
            break
            
        # Get all values in the row range, preserving column positions
        row_values = []
        row_formats = []
        has_data = False
        
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=current_row, column=col)
            cell_value, cell_format = get_cell_value_and_format(cell)

            # Treat non-empty raw values as data; keep raw value types
            if cell_value not in (None, ""):
                row_values.append(cell_value)
                row_formats.append(cell_format)
                has_data = True
            else:
                row_values.append("")
                row_formats.append(None)
        
        if has_data:
            # Found data in this row
            consecutive_empty_rows = 0
            table_data.append((row_values, row_formats))  # Store as tuple of (values, formats)
            last_data_row = current_row
            print(f"    Row {current_row}: {len([v for v in row_values if v])} non-empty cells")
        else:
            # Empty row
            consecutive_empty_rows += 1
            if table_data and consecutive_empty_rows >= max_consecutive_empty:
                # Stop if we've found some content and hit too many empty rows
                print(f"    Stopping at row {current_row} due to {consecutive_empty_rows} consecutive empty rows")
                break
        
        current_row += 1
    
    print(f"  Extracted {len(table_data)} rows of structured data")
    return table_data, last_data_row

def detect_merged_regions_for_question(ws, question_row, start_col=4, end_col=12):
    """
    Detect if there are merged cells in the answer area and return the extent
    """
    merged_ranges = []
    for merged_range in ws.merged_cells.ranges:
        # Check if the merged range intersects with our answer area
        if (merged_range.min_row <= question_row <= merged_range.max_row and
            merged_range.min_col <= end_col and merged_range.max_col >= start_col):
            merged_ranges.append(merged_range)
    
    if merged_ranges:
        # Find the extent of all relevant merged ranges
        min_row = min(r.min_row for r in merged_ranges)
        max_row = max(r.max_row for r in merged_ranges)
        return min_row, max_row
    
    return question_row, question_row

def create_enhanced_excel(master_data, all_suppliers):
    """
    Create Excel file with reserved columns for each supplier (8 columns each)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated Data"
    
    # Define styling
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    sub_header_font = Font(bold=True, color="366092", size=11)
    data_font = Font(name="Calibri", size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Calculate column layout
    COLS_PER_SUPPLIER = 8
    SPACING_COLS = 2  # Spacing columns between suppliers
    base_cols = 2  # Sheet Name + Question columns
    
    # Create headers
    current_row = 1
    
    # Main headers - merge across supplier columns
    ws.cell(row=current_row, column=1, value="Sheet Name")
    ws.cell(row=current_row, column=2, value="Question")
    
    col_start = base_cols + 1
    for i, supplier in enumerate(all_suppliers):
        # Merge cells for supplier header
        start_col = col_start
        end_col = col_start + COLS_PER_SUPPLIER - 1
        
        ws.merge_cells(f"{get_column_letter(start_col)}{current_row}:{get_column_letter(end_col)}{current_row}")
        ws.cell(row=current_row, column=start_col, value=supplier)
        
        # Move to next supplier position (including spacing)
        col_start = end_col + 1 + SPACING_COLS
    
    # Apply header styling
    for col in range(1, col_start):
        cell = ws.cell(row=current_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Sub-headers for supplier columns
    current_row += 1
    ws.cell(row=current_row, column=1, value="")  # Empty for sheet name
    ws.cell(row=current_row, column=2, value="")  # Empty for question

    col_start = base_cols + 1
    for supplier in all_suppliers:
        for i in range(COLS_PER_SUPPLIER):
            col_num = col_start + i
            ws.cell(row=current_row, column=col_num, value=f"Col {i+1}")
            cell = ws.cell(row=current_row, column=col_num)
            cell.font = sub_header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        # Add spacing columns (empty)
        for s in range(SPACING_COLS):
            col_num = col_start + COLS_PER_SUPPLIER + s
            ws.cell(row=current_row, column=col_num, value="")
            cell = ws.cell(row=current_row, column=col_num)
            cell.border = border
        # Move to next supplier position (including spacing)
        col_start += COLS_PER_SUPPLIER + SPACING_COLS

    # Style the first two columns in sub-header row
    for col in [1, 2]:
        cell = ws.cell(row=current_row, column=col)
        cell.font = sub_header_font
        cell.border = border
    
    # Data rows
    current_row += 1

    for (sheet, question), supplier_answers in master_data.items():
        row_start = current_row
        max_rows_needed = 1

        # First, determine how many rows we need for this question
        for supplier in all_suppliers:
            answer = supplier_answers.get(supplier, "")
            if answer:
                # Handle list-based tabular data (from is_tabular_data detection)
                if isinstance(answer, list) and len(answer) > 0:
                    # Check if this is new tuple format or legacy list format
                    if isinstance(answer[0], tuple):
                        # New format with values and formats - count the tuples
                        max_rows_needed = max(max_rows_needed, len(answer))
                    elif isinstance(answer[0], list):
                        # Legacy structured tabular data (list of lists)
                        max_rows_needed = max(max_rows_needed, len(answer))
                elif isinstance(answer, str):
                    answer_parts = answer.split('\n')
                    # Check if it's tabular data that was processed
                    if '|' in answer and ('┌' in answer or '├' in answer):
                        # This is a formatted table, count actual data rows
                        data_lines = [line for line in answer_parts if line.strip() and not line.startswith('┌') and not line.startswith('├') and not line.startswith('└') and '─' in line]
                        max_rows_needed = max(max_rows_needed, len(answer_parts))
                    else:
                        # Check if original data was tabular
                        original_parts = answer.split('\n') if isinstance(answer, str) else []
                        if any('|' in part for part in original_parts):
                            # Parse as table
                            table_data = parse_tabular_data([part for part in original_parts if '|' in part])
                            max_rows_needed = max(max_rows_needed, len(table_data))
                        else:
                            max_rows_needed = max(max_rows_needed, len(answer_parts))
                else:
                    # If it's some other format, convert to string and count lines
                    answer_str = str(answer) if answer else ""
                    answer_parts = answer_str.split('\n') if answer_str else []
                    max_rows_needed = max(max_rows_needed, len(answer_parts))

        # Set sheet name and question (merge across all rows if multi-row)
        if max_rows_needed > 1:
            ws.merge_cells(f"A{row_start}:A{row_start + max_rows_needed - 1}")
            ws.merge_cells(f"B{row_start}:B{row_start + max_rows_needed - 1}")

        ws.cell(row=row_start, column=1, value=sheet)
        ws.cell(row=row_start, column=2, value=question)

        # Style sheet and question cells
        for col in [1, 2]:
            cell = ws.cell(row=row_start, column=col)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = border

        # Fill supplier data
        col_start = base_cols + 1
        for supplier in all_suppliers:
            answer = supplier_answers.get(supplier, "")

            if answer:
                # Handle list-based data (both tabular and regular with formats)
                if isinstance(answer, list) and len(answer) > 0:
                    # Check if this is new tuple format (values, formats) or legacy list format
                    if isinstance(answer[0], tuple):
                        # New format with values and formats - could be tabular or regular
                        if is_tabular_data(answer):
                            print(f"    Processing tabular data with formats for {supplier}: {len(answer)} rows")
                            
                            # Place table data in grid with proper formatting
                            for row_idx, (row_values, row_formats) in enumerate(answer):
                                for col_idx, (cell_value, cell_format) in enumerate(zip(row_values[:COLS_PER_SUPPLIER], row_formats[:COLS_PER_SUPPLIER])):
                                    # Place all values, including empty strings to maintain structure
                                    target_row = row_start + row_idx
                                    target_col = col_start + col_idx

                                    # Write raw value when possible so Excel renders formatting
                                    set_value = cell_value if cell_value is not None else ""
                                    ws.cell(row=target_row, column=target_col, value=set_value)

                                    cell = ws.cell(row=target_row, column=target_col)
                                    cell.font = data_font
                                    cell.alignment = Alignment(horizontal='left', vertical='center')
                                    cell.border = border

                                    # Apply the original number format if available and value is not empty
                                    if cell_format and set_value != "":
                                        cell.number_format = cell_format
                        else:
                            # Regular data with formats - extract values and formats for proper display
                            print(f"    Processing regular data with formats for {supplier}: {len(answer)} rows")
                            all_values = []
                            all_formats = []
                            
                            for row_values, row_formats in answer:
                                for cell_value, cell_format in zip(row_values, row_formats):
                                    # Render to text respecting percent formats
                                    text = _format_value_for_text(cell_value, cell_format)
                                    if text.strip():
                                        all_values.append(text.strip())
                                        all_formats.append(cell_format)
                            
                            if all_values:
                                # For regular data, merge across all supplier columns and apply format to first non-empty value
                                combined_text = " ".join(all_values)
                                ws.merge_cells(f"{get_column_letter(col_start)}{row_start}:{get_column_letter(col_start + COLS_PER_SUPPLIER - 1)}{row_start + max_rows_needed - 1}")
                                ws.cell(row=row_start, column=col_start, value=combined_text)
                                
                                cell = ws.cell(row=row_start, column=col_start)
                                cell.font = data_font
                                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                                cell.border = border
                                # Do not apply number_format to a combined text cell
                    
                    elif isinstance(answer[0], list):
                        # Legacy structured tabular data (list of lists)
                        print(f"    Processing legacy tabular data for {supplier}: {len(answer)} rows")
                        
                        # Place table data in grid
                        for row_idx, row_data in enumerate(answer):
                            for col_idx, cell_value in enumerate(row_data[:COLS_PER_SUPPLIER]):
                                target_row = row_start + row_idx
                                target_col = col_start + col_idx
                                # Place all values, including empty strings to maintain table structure
                                display_value = str(cell_value).strip() if cell_value else ""
                                ws.cell(row=target_row, column=target_col, value=display_value)

                                cell = ws.cell(row=target_row, column=target_col)
                                cell.font = data_font
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                                cell.border = border
                
                elif isinstance(answer, str):
                    # Check if this looks like processed tabular data or original tabular data
                    if ('|' in answer and not ('┌' in answer or '├' in answer)) or answer.count('|') > answer.count('\n'):
                        # This is raw tabular data, parse it properly
                        original_parts = [part.strip() for part in answer.split('\n') if part.strip()]
                        table_parts = [part for part in original_parts if '|' in part]

                        if table_parts:
                            table_data = parse_tabular_data(table_parts)
                            print(f"    Processing string-based tabular data for {supplier}: {len(table_data)} rows")

                            # Place table data in grid
                            for row_idx, row_data in enumerate(table_data):
                                for col_idx, cell_value in enumerate(row_data[:COLS_PER_SUPPLIER]):
                                    target_row = row_start + row_idx
                                    target_col = col_start + col_idx
                                    # Place all values, including empty strings to maintain table structure
                                    display_value = cell_value.strip() if cell_value else ""
                                    ws.cell(row=target_row, column=target_col, value=display_value)

                                    cell = ws.cell(row=target_row, column=target_col)
                                    cell.font = data_font
                                    cell.alignment = Alignment(horizontal='left', vertical='center')
                                    cell.border = border
                        else:
                            # Non-tabular multi-line data
                            lines = answer.split('\n')
                            for line_idx, line in enumerate(lines[:max_rows_needed]):
                                if line.strip():
                                    target_row = row_start + line_idx
                                    ws.merge_cells(f"{get_column_letter(col_start)}{target_row}:{get_column_letter(col_start + COLS_PER_SUPPLIER - 1)}{target_row}")
                                    ws.cell(row=target_row, column=col_start, value=line)

                                    cell = ws.cell(row=target_row, column=col_start)
                                    cell.font = data_font
                                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                                    cell.border = border
                    else:
                        # Single value or simple text - merge across all supplier columns
                        ws.merge_cells(f"{get_column_letter(col_start)}{row_start}:{get_column_letter(col_start + COLS_PER_SUPPLIER - 1)}{row_start + max_rows_needed - 1}")
                        ws.cell(row=row_start, column=col_start, value=answer)

                        cell = ws.cell(row=row_start, column=col_start)
                        cell.font = data_font
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                        cell.border = border

            # Add borders to empty cells in this supplier's columns
            for col_offset in range(COLS_PER_SUPPLIER):
                for row_offset in range(max_rows_needed):
                    target_row = row_start + row_offset
                    target_col = col_start + col_offset
                    cell = ws.cell(row=target_row, column=target_col)
                    if not cell.border.left.style:  # Only add border if not already set
                        cell.border = border
            # Add borders to spacing columns
            for s in range(SPACING_COLS):
                for row_offset in range(max_rows_needed):
                    target_row = row_start + row_offset
                    target_col = col_start + COLS_PER_SUPPLIER + s
                    cell = ws.cell(row=target_row, column=target_col)
                    cell.border = border

            col_start += COLS_PER_SUPPLIER + SPACING_COLS

        current_row += max_rows_needed + 3  # Add 3 empty rows between questions
    
    # Adjust column widths
    # Fixed width for first two columns
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40

    # Set width for supplier columns and spacing columns
    col_start = base_cols + 1
    for supplier in all_suppliers:
        for i in range(COLS_PER_SUPPLIER):
            col_letter = get_column_letter(col_start + i)
            ws.column_dimensions[col_letter].width = 15
        # Set width for spacing columns
        for s in range(SPACING_COLS):
            col_letter = get_column_letter(col_start + COLS_PER_SUPPLIER + s)
            ws.column_dimensions[col_letter].width = 5
        col_start += COLS_PER_SUPPLIER + SPACING_COLS
    
    return wb

folder = "questionaries"
master_data = {}  # key: (sheet_name, question_text), value: {supplier: answer}

print("Processing files...")
for file in os.listdir(folder):
    if not file.endswith(".xlsx"):
        continue

    supplier = os.path.splitext(file)[0].split('--')[-1]
    filepath = os.path.join(folder, file)
    print(f"Processing {supplier}...")
    
    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"Error loading {file}: {e}")
        continue

    for sheetname in wb.sheetnames[1:]:  # Skip Sheet1 (index 0)
        ws = wb[sheetname]
        print(f"  Processing sheet: {sheetname}")
        
        row = 1
        while row <= ws.max_row:
            q_cell = ws[f"C{row}"]
            question = get_cell_display_value(q_cell)
            if question:
                print(f"  Found question at row {row}: {question[:50]}...")

                # Extract answer data starting from the same row as the question
                # Questions are in column C, answers are in columns D-L (4-12)
                table_data, last_processed_row = extract_complete_answer_from_merged_area(
                    ws, 
                    start_row=row,  # Start from the question row itself
                    start_col=4, 
                    end_col=12, 
                    max_look_ahead=30
                )

                # Smart formatting based on data structure
                if is_tabular_data(table_data):
                    # For tabular data, we need to preserve the exact structure
                    # Convert to a format that maintains column alignment
                    if isinstance(table_data[0], tuple):
                        # New format with values and formats - keep the complete structure
                        formatted_answer = table_data  # Keep as list of tuples
                    elif isinstance(table_data[0], list):
                        # Legacy format - Store as structured data for proper placement
                        formatted_answer = table_data  # Keep as list of lists
                    else:
                        # Fallback to string format
                        formatted_answer = "\n".join(table_data)
                    print(f"    Detected as tabular data with {len(table_data)} rows")
                else:
                    formatted_answer = format_regular_data(table_data)
                    print(f"    Detected as regular data")

                # store the data
                key = (sheetname, question)
                if key not in master_data:
                    master_data[key] = {}
                master_data[key][supplier] = formatted_answer
                
                # Move to next row after the answer data to avoid misalignment
                if table_data:
                    # Skip to after the last processed row
                    row = last_processed_row + 1
                    print(f"    Advanced to row {row} after processing answer data")
                else:
                    # If no data found, just advance by 1
                    row += 1
            else:
                row += 1

print("Building consolidated data...")

# Build data
all_suppliers = sorted({supplier for answers in master_data.values() for supplier in answers})

# Create enhanced Excel file
print("Creating enhanced Excel file with reserved columns...")
wb = create_enhanced_excel(master_data, all_suppliers)

# Save file
output_path = "Consolidated Questionaries WO Transation_v3.xlsx"
wb.save(output_path)

print(f"Done! Enhanced file saved at: {output_path}")
print(f"Processed {len(master_data)} questions from {len(all_suppliers)} suppliers.")
print(f"Each supplier has {8} reserved columns for proper table display.")

# Create summary
print("Creating summary sheet...")
summary_data = []
for sheet_name in set(key[0] for key in master_data.keys()):
    sheet_questions = [key for key in master_data.keys() if key[0] == sheet_name]
    summary_data.append({
        'Sheet Name': sheet_name,
        'Number of Questions': len(sheet_questions),
        'Suppliers with Answers': len(set(supplier for key in sheet_questions for supplier in master_data[key].keys()))
    })

# Add summary sheet
summary_ws = wb.create_sheet("Summary")
summary_ws.append(['Sheet Name', 'Number of Questions', 'Suppliers with Answers'])

for data in summary_data:
    summary_ws.append([data['Sheet Name'], data['Number of Questions'], data['Suppliers with Answers']])

# Style summary sheet
for row in summary_ws.iter_rows():
    for cell in row:
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
        if cell.row == 1:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

wb.save(output_path)
print("Summary sheet added successfully!")