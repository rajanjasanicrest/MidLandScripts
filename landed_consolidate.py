import pandas as pd
import numpy as np
import re
import os
import time
from datetime import datetime
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

start_time = time.time()

wierd_list = ["7000-04-06","7000-08-04","7000-08-16","7000-10-06","7000-12-16","7002-06-04","7002-06-06","7002-06-08","7002-08-06","7002-08-08","7002-08-12","7002-12-08","7003-04-04","7003-06-06","7003-06-08","7003-08-08","7003-12-12","7004-04-02","7004-04-04","7004-04-06","7004-04-08","7004-06-04","7004-06-06","7004-06-08","7004-08-06","7004-08-08","7004-08-10","7004-08-12","7004-10-08","7004-10-10","7004-10-12","7004-12-12","7004-12-16","7012-05-04","7022-06-08","7022-10-06","7022-10-10","7022-12-08","7022-12-16","7032-02-02","7032-02-04","7032-02-06","7032-04-02","7032-04-04","7032-04-06","7032-04-08","7032-04-12","7032-06-02","7032-06-04","7032-06-06","7032-06-08","7032-06-12","7032-06-16","7032-08-04","7032-08-06","7032-08-08","7032-08-12","7032-08-16","7032-12-04","7032-12-06","7032-12-08","7032-12-12","7032-12-16","7033-02-04","7033-04-02","7033-04-04","7033-04-06","7033-06-04","7033-06-06","7033-06-08","7033-08-04","7033-08-06","7033-08-08","7033-08-12","7033-12-04","7033-12-06","7033-12-08","7033-12-12","7034-04-02","7034-04-04","7034-06-06","7034-08-06","7034-08-08","7034-12-12","7040-04-04","7040-08-12","7040-12-12","7042-02-04","7042-04-06","7042-04-08","7042-06-04","7042-06-08","7042-08-06","7042-12-08","7042-12-16","7062-04-04","7062-06-04","7062-06-06","7062-06-08","7062-08-06","7062-08-10","7062-10-08","7062-10-12","7062-12-06","7062-12-08","7062-12-12","7062-12-16","7202-04-06","7202-04-08","7202-05-04","7202-06-02","7202-08-04","7202-10-06","7202-10-12","7204-04-04","7204-06-06","7204-12-12","7802-04-06","7802-06-04","7802-06-06","7802-06-08","7802-08-06","9000-02-02","9000-04-02","9000-04-04","9000-06-04","9000-06-06","9000-08-04","9000-08-06","9000-08-08","9000-10-10","9000-12-08","9000-12-12","9001-06-02","9001-06-04","9001-08-04","9001-08-06","9001-12-06","9001-12-08","9020-04-02","9020-04-04","9020-04-06","9020-04-08","9020-06-04","9020-06-06","9020-06-08","9020-06-12","9020-06-16","9020-08-04","9020-08-06","9020-08-08","9020-08-12","9020-10-08","9020-10-12","9020-12-04","9020-12-06","9020-12-08","9020-12-12","9020-12-16","9022-04-02","9022-04-04","9022-06-04","9022-06-06","9022-08-04","9022-08-06","9022-08-08","9022-10-08","9022-10-10","9022-12-04","9022-12-06","9022-12-08","9022-12-10","9022-12-12","9023-02-04","9023-02-08","9023-04-06","9023-04-08","9023-04-12","9023-06-02","9023-06-04","9023-06-08","9023-06-16","9024-02-04","9024-02-06","9024-04-02","9024-04-04","9024-04-06","9024-04-08","9024-04-12","9024-06-02","9024-06-04","9024-06-06","9024-06-08","9024-06-12","9024-08-04","9024-08-06","9024-08-08","9024-08-10","9024-08-12","9024-08-16","9024-10-08","9024-10-10","9024-10-12","9024-12-06","9024-12-08","9024-12-12","9024-12-16","9025-04-12","9025-06-16","9025-08-18","9025-08-20","9025-10-22","9025-12-26","9033-02-02","9033-02-06","9033-04-02","9033-04-04","9033-06-04","9033-06-06","9033-08-06","9033-08-08","9033-10-08","9033-10-10","9033-12-08","9033-12-12","9044-04-04","9044-06-06","9044-08-08","9222-04-04","9222-06-04","9222-06-06","9222-08-06","9222-08-08","9222-12-08","9222-12-10","9222-12-12","2501-08-02","2502-03-02","2503-04-08","2701-03-03","2702-05-05","5000-12-04","6400-08-05","6400-10-04","6400-10-14","6404-08-12","6500-06-04","6500-08-12","6505-10-06","7005-08-16","7005-08-18","7005-08-22","7062-08-12","7205-12-26","2403-04-02","2403-05-04","2404-04-12","2404-05-08","2404-06-16","2404-08-02","2404-12-24","2405-05-06","2406-04-04","2406-04-05","2406-05-04","2406-08-05","2406-08-08","7001-04-10","7001-04-12","7001-04-14","7001-04-16","7001-05-14","7001-06-12","7001-06-14","7001-06-16","7001-06-18","7001-08-14","7001-08-16","7001-08-18","7001-08-22","7001-10-18","7001-10-20","7001-10-22","7001-12-22","7001-12-26","7001-12-27","7005-04-18","7005-06-12","7005-06-14","7005-06-16","7005-06-18","7005-06-20","7005-06-22","7005-08-14","7005-08-20","7005-10-16","7005-10-26","7045-02-18","7045-04-16","7045-12-30"]

bidsheet_file = "new/bidsheet_master_consolidate.xlsx"
wapp_file = "wapp2.xlsx"
p21_file = "P21 supplier bid supplier norm 070725v3.xlsx"
supplier_port_file = "Supplier Port per Part table 070925.csv"
frieght_file = "Freight cost mutipliers table 071025v2.csv"
port_country_map = {
    'DALIAN': 'China', 
    'NINGBO': 'China', 
    'QINGDAO2': 'China2', 
    'QINGDAO': 'China', 
    'SHANGHAI': 'China',
    'SHENZHEN': 'China', 
    'TIANJIN': 'China', 
    'XINGANG': 'China', 
    'XIAMEN': 'China',
    'AHMEDABAD': 'India', 
    'CHENNAI': 'India', 
    'DADRI': 'India', 
    'MUMBAI': 'India',
    'MUNDRA': 'India', 
    'NHAVA SHEVA': 'India',
    'SURABAYA': 'Indonesia',
    'PORT KLANG': 'Malaysia', 
    'PASIR GUDANG': 'Malaysia', 
    'TANJUNG PELAPAS': 'Malaysia',
    'BUSAN': 'South Korea',
    'KAOHSIUNG': 'Taiwan', 
    'KEELUNG': 'Taiwan', 
    'TAICHUNG': 'Taiwan', 
    'TAIPEI': 'Taiwan',
    'BANGKOK': 'Thailand',
    'LAEM CHABANG': 'Thailand',
    'HO CHI MINH CITY': 'Vietnam', 
    'VUNG TAU': 'Vietnam', 
    'HAI PHONG': 'Vietnam'
}

tariff_data = [
    ('China2', 'Aluminum', 0.25), 
    ('China2', 'brass', 0.00), 
    ('China2', 'steel', 0.50), 
    ('China2', 'Bronze', 0.55), 
    ('China2', 'Iron', 0.55), 
    ('China2', 'Lead-free bronze', 0.55), 
    ('China2', 'Zinc', 0.55),
    ('China2', 'Copper', 0.55),

    ('China', 'Aluminum', 0.50), 
    ('China', 'brass', 0.00), 
    ('China', 'steel', 0.50), 
    ('China', 'Bronze', 0.55), 
    ('China', 'Iron', 0.55), 
    ('China', 'Lead-free bronze', 0.55), 
    ('China', 'Zinc', 0.55),
    ('China', 'Copper', 0.55),

    ('India', 'Aluminum', 0.50), 
    ('India', 'brass', 0.00),
    ('India', 'steel', 0.50), 
    ('India', 'Bronze', 0.10), 
    ('India', 'Iron', 0.10), 
    ('India', 'Lead-free bronze', 0.10), 
    ('India', 'Zinc', 0.10),
    ('India', 'Copper', 0.10),

    ('Indonesia', 'Aluminum', 0.50), 
    ('Indonesia', 'brass', 0.00), 
    ('Indonesia', 'steel', 0.50), 
    ('Indonesia', 'Bronze', 0.10), 
    ('Indonesia', 'Iron', 0.10), 
    ('Indonesia', 'Lead-free bronze', 0.10), 
    ('Indonesia', 'Zinc', 0.10),
    ('Indonesia', 'Copper', 0.10),

    ('Malaysia', 'Aluminum', 0.50), 
    ('Malaysia', 'brass', 0.00), 
    ('Malaysia', 'steel', 0.50), 
    ('Malaysia', 'Bronze', 0.10),
    ('Malaysia', 'Iron', 0.10),
    ('Malaysia', 'Lead-free bronze', 0.10),
    ('Malaysia', 'Zinc', 0.10),
    ('Malaysia', 'Copper', 0.10),

    ('Taiwan', 'Aluminum', 0.50), 
    ('Taiwan', 'brass', 0.00), 
    ('Taiwan', 'steel', 0.50), 
    ('Taiwan', 'Zinc', 0.10),
    ('Taiwan', 'Bronze', 0.10),
    ('Taiwan', 'Iron', 0.10),
    ('Taiwan', 'Lead-free bronze', 0.10),
    ('Taiwan', 'Copper', 0.10),

    ('Thailand', 'Aluminum', 0.50), 
    ('Thailand', 'brass', 0.00), 
    ('Thailand', 'steel', 0.50), 
    ('Thailand', 'Zinc', 0.10),
    ('Thailand', 'Bronze', 0.10),
    ('Thailand', 'Iron', 0.10),
    ('Thailand', 'Lead-free bronze', 0.10),
    ('Thailand', 'Copper', 0.10),

    ('Vietnam', 'Aluminum', 0.50),
    ('Vietnam', 'brass', 0.00), 
    ('Vietnam', 'steel', 0.50), 
    ('Vietnam', 'Zinc', 0.10),
    ('Vietnam', 'Iron', 0.10),
    ('Vietnam', 'Lead-free bronze', 0.10),
    ('Vietnam', 'Copper', 0.10),
]
tariff_data_v2 = [
    ('China', 'Aluminum', 0.50), 
    ('China', 'brass', 0.50), 
    ('China', 'steel', 0.50), 
    ('China', 'Bronze', 0.55), 
    ('China', 'Iron', 0.55), 
    ('China', 'Lead-free bronze', 0.55), 
    ('China', 'Zinc', 0.55),
    ('China', 'Copper', 0.55),

    ('India', 'Aluminum', 0.50), 
    ('India', 'brass', 0.50),
    ('India', 'steel', 0.50), 
    ('India', 'Bronze', 0.26), 
    ('India', 'Iron', 0.26), 
    ('India', 'Lead-free bronze', 0.26), 
    ('India', 'Zinc', 0.26),
    ('India', 'Copper', 0.26),

    ('Indonesia', 'Aluminum', 0.50), 
    ('Indonesia', 'brass', 0.50), 
    ('Indonesia', 'steel', 0.50), 
    ('Indonesia', 'Bronze', 0.32), 
    ('Indonesia', 'Iron', 0.32), 
    ('Indonesia', 'Lead-free bronze', 0.32), 
    ('Indonesia', 'Zinc', 0.32),
    ('Indonesia', 'Copper', 0.32),

    ('Malaysia', 'Aluminum', 0.50), 
    ('Malaysia', 'brass', 0.50), 
    ('Malaysia', 'steel', 0.50), 
    ('Malaysia', 'Bronze', 0.25),
    ('Malaysia', 'Iron', 0.25),
    ('Malaysia', 'Lead-free bronze', 0.25),
    ('Malaysia', 'Zinc', 0.25),
    ('Malaysia', 'Copper', 0.25),

    ('Taiwan', 'Aluminum', 0.50), 
    ('Taiwan', 'brass', 0.50), 
    ('Taiwan', 'steel', 0.50), 
    ('Taiwan', 'Zinc', 0.32),
    ('Taiwan', 'Bronze', 0.32),
    ('Taiwan', 'Iron', 0.32),
    ('Taiwan', 'Lead-free bronze', 0.32),
    ('Taiwan', 'Copper', 0.32),

    ('Thailand', 'Aluminum', 0.50), 
    ('Thailand', 'brass', 0.50), 
    ('Thailand', 'steel', 0.50), 
    ('Thailand', 'Zinc', 0.36),
    ('Thailand', 'Bronze', 0.36),
    ('Thailand', 'Iron', 0.36),
    ('Thailand', 'Lead-free bronze', 0.36),
    ('Thailand', 'Copper', 0.36),

    ('Vietnam', 'Aluminum', 0.50),
    ('Vietnam', 'brass', 0.50), 
    ('Vietnam', 'steel', 0.50), 
    ('Vietnam', 'Zinc', 0.20),
    ('Vietnam', 'Iron', 0.20),
    ('Vietnam', 'Lead-free bronze', 0.20),
    ('Vietnam', 'Copper', 0.20),
]

tariff_df = pd.DataFrame(tariff_data, columns=['Country', 'Metal Type', 'Tariff Multiplier'])
tariff_df_2 = pd.DataFrame(tariff_data_v2, columns=['Country', 'Metal Type', 'Tariff Multiplier'])

part_numbers = ["39877","CGC-050-A1","CDC-050-A1","CDC-075-A1","CDC-100-A1","CDC-100-A1S","CDC-125-A1","CDC-150-A1","CDC-150-A1S","CDC-200-A1","CDC-200-A1S","CDC-250-A1","CDC-300-A1","CDC-300-A1S","CDC-400-A1","CDC-400-A1S","CDC-500-A1","CDC-600-A1","CDC-600-A1S","CDC-800-A1","CDCHK5-400-A1","CDCL-200-A","CDCL-400-A","CDCL-500-A","CDCL-600-A","CDCSL-200-A1","CDCSL-300-A1","CDCSL-400-A1","CDCSL-600-A1","CDCWC-250-A1","CDCWC-400-A1","CDP-050-A1","CDP-075-A1","CDP-100-A1","CDP-125-A1","CDP-200-A1","CDP-300-A1","CDP-400-A1","CDP-500-A1","CDP-600-A1","CDP-800-A1","CDPWC-300-A1","CDPWC-400-A1","CGA-050-A1","CGA-075-A1","CGA-1000-A","CGA-100-A1","CGA-125-A1","CGA-150-A1","CGA-200-A1","CGA-2015-A","CGA-250-A1","CGA-300-A1","CGA-3020-A","CGA-3040-A","CGA-400-A1","CGA-4030-A","CGA-4060-A","CGA-500-A1","CGA-600-A1","CGA-800-A1","CGABSP-300-A1","CGABSP-400-A1","CGAF-300-A","CGAF-600-A","CGAF-800-AA","CGANPS-200-A","CGASV-300-A","CGASV-400-A","CGAW-300-A","CGAW-400-A","CGAW-600-A","CGB-075-A1","CGB-100-A1","CGB-100-A1S","CGB-125-A1","CGB-150-A1","CGB-150-A1S","CGB-200-A1","CGB-200-A1S","CGB-2015-A","CGB-250-A1","CGB-300-A1","CGB-3020-A","CGB-3040-A","CGB-400-A1","CGB-500-A1","CGB-600-A1","CGB-800-A1","CGBSL-300-A1","CGBSL-400-A1","CGBWW-200-A1","CGBWW-300-A1","CGC-075-A1","CGC-075-A1S","CGC-1000-A","CGC-100-A1","CGC-100-A1S","CGC-125-A1","CGC-125-A1S","CGC-150-A1","CGC-150-A1S","CGC-150CR-A1","CGC-1590-A","CGC-200-A1","CGC-200-A1S","CGC-200CR-A1","CGC-2015-A","CGC-2090-A","CGC-250-A1","CGC-300-A1","CGC-300-A1S","CGC-300CR-A1","CGC-3025-A","CGC-400-A1","CGC-400-A1S","CGC-400CR-A1","CGC-4030-A","CGC-500-A1","CGC-500-A1S","CGC-600-A1","CGC-600-A1S","CGC-800-A1","CGC-804-A1","CGCHK5-150-A1","CGCP-300-A","CGCSL-200CR-A1","CGCSL-300CR-A1","CGCSL-400CR-A1","CGCSL-600CR-A1","CGCT-400-A","CGD-050-A1","CGD-075-A1","CGD-1000-A","CGD-100-A1","CGD-100-A1S","CGD-125-A1","CGD-150-A1","CGD-1510-A","CGD-200-A1","CGD-200-A1S","CGD-2015-A","CGD-250-A1","CGD-250-A1S","CGD-300-A1","CGD-3020-A","CGD-400-A1","CGD-400-A1S","CGD-500-A1","CGD-600-A1","CGD-600-A1S","CGD-800-A1","CGDBSP-150-A1","CGDBSP-200-A1","CGDBSP-300-A1","CGDF-600-A","CGDHK5-200-A1","CGDSL-200-A1","CGDSL-300-A1","CGDSL-400-A1","CGDSL-600-A1","CGDSV-300-A","CGDSV-400-A","CGDW-400-A","CGDW-500-A","CGDW-600-A","CGE-075-A1","CGE-100-A1","CGE-150-A1","CGE-150CR-A1","CGE-200-A1","CGE-200CR-A1","CGE-2015-A","CGE-2030-A","CGE-250-A1","CGE-300-A1","CGE-300CR-A1","CGE-3025-A","CGE-400-A1","CGE-400CR-A1","CGE-500-A1","CGE-600-A1","CGE-600CR-A1","CGE-800-A1","CGF-050-A1","CGF-075-A1","CGF-100-A1","CGF-125-A1","CGF-150-A1","CGF-1520-A","CGF-200-A1","CGF-2015-A","CGF-250-A1","CGF-300-A1","CGF-3020-A","CGF-3040-A","CGF-400-A1","CGF-4030-A","CGF-500-A1","CGF-600-A1","CGF-800-A1","CGFBSP-400-A1","CGFSW-400-A","CGFSW-500-A","CGFSW-600-A","CGFWW-200-A1","CNT-200-A","CNT-400-A","CSL-40415-A","DA-1510-A","DA-2015-A","DA-2015-AA","DA-2030-A","DA-2030-AA","DA-2040-A","DA-2040-AA","DA-3015-A","DA-3020-A","DA-3020-AA","DA-3040-A","DA-3040-AA","DA-4020-A","DA-4020-AA","DA-4030-A","DA-4030-AA","DA-4060-A","DA-4060-AA","DA-5040-A","DA-6040-A","DA-6040-AA","DA-6050-A","DA-8060-AA","DASG-3030-A","DASL-3020-A1","DASL-3040-A1","DASL-4030-A1","DASL-6040-A1","DD-2020-A","DD-2020-AA","DD-2030-A","DD-2030-AA","DD-3030-A","DD-3030-AA","DD-3040-A","DD-3040-AA","DD-4040-A","DD-4040-AA","DDSL-2020-A1","DDSL-3030-A1","DDSL-4040-A1","FCF-150-A","FFF-075-A","FFF-100-A","FFSV-075-A","FFSV-100-A","PF-200-AD","PL-150-AD","PL-200-AD","PL-300-AD","PM-150-A","PM-200-A","PM-300-A","PM-400-A","SA-100-A","SA-1520-A","SA-200-A","SA-200-AA","SA-2030-AA","SA-300-A","SA-300-AA","SA-3040-AA","SA-400-AA","SA-4050-A","SA-4060-A","SFL-200-A","SFL-300-A","SFL-400-A","SFL-500-A","SFL-600-A","SFL-800-A","SHANK-800-S","SNH-075-A","SNH-075-ANPT","SNH-100-A","SNH-125-A","SQC-100-A","SQC-151-A","SRC-2003-A","SRC-3003-A","STA-125-ANPS","STA-151-A","WNH40-300-A","39531","39537","39870","39871","39874","39875","39877","39878","03C03395","320054","320120","320121","940825","940826","940827","CGC-250-A1","DASL-2030-A1","03C03392","03C03393","03C03394","03C03396","CGD-075-A1","28560SC","28579SC","CGBSL-200-A1","46523A","AF4","AFR1","CDC-075-A1","CDC-100-A1","CDC-125-A1","CDC-150-A1","CDC-200-A1","CDC-300-A1","CDC-400-A1","CDC-800-A1","CDP-075-A1","CDP-100-A1","CDP-150-A1","CDP-200-A1","CDP-300-A1","CDP-400-A1","CDP-600-A1","CGA-050-A1","CGA-075-A1","CGA-100-A1","CGA-125-A1","CGA-150-A1","CGA-200-A1","CGA-250-A1","CGA-300-A1","CGA-400-A1","CGA-500-A1","CGA-600-A1","CGA-6040-A","CGA-800-A1","CGB-075-A1","CGB-100-A1","CGB-150-A1","CGB-200-A1","CGB-250-A1","CGB-300-A1","CGB-600-A1","CGC-075-A1","CGC-100-A1","CGC-125-A1","CGC-150-A1","CGC-200-A1","CGC-2015-A","CGC-300-A1","CGC-3025-A","CGC-400-A1","CGC-400CR-A1","CGC-4030-A","CGC-500-A1","CGC-600CR-A1","CGC-804-A1","CGCSL-200CR-A1","CGD-100-A1","CGD-125-A1","CGD-150-A1","CGD-200-A1","CGD-2015-A","CGD-250-A1","CGD-300-A1","CGD-3020-A","CGD-400-A1","CGD-4030-A","CGD-500-A1","CGD-600-A1","CGD-800-A1","CGDSL-200-A1","CGE-050-A1","CGE-075-A1","CGE-100-A1","CGE-125-A1","CGE-150-A1","CGE-200-A1","CGE-300-A1","CGE-300CR-A1","CGE-400-A1","CGF-075-A1","CGF-100-A1","CGF-125-A1","CGF-150-A1","CGF-200-A1","CGF-2015-A","CGF-250-A1","CGF-300-A1","CGF-3040-A","CGF-400-A1","CGF-500-A1","CGF-600-A1","CGF-800-A1","CNT-300-A","CNT-400-A","CSL-20210-A","CSL-30310-A","CNT-125-A","CGD-050-A1","PF-150-AD","CDC-250-A1","SA-400-A","CSL-40414-A","DA-1520-A","DA-2015-A","DA-2030-A","DA-2040-A","DA-3015-A","DA-3020-A","DA-3020-AA","DA-3040-A","DA-4020-A","DA-4030-A","DA-4030-AA","DA-4060-A","DA-6040-A","DD-1515-A","DD-2030-A","DD-3040-A","DD-3040-AA","E-CDC-050-A1","E-CDC-075-A1","E-CDC-100-A1","E-CDC-125-A1","E-CDC-150-A1","E-CDC-200-A1","E-CDC-250-A1","E-CDC-300-A1","E-CDC-400-A1","E-CDC-500-A1","E-CDC-600-A1","E-CDC-800-A1","E-CDP-050-A1","E-CDP-075-A1","E-CDP-100-A1","E-CDP-125-A1","E-CDP-150-A1","E-CDP-200-A1","E-CDP-250-A1","E-CDP-300-A1","E-CDP-400-A1","E-CDP-500-A1","E-CDP-600-A1","E-CDP-800-A1","E-CGA-050-A1","E-CGA-075-A1","E-CGA-100-A1","E-CGA-125-A1","E-CGA-150-A1","E-CGA-200-A1","E-CGA-250-A1","E-CGA-300-A1","E-CGA-3040-A","E-CGA-400-A1","E-CGA-4030-A","E-CGA-500-A1","E-CGA-600-A1","E-CGA-800-A1","E-CGB-050-A1","E-CGB-075-A1","E-CGB-100-A1","E-CGB-125-A1","E-CGB-150-A1","E-CGB-200-A1","E-CGB-250-A1","E-CGB-300-A1","E-CGB-400-A1","E-CGB-500-A1","E-CGB-600-A1","E-CGB-800-A1","E-CGC-050-A1","E-CGC-075-A1","E-CGC-100-A1","E-CGC-125-A1","E-CGC-150-A1","E-CGC-200-A1","E-CGC-250-A1","E-CGC-300-A1","E-CGC-400-A1","E-CGC-500-A1","E-CGC-600-A1","E-CGC-800-A1","E-CGD-050-A1","E-CGD-075-A1","E-CGD-100-A1","E-CGD-125-A1","E-CGD-150-A1","E-CGD-200-A1","E-CGD-250-A1","E-CGD-300-A1","E-CGD-3020-A","E-CGD-400-A1","E-CGD-4030-A","E-CGD-500-A1","E-CGD-600-A1","E-CGD-800-A1","E-CGE-050-A1","E-CGE-075-A1","E-CGE-100-A1","E-CGE-125-A1","E-CGE-150-A1","E-CGE-200-A1","E-CGE-250-A1","E-CGE-300-A1","E-CGE-400-A1","E-CGE-500-A1","E-CGE-600-A1","E-CGE-800-A1","E-CGF-050-A1","E-CGF-075-A1","E-CGF-100-A1","E-CGF-125-A1","E-CGF-150-A1","E-CGF-200-A1","E-CGF-250-A1","E-CGF-300-A1","E-CGF-400-A1","E-CGF-500-A1","E-CGF-600-A1","E-CGF-800-A1","E-DA-2015-A","E-DA-2030-A","E-DA-3015-A","E-DA-3020-A","E-DA-3040-A","E-DA-4030-A","E-DA-6040-A","E-DD-2020-A","E-DD-2030-A","E-DD-3030-A","E-DD-3040-A","E-DD-4040-A","E-SA-200-A","E-SA-2030-A","E-SA-300-A","E-SA-3040-A","E-SA-400-A","K601","PF-200-AD","PL-200-AD","PM-200-A","SA-200-A","SA-200-AA","SA-2030-A","SA-300-A","SA-3040-A","SA-400-A","CDP-150-A1","CDP-125-A1","CGBSW-400-A","CDP-500-A1","SA-250-A","PF-300-AD","CGB-300-A1S","CGB-400-A1S","DA-1520-A","CGE-250CR-A1","CGC-3020-A","CDCWC-300-A1","SA-4060-AA","CGC-600CR-A1","SFL-1200-A","CGDF-400-A","44474","44276","44101","44103","44105","44106","44108","44163","44164","44165","44166","44183","44184","44186","44205","44252","44253","44255","44257","44258","44282","44394","44395","44411","44414","44603","44438","44441","44442","44446","44450","44454","44462","44473","44511","44512","44514","44516","44521","44527","44528","44529","44531","44605","44635","44656","44657","44675","44671","738108-32","738119-1204","44203","44505","44410","45472LFCP","44102","44103","44105","44106","44108","44109","44110","44127","44131","44160","44162","44163","44164","44165","44166","44167","44168","44169","44170","44183","44187","44188","44190","44200","44205","44206","44207","44208","44252","44253","44255","44256","44257","44258","44259","44280","44282","44291","44298","44301","44315","44392","44394","44396","44398","44411","44413","44414","44415","44416","44417","44418","44419","44431","44432","44435","44437","44438","44442","44445","44446","44450","44451","44452","44453","44456","44462","44471","44473","44475","44476","44477","44478","44480","44500","44501","44503","44504","44507","44508","44509","44510","44511","44513","44514","44516","44517","44518","44519","44522","44523","44524","44528","44529","44530","44531","44535","44537","44538","44543","44602","44604","44605","44606","44607","44608","44609","44634","44635","44636","44637","44638","44651","44653","44654","44655","44656","44657","44658","44659","44660","44672","44673","44676","44678","44706","44707","44724","44725","44726","44727","44728","44250","45412LFCP","45181LFCP","44433","44434","44508","300020","300021","300022","300023","300024","300026","44434","947141","947144","973948","44185","973966","973982","45435LFCP","738110-2412","45254LFCP","45513LFCP","44437","44107","45438LFCP","44276","44472LF","44125","44279","44840LF","44843LF","45184LFCP","44449","45473LFCP","45474LFCP","45603LFCP","44440","9621212LF","9621616LF","9631618LF","9641212LF","9641216LF","9651212LF","973960LF","973961LF","973962LF","973965LF","973976LF","44283","45475LFCP","44420","9621216LF","44677","44444","45205LFCP","44254","738110-3212","44185","44395","45394LFCP","738119-2420","44539","44140","44610","44420","44463","44512","10085","10097","35220","35221","35222","35223","AF1","AF2","AF3","AF5","AL2","AL3","AR1","AR7","64103","64104","64105","64107","64108","64109","64110","64131","64134","64162","64163","64164","64165","64166","64167","64168","64181","64183","64184","64186","64187","64188","64204","64205","64208","64253","64254","64255","64256","64257","64258","64260","64280","64291","64298","64316","64410","64412","64414","64418","64419","64420","64421","64434","64441","64442","64449","64450","64451","64452","64453","64454","64473","64474","64475","64476","64477","64478","64480","64481","64505","64509","64513","64517","64521","64522","64523","64524","64526","64528","64529","64530","64531","64534","64535","64537","64538","64553","64601","64603","64604","64605","64606","64607","64608","64609","64610","64654","64658","64660","64694","64696","65100","65101","65102","65103","65104","65105","65106","65107","65108","65109","65110","65112","65114","65115","65126","65127","65130","65131","65134","65136","65137","65138","65142","65146","65151","65160","65161","65162","65163","65164","65165","65166","65167","65168","65169","65170","65171","65181","65182","65183","65184","65185","65186","65187","65188","65189","65190","65191","65192","65194","65201","65202","65203","65204","65205","65206","65207","65208","65209","65210","65211","65250","65251","65252","65253","65254","65255","65256","65257","65258","65259","65260","65261","65262","65264","65276","65286","65287","65288","65291","65292","65294","65295","65296","65297","65298","65301","65303","65306","65309","65311","65316","65318","65322","65327","65328","65330","65334","65336","65363","65391","65393","65395","65396","65397","65398","65401","65410","65411","65412","65413","65414","65415","65416","65417","65418","65419","65420","65421","65422","65432","65433","65434","65435","65436","65437","65438","65440","65441","65442","65444","65445","65446","65447","65448","65449","65450","65451","65452","65453","65454","65456","65457","65458","65459","65460","65461","65462","65463","65464","65465","65472","65473","65474","65475","65476","65477","65478","65479","65480","65481","65482","65486","65491","65500","65502","65503","65504","65505","65507","65508","65509","65510","65511","65512","65513","65514","65516","65517","65518","65519","65520","65521","65522","65523","65524","65525","65526","65527","65528","65529","65530","65531","65532","65533","65534","65535","65537","65538","65539","65540","65541","65542","65543","65544","65545","65549","65550","65553","65601","65602","65603","65604","65605","65606","65607","65608","65609","65610","65650","65651","65652","65653","65654","65655","65656","65657","65658","65659","65660","65661","65662","65694","65695","65697","65721","65965","65966","65967","65968","65971","69103","69104","69105","69107","69108","69254","69258","69413","69416","69432","69603","69605","69607","69608","108028","64654","57228V","59160SMLSXX","BC2-050A-DP","BC2-075A-DP","BC2-075C-DP","BC4-100A-DP","BC4-100-DP","BC4-200A-DP","BC4-200-DP","BC4-300-DP","CGA-150-DP","CGA-400-DP","CGASW-400-D","CGASW-600-D","CGD-300-DP","CGD-400-DP","CGE-400-DP","DBC-1125-DP","DBC-1360-DP","DBC-525-DP","DBC-550-DP","DBC-60-DP","DBC-675-DP","DBC-769-DP","DBC-875-DP","DBC-988-DP","GDS-200-DP","GJF-075-DP","GJF-100-DP","GJF-150-DP","GJF-200-DP","GJF-300-DP","GJF-400-DP","GJM-125-DP","GJM-150-DP","GJM-200CR-DP","GJM-200-DP","GJN-200-DP","GJN-300-DP","GJN-400-DP","GJS-075-DP","GJS-150-DP","GJS-200CR-DP","GJS-200-DP","GJS-300CR-DP","GJS-300-DP","GJS-400-DP","GMS-075-DP","GMS-200-DP","GMS-300-DP","HE-200-SP","ME-025-SP","ME-038-SP","ME-050-SP","ME-075-SP","ME-100-SP","PF-250-DP","PM-200-DP","R88DB-12-DP","R88DB-16-DP","R88DB-20-DP","R88DB-24-DP","R88DB-32-DP","SFL-400-DP","SRG-NY2","21178","39506","39509","39510","64705","64100","64101","64102","64103","64104","64105","64106","64107","64108","64109","64110","64112","64114","64124","64126","64130","64131","64134","64138","64139","64142","64145","64146","64160","64161","64162","64163","64164","64165","64166","64167","64168","64169","64170","64171","64181","64183","64184","64185","64186","64187","64188","64189","64190","64191","64201","64202","64203","64204","64205","64206","64207","64208","64209","64210","64250","64251","64252","64253","64254","64255","64256","64257","64258","64259","64260","64261","64269","64279","64280","64283","64286","64288","64289","64291","64296","64297","64298","64303","64306","64316","64318","64328","64330","64334","64351","64363","64383","64385","64391","64392","64393","64394","64395","64397","64398","64410","64411","64412","64413","64414","64415","64416","64417","64418","64419","64420","64421","64422","64430","64431","64432","64434","64435","64437","64438","64439","64440","64441","64442","64445","64446","64447","64448","64449","64450","64451","64452","64453","64454","64456","64457","64458","64459","64461","64462","64464","64465","64466","64473","64474","64475","64476","64477","64478","64479","64480","64481","64486","64491","64502","64503","64504","64505","64506","64507","64508","64509","64510","64511","64512","64513","64514","64516","64517","64518","64519","64521","64522","64523","64524","64525","64526","64527","64528","64529","64530","64531","64532","64533","64534","64535","64537","64538","64539","64541","64542","64543","64545","64549","64550","64553","64601","64602","64603","64604","64605","64606","64607","64608","64609","64610","64611","64653","64655","64656","64657","64658","64660","64661","64692","64693","64694","64695","64696","64697","64706","64707","64708","64964","64965","64966","64968","64969","65100","65101","65102","65103","65104","65105","65106","65107","65108","65109","65110","65112","65113","65114","65115","65122","65126","65127","65131","65133","65134","65136","65137","65138","65139","65140","65142","65150","65161","65162","65163","65164","65165","65166","65167","65168","65169","65170","65180","65181","65182","65183","65184","65185","65186","65187","65188","65189","65190","65191","65192","65200","65201","65202","65203","65204","65205","65206","65207","65208","65209","65210","65211","65250","65251","65252","65253","65254","65255","65256","65257","65258","65259","65260","65261","65262","65264","65279","65280","65283","65286","65287","65288","65289","65291","65293","65295","65296","65298","65301","65303","65305","65306","65309","65312","65313","65316","65317","65318","65319","65321","65322","65323","65325","65327","65328","65330","65334","65336","65339","65347","65348","65349","65350","65351","65353","65363","65364","65372","65375","65386","65391","65392","65393","65394","65395","65396","65397","65398","65399","65400","65401","65410","65411","65412","65413","65414","65415","65416","65417","65418","65419","65420","65421","65422","65430","65431","65432","65433","65434","65435","65436","65437","65438","65439","65440","65441","65442","65443","65444","65445","65446","65447","65448","65449","65450","65451","65452","65453","65454","65456","65457","65458","65459","65460","65461","65462","65463","65464","65465","65466","65470","65471","65472","65473","65474","65475","65476","65477","65478","65479","65480","65481","65482","65486","65491","65500","65501","65502","65503","65504","65505","65506","65507","65508","65509","65510","65511","65512","65513","65514","65515","65516","65517","65518","65519","65520","65521","65522","65523","65524","65525","65526","65527","65528","65529","65530","65531","65532","65533","65534","65535","65537","65538","65539","65540","65541","65542","65543","65544","65545","65547","65549","65550","65553","65559","65575","65600","65601","65602","65603","65604","65605","65606","65607","65608","65609","65610","65611","65650","65652","65653","65654","65655","65656","65657","65658","65659","65660","65661","65662","65664","65692","65693","65694","65695","65696","65697","65964","65965","65966","65967","65968","65970","65971","66005","66006","66007","66008","66009","66011","66012","66013","69101","69103","69105","69107","69108","69163","69165","69166","69167","69168","69181","69182","69183","69184","69186","69187","69188","69252","69253","69254","69255","69256","69257","69258","69413","69414","69415","69416","69418","69432","69472","69473","69475","69477","69478","69601","69602","69603","69604","69605","69606","69608","108025","108027","108028","108029","108032","108035","108038","108067","108069","108071","108075","108088","108095","108096","108097","108098","64122","64433","64182","962004","962006","962010","963003","963004","963006","963008","963012","64124","65123","64439","66010","108067BF","64126","64127","64459","64659","108065","64137","64463","57201TB","57221TB","57228V","9600FJ10","9600FJ12","9600FJ3","9600FJ4","9600FJ6","9600FJ8","9600FL10","9600FL12","9600FL14","9600FL3","9600FL4","9600FL6","9600FL8","9600IPA412","9600MJ10","9600MJ12","9600MJ16","9600MJ3","9600MJ6","9600MJ8","9610F10","9610F12","9610F3","962010LW","962012LW","96203LW","96204LW","96204UL","96206LW","96206UL","96208LW","9640FL10","9640FL2","9640FL8","9640TH2","9640TH3","9640TH4","9650EX1248","9650EX4648","9650G10","9650G10SE","9650G12","9650G3","9650G3SE","9650G4","9650G4SE","9650G8","9650G812","9650L2.5","9650L23","9650L3SE","9650L4","9650L46","9650L4SE","9650L5","9650L6","9650L6SE","9650L8","9650L8SE","9660G10","9660G10SE","9660G12","9660G12SE","9660G2.5SE","9660G4","9660G4SE","9660G6SE","9660G8SE","9660L2","9660L2.5","9660L3","9660L3SE","9660L4SE","9660L6","9660L6SE","9660L8","9660L8SE","BC2-075A-DP","BC4-100B-DP","BC4-200A-DP","CGA-300-DP","CGA-400-DP","CGB-200-DP","CGC-300-DP","CGE-300-DP","65383","BFVGE-200","64282","64393","69185","BFVGE-300","BFVGH-600-NBR","108094","57600GBS","65133","65135","65384","BFVGE-600","BFVGH-800-NBR","GJN-075-DP","DBC-1275-DP","DBC-400-DP","DBC-525-DP","DBC-550-DP","DBC-675-DP","DBC-76-DP","DBC-875-DP","DBC-94-DP","DBC-988-DP","GJM-200-DP","GJN-300-DP","ME-025-SP","ME-075-SP","ME-100-SP","R88DB-16-DP","R88DB-20-DP","R88DB-24-DP","R88DB-32-DP","64261","64463","GDS-300-DP","108092","64971","BFVGH-200-NBR","GJN-150-DP","9660G8","65317","65386","BFVGH-200-VI","64141","BC4-125-DP","65319","64336","65151","108037","9600MJ4","CGASW-300-D","65145","CDP-400-DP","CGD-300-DP","65343","45652LFCP","45502LFCP","45504LFCP","44442LF","44454LF","44513LF","44127LF","45201LFCP","45411LFCP","45161LFCP","45202LFCP","738102-04","45102LFCP","45654LFCP","44506LF","45182LFCP","45251LFCP","45509LFCP","45124LFCP","45162LFCP","45391LFCP","45601LFCP","45655LFCP","45252LFCP","44109LF","44110LF","44122LF","44123LF","44128LF","44138LF","44167LF","44168LF","44169LF","44180LF","44186LF","44187LF","44188LF","44201LF","44202LF","44205LF","44206LF","44257LF","44258LF","44276LF","44279LF","44280LF","44282LF","44287LF","44288LF","44291LF","44296LF","44297LF","44318LF","44328LF","44330LF","44351LF","44390LF","44391LF","44392LF","44393LF","44395LF","44396LF","44398LF","44419LF","44420LF","44421LF","44431LF","44432LF","44436LF","44444LF","44445LF","44447LF","44448LF","44449LF","44450LF","44451LF","44452LF","44453LF","44454LF","44456LF","44459LF","44461LF","44462LF","44470LF","44476LF","44477LF","44478LF","44481LF","44510LF","44511LF","44512LF","44516LF","44517LF","44519LF","44521LF","44522LF","44524LF","44527LF","44528LF","44529LF","44530LF","44533LF","44534LF","44535LF","44538LF","44539LF","44541LF","44543LF","44600LF","44610LF","44634LF","44651LF","44652LF","44656LF","44657LF","44670LF","44673LF","44677LF","44678LF","44790LF","44794LF","44795LF","44796LF","44797LF","44798LF","45103LFCP","45104LFCP","45127LFCP","45163LFCP","45164LFCP","45183LFCP","45203LFCP","45253LFCP","45392LFCP","45413LFCP","45414LFCP","45434LFCP","45604LFCP","45505LFCP","45653LFCP","45204LFCP","45255LFCP","738108-20","44792LF","45165LFCP","45393LFCP","45605LFCP","44639LF","45185LFCP","44207LF","45395LFCP","44327LF","44454","44100LF","44101LF","44102LF","44103LF","44104LF","44105LF","44106LF","44107LF","44108LF","44124LF","44126LF","44131LF","44160LF","44161LF","44162LF","44163LF","44164LF","44165LF","44166LF","44181LF","44182LF","44183LF","44184LF","44185LF","44203LF","44204LF","44250LF","44251LF","44252LF","44253LF","44254LF","44255LF","44256LF","44268LF","44286LF","44298LF","44394LF","44410LF","44411LF","44412LF","44413LF","44414LF","44415LF","44416LF","44417LF","44418LF","44430LF","44434LF","44435LF","44437LF","44438LF","44441LF","44442LF","44446LF","44471LF","44473LF","44474LF","44475LF","44500LF","44502LF","44503LF","44504LF","44505LF","44507LF","44508LF","44509LF","44513LF","44518LF","44523LF","44531LF","44601LF","44602LF","44603LF","44604LF","44605LF","44606LF","44607LF","44608LF","44653LF","44654LF","44655LF","44658LF","44671LF","44672LF","44674LF","44675LF","44676LF","30600","30609","30626","30629","46270","320050","320053","AR10","AR3","AR4","AR5","AR6","AR8","AR9" ]
metals = ["Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Aluminum","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Bronze","Copper ","Copper ","Copper ","Copper ","Copper ","Copper ","Copper ","Copper ","Copper ","Copper ","Copper ","Copper ","Copper ","Copper ","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Iron","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Lead-free bronze","Zinc ","Zinc ","Zinc ","Zinc ","Zinc ","Zinc ","Zinc ","Zinc ","Zinc ","Zinc ","Zinc ","Zinc ","Zinc ","Zinc ",]

part_metal = list(zip(part_numbers, metals))

output_file = "new/Bidsheet Master Consolidate Landed3.xlsx"

# Update 'type' column only for part numbers in part_metal
part_to_metal = dict(part_metal)

bidsheet_df = pd.read_excel(bidsheet_file)
wapp_df = pd.read_excel(wapp_file)
p21_df = pd.read_excel(p21_file, skiprows=2)
supplier_port_df = pd.read_csv(supplier_port_file)
frieght_file_df = pd.read_csv(frieght_file)

supplier_port_long = supplier_port_df.melt(
    id_vars=['ROW ID #', 'Division', 'Part #'],
    var_name='Supplier',
    value_name='Port'
)
supplier_port_long['Country'] = supplier_port_long['Port'].map(port_country_map)

freight_long = frieght_file_df.melt(
    id_vars=['Reference'],
    var_name='Division',
    value_name='Freight Multiplier'
)

freight_lookup_df = (
    supplier_port_long
    .merge(freight_long, left_on=['Port', 'Division'], right_on=['Reference', 'Division'], how='left')
    .drop(columns=['Reference'])
)

bidsheet_df['type'] = bidsheet_df.apply(
    lambda row: part_to_metal[row['Part #']] if row['Part #'] in part_to_metal else row['type'],
    axis=1
)

# Add Duty Multiplier column based on type
def get_duty_multiplier(metal_type, supplier, part = []):
    
    if supplier in ['Luxecasting']:
        return 0
    else:
        if isinstance(metal_type, str):
            t = metal_type.strip().lower()
            if t == 'steel':
                return 0.05
            elif t == 'brass':
                return 0.03
            
    return 0

# bidsheet_df['Duty Multiplier'] = bidsheet_df['type'].apply(get_duty_multiplier)

wapp_df['Norm Item ID'] = wapp_df['Norm Item ID'].astype(str).str.strip().str.upper()

def date_to_excel_serial(date_str):
    base_date = datetime(1899, 12, 30)  # Excel's day 0
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    delta = date_obj - base_date
    return delta.days

band_columns = {
    (1, 50): "1-50",
    (51, 200): "51-200",
    (201, 500): "201-500",
    (501, 1000): "501-1000",
    (1001, 2000): "1001-2000",
    (2001, 5000): "2001-5000",
    (5001, 10000): "5001-10000",
    (10001, 25000): "10001-25000",
    (25001, 100000): "25001-100000",
    (100001, 250000): "100001-250000",
    (250001, float('inf')): "250001+"
}

def get_band_column(aoq):
    for (low, high), col in band_columns.items():
        if low <= aoq <= high:
            return col
    return None

# --- Refined logic for fetching Volume-banded WAPP and Most common supplier ---
volume_wapp, mcs_list = [], []
norm_part_ids = bidsheet_df['Part #'].astype(str).str.strip().str.upper()
wapp_df['Norm Item ID'] = wapp_df['Norm Item ID'].astype(str).str.strip().str.upper()
bidsheet_df['Average Order Quantity (per UOM)'] = pd.to_numeric(bidsheet_df['Average Order Quantity (per UOM)'], errors='coerce')

# Find the actual column name for Most Common Supplier in wapp_df (case-insensitive, stripped)
def find_mcs_column(wapp_df):
    for col in wapp_df.columns:
        if str(col).strip().lower() == 'most common supplier':
            return col
    # Try partial match if exact not found
    for col in wapp_df.columns:
        if 'most common supplier' in str(col).strip().lower():
            return col
    return None

mcs_col_name = find_mcs_column(wapp_df)

for i, row in tqdm(bidsheet_df.iterrows(), total=len(bidsheet_df), desc='Volume WAPP'):
    part_id = norm_part_ids[i]
    aoq = row['Average Order Quantity (per UOM)']

    # Handle weird part numbers
    if part_id in [x.upper() for x in wierd_list]:
        try:
            excel_serial = str(date_to_excel_serial(row['Part #']))
        except Exception:
            excel_serial = None

        wapp_rows = wapp_df[wapp_df['Norm Item ID'] == excel_serial]
        if wapp_rows.empty:
            wapp_rows = wapp_df[wapp_df['Norm Item ID'] == part_id]
    else:
        wapp_rows = wapp_df[wapp_df['Norm Item ID'] == part_id]

    if wapp_rows.empty or pd.isna(aoq):
        volume_wapp.append('-')
        mcs_list.append('')
        continue

    wapp_row = wapp_rows.iloc[0]
    band_col = get_band_column(aoq)
    val = wapp_row.get(band_col, None)
    if val is None or (isinstance(val, float) and np.isnan(val)):
        val = wapp_row.get('Raw WAPP', '-')

    volume_wapp.append(round(val, 4) if isinstance(val, (int, float, np.floating)) and not pd.isna(val) else '-')

    # Fetch Most Common Supplier robustly
    mcs_val = wapp_row.get(mcs_col_name, '') if mcs_col_name else ''
    mcs_list.append(mcs_val if pd.notna(mcs_val) else '')

min_bid_idx = bidsheet_df.columns.get_loc("Final Min Bid")
bidsheet_df.insert(min_bid_idx, "Volume-banded WAPP", volume_wapp)
bidsheet_df.insert(min_bid_idx+1, "Most common supplier", mcs_list)

p21_df['p21_supplier_lower'] = p21_df['P21 supplier'].astype(str).str.lower().str.strip()
bidsheet_df['most_common_supplier_lower'] = bidsheet_df['Most common supplier'].astype(str).str.lower().str.strip()
mapping_dict = dict(zip(p21_df['p21_supplier_lower'], p21_df['Normalized to match bid supplier ']))
def get_normalized_supplier(mcs):
    if mcs in mapping_dict:
        return mapping_dict[mcs].strip() if isinstance(mapping_dict[mcs], str) else mapping_dict[mcs]
    else:
        return "-"
    
# Apply the mapping
normalized_incumbent_supplier = bidsheet_df['most_common_supplier_lower'].apply(get_normalized_supplier)

# Insert the new column next to "Most common supplier"
mcs_col_idx = bidsheet_df.columns.get_loc("Most common supplier")
bidsheet_df.insert(mcs_col_idx + 1, "Normalized incumbent supplier", normalized_incumbent_supplier)

# Optionally, drop the helper lowercase column if not needed anymore
bidsheet_df.drop(columns=['most_common_supplier_lower'], inplace=True)

# Drop all rows where Normalized incumbent supplier is "Bugatti Group"
bidsheet_df = bidsheet_df[bidsheet_df['Normalized incumbent supplier'] != "Bugatti Group"].reset_index(drop=True)

bidsheet_df['Annual Volume (per UOM)'] = pd.to_numeric(bidsheet_df['Annual Volume (per UOM)'], errors='coerce')
bidsheet_df['Volume-banded WAPP'] = pd.to_numeric(bidsheet_df['Volume-banded WAPP'], errors='coerce')

# Calculate Extended Cost USD
bidsheet_df['Extended Cost USD'] = (bidsheet_df['Annual Volume (per UOM)'] * bidsheet_df['Volume-banded WAPP']).round(4)

# Move Extended Cost USD next to Volume-banded WAPP
wapp_idx = bidsheet_df.columns.get_loc("Volume-banded WAPP")
ext_cost = bidsheet_df.pop("Extended Cost USD")
bidsheet_df.insert(wapp_idx + 1, "Extended Cost USD", ext_cost)

# point 6 & 7 from the mail.
valid_supplier_idx = bidsheet_df.columns.get_loc("Valid Supplier")

def calculate_as_is_r1(row):
    vol_wapp = row["Volume-banded WAPP"]
    valid_sup = row["Valid Supplier"]
    norm_inc_supplier = row["Normalized incumbent supplier"]

    # Check Volume-banded WAPP and Valid Supplier non-zero and not NaN
    if pd.isna(vol_wapp) or vol_wapp == 0 or pd.isna(valid_sup) or valid_sup == 0:
        return "-"

    if not isinstance(norm_inc_supplier, str) or norm_inc_supplier.strip() == "":
        return "-"

    supplier_col = f"{norm_inc_supplier} - R1 - Total Cost Per UOM FOB Port of Origin/Departure (USD)"

    if supplier_col not in bidsheet_df.columns:
        # print(f"Warning: Supplier column '{supplier_col}' not found in DataFrame for row {row.name}.")
        return "-"

    supplier_r1_cost = row[supplier_col]

    # Check supplier cost non-zero and not NaN
    if pd.isna(supplier_r1_cost) or supplier_r1_cost == 0:
        return "-"

    try:
        result = (vol_wapp - supplier_r1_cost) / vol_wapp
        return round(result, 4)  # rounded to 6 decimals, change if needed
    except Exception as e:
        print(e)
        return "-"
    
bidsheet_df.insert(valid_supplier_idx + 1, "As Is R1 %", bidsheet_df.apply(calculate_as_is_r1, axis=1))
# Get the index of "As Is R1 %" column (which was just inserted)
as_is_r1_pct_idx = bidsheet_df.columns.get_loc("As Is R1 %")

def calculate_as_is_r1_usd(row):
    vol_wapp = row["Volume-banded WAPP"]
    valid_sup = row["Valid Supplier"]
    as_is_r1_pct = row["As Is R1 %"]
    ext_cost_usd = row.get("Extended Cost USD", None)

    # Check required fields
    if pd.isna(vol_wapp) or vol_wapp == 0 or pd.isna(valid_sup) or valid_sup == 0:
        return "-"
    if as_is_r1_pct == "-" or pd.isna(as_is_r1_pct):
        return "-"
    if ext_cost_usd is None or pd.isna(ext_cost_usd):
        return "-"

    try:
        result = as_is_r1_pct * ext_cost_usd
        return round(result, 4)  # round as needed
    except Exception as e:
        print(e)
        return "-"

# Insert new column next to "As Is R1 %"
bidsheet_df.insert(as_is_r1_pct_idx + 1, "As Is R1 USD", bidsheet_df.apply(calculate_as_is_r1_usd, axis=1))
# AS IS USING R2
normalized_incumbent_supplier_idx = bidsheet_df.columns.get_loc("Normalized incumbent supplier")

def calculate_as_is_final(row):
    vol_wapp = row["Volume-banded WAPP"]
    valid_sup = row["Valid Supplier"]
    norm_inc_supplier = row["Normalized incumbent supplier"]

    # Check Volume-banded WAPP and Valid Supplier non-zero and not NaN
    if pd.isna(vol_wapp) or vol_wapp == 0 or pd.isna(valid_sup) or valid_sup == 0:
        return "-"

    if not isinstance(norm_inc_supplier, str) or norm_inc_supplier.strip() == "":
        return "-"

    supplier_col = f"{norm_inc_supplier} - R2 - Total Cost Per UOM FOB Port of Origin/Departure (USD)"

    if supplier_col not in bidsheet_df.columns:
        # print(f"Warning: Supplier column '{supplier_col}' not found in DataFrame for row {row.name}.")
        return "-"

    supplier_r2_cost = row[supplier_col]

    # Check supplier cost non-zero and not NaN
    if pd.isna(supplier_r2_cost) or supplier_r2_cost == 0:
        return "-"

    try:
        result = (vol_wapp - supplier_r2_cost) / vol_wapp
        return round(result, 4)  # rounded to 6 decimals, change if needed
    except Exception as e:
        print(e)
        return "-"
    
bidsheet_df.insert(normalized_incumbent_supplier_idx + 1, "As Is Final %", bidsheet_df.apply(calculate_as_is_final, axis=1))
# Get the index of "As Is Final %" column (which was just inserted)

as_is_final_pct_idx = bidsheet_df.columns.get_loc("As Is Final %")

def calculate_as_is_final_usd(row):
    vol_wapp = row["Volume-banded WAPP"]
    valid_sup = row["Valid Supplier"]
    as_is_final_pct = row["As Is Final %"]
    ext_cost_usd = row.get("Extended Cost USD", None)

    # Check required fields
    if pd.isna(vol_wapp) or vol_wapp == 0 or pd.isna(valid_sup) or valid_sup == 0:
        return "-"
    if as_is_final_pct == "-" or pd.isna(as_is_final_pct):
        return "-"
    if ext_cost_usd is None or pd.isna(ext_cost_usd):
        return "-"

    try:
        result = as_is_final_pct * ext_cost_usd
        return round(result, 4)  # round as needed
    except Exception as e:
        print(e)
        return "-"

# Insert new column next to "As Is R1 %"
bidsheet_df.insert(as_is_final_pct_idx + 1, "As Is Final USD", bidsheet_df.apply(calculate_as_is_final_usd, axis=1))


bidsheet_df['Final Min Bid'] = pd.to_numeric(bidsheet_df['Final Min Bid'], errors='coerce')
bidsheet_df['Cherry Pick min Final %'] = ((bidsheet_df['Volume-banded WAPP'] - bidsheet_df['Final Min Bid']) / bidsheet_df['Volume-banded WAPP']).round(4)
bidsheet_df.loc[bidsheet_df['Volume-banded WAPP'] == 0, 'Cherry Pick min Final %'] = np.nan

bidsheet_df['Cherry Pick min Final USD'] = (pd.to_numeric(bidsheet_df['Cherry Pick min Final %'], errors='coerce') * bidsheet_df['Extended Cost USD']).round(4)
bidsheet_df.loc[bidsheet_df['Cherry Pick min Final %'].isna(), 'Cherry Pick min Final USD'] = np.nan

mbs_idx = bidsheet_df.columns.get_loc("Final Minimum Bid Supplier")
bidsheet_df.insert(mbs_idx, "Cherry Pick min Final %", bidsheet_df.pop("Cherry Pick min Final %"))
bidsheet_df.insert(mbs_idx+1, "Cherry Pick min Final USD", bidsheet_df.pop("Cherry Pick min Final USD"))

# cherry pick for min R1
# min_bid_r1_idx = bidsheet_df.columns.get_loc("Min Bid R1")
# bidsheet_df.insert(
#     min_bid_r1_idx + 1,
#     "Cherry Pick min R1 %",
#     bidsheet_df.apply(
#         lambda row: (
#             (row['Volume-banded WAPP'] - row['Min Bid R1']) / row['Volume-banded WAPP']
#         ) if (
#             pd.notna(row['Volume-banded WAPP']) and row['Volume-banded WAPP'] != 0
#             and pd.notna(row['Valid Supplier']) and row['Valid Supplier'] != 0
#             and pd.notna(row['Min Bid R1'])
#         ) else '-',
#         axis=1
#     )
# )
# # Find index of "Cherry Pick min R1 %" column
# cherry_pick_min_r1_pct_idx = bidsheet_df.columns.get_loc("Cherry Pick min R1 %")

# bidsheet_df.insert(
#     cherry_pick_min_r1_pct_idx + 1,
#     "Cherry Pick min R1 USD",
#     bidsheet_df.apply(
#         lambda row: (
#             row["Cherry Pick min R1 %"] * row["Extended Cost USD"]
#         ) if (
#             pd.notna(row["Volume-banded WAPP"]) and row["Volume-banded WAPP"] != 0
#             and pd.notna(row["Valid Supplier"]) and row["Valid Supplier"] != 0
#             and pd.notna(row["Cherry Pick min R1 %"]) and row["Cherry Pick min R1 %"] != "-"
#             and pd.notna(row["Extended Cost USD"])
#         ) else "-",
#         axis=1
#     )
# )

# # Find index of "Cherry Pick min R1 USD" column
# awardable_col_idx = bidsheet_df.columns.get_loc("Cherry Pick min R1 USD")

# Function to determine if a bid is awardable based on Cherry Pick min R1 %
# def awardable_min_bid_r1(row):
#     vol_wapp = row["Volume-banded WAPP"]
#     valid_sup = row["Valid Supplier"]
#     cherry_pick_pct = row["Cherry Pick min R1 %"]

#     if pd.isna(vol_wapp) or vol_wapp == 0 or pd.isna(valid_sup) or valid_sup == 0:
#         return "No baseline or bid"

#     if cherry_pick_pct == "-" or pd.isna(cherry_pick_pct):
#         return "No baseline or bid"

#     try:
#         if cherry_pick_pct > 0:
#             return "Yes"
#         else:
#             return "No"
#     except Exception:
#         return "No baseline or bid"

# # Insert new column next to "Cherry Pick min R1 USD" called "Awardable Min Bid R1 (+0% savings)"
# bidsheet_df.insert(
#     awardable_col_idx + 1,
#     "Awardable Min Bid R1 (+0% savings)",
#     bidsheet_df.apply(awardable_min_bid_r1, axis=1)
# )

'''
add a column "Min improved R2 vs R1" right next to "Final Min Bid" with this logic:
If "Volume-banded WAPP" is missing or zero OR "Valid Supplier" is zero or missing → "No baseline or bid"
Else if "Final Min Bid" < "Min Bid R1" → "Yes"
Else if "Final Min Bid" ≥ "Min Bid R1" → "No"
'''
# final_min_bid_idx = bidsheet_df.columns.get_loc("Final Min Bid")

# def min_improved_r2_vs_r1(row):
#     vol_wapp = row["Volume-banded WAPP"]
#     valid_sup = row["Valid Supplier"]
#     final_min_bid = row["Final Min Bid"]
#     min_bid_r1 = row["Min Bid R1"]

#     if pd.isna(vol_wapp) or vol_wapp == 0 or pd.isna(valid_sup) or valid_sup == 0:
#         return "No baseline or bid"
#     if pd.isna(final_min_bid) or pd.isna(min_bid_r1):
#         return "No baseline or bid"

#     try:
#         if final_min_bid < min_bid_r1:
#             return "Yes"
#         else:
#             return "No"
#     except Exception:
#         return "No baseline or bid"

# bidsheet_df.insert(
#     final_min_bid_idx + 1,
#     "Min improved R2 vs R1",
#     bidsheet_df.apply(min_improved_r2_vs_r1, axis=1)
# )

'''
Then next to the Cherry Pick min Final USD can we add a column called "Awardable Min Bid Final (+0% savings)" 
and in it have the value be Yes if Cherry Pick min Final %  > 0%, No if <= 0% and "No baseline or bid" if there is no Volume Banded WAPP and/or Valid Supplier is 0

'''

cherry_pick_final_usd_idx = bidsheet_df.columns.get_loc("Cherry Pick min Final USD")

def awardable_min_bid_final(row):
    vol_wapp = row["Volume-banded WAPP"]
    valid_sup = row["Valid Supplier"]
    cherry_pick_final_pct = row.get("Cherry Pick min Final %", None)

    if pd.isna(vol_wapp) or vol_wapp == 0 or pd.isna(valid_sup) or valid_sup == 0:
        return "No baseline or bid"
    if cherry_pick_final_pct == "-" or pd.isna(cherry_pick_final_pct):
        return "No baseline or bid"

    try:
        return "Yes" if cherry_pick_final_pct > 0 else "No"
    except Exception:
        return "No baseline or bid"

bidsheet_df.insert(
    cherry_pick_final_usd_idx + 1,
    "Awardable Min Bid Final (+0% savings)",
    bidsheet_df.apply(awardable_min_bid_final, axis=1)
)

volume_banded_wapp_freight_idx = bidsheet_df.columns.get_loc("Volume-banded WAPP")
def calculate_volume_banded_wapp_with_freight(row):
    vol_wapp = row["Volume-banded WAPP"]
    row_id = row['ROW ID #']
    
    division = row['Division']
    incumbent_supplier = row['Normalized incumbent supplier']

    if incumbent_supplier == '-':
        return "-"

    supplier_port = supplier_port_df.loc[
        (supplier_port_df['ROW ID #'] == row_id),
        incumbent_supplier
    ].values[0]

    freight_multiplier = frieght_file_df.loc[
        (frieght_file_df['Reference'] == supplier_port),
        division
    ].values[0]

    # Get country from port-country map
    country = port_country_map.get(supplier_port, None)
    # Normalize metal type for lookup
    def normalize_metal_type(metal):
        if not isinstance(metal, str):
            return 'zinc, copper, iron and all other'
        m = metal.strip().lower()
        if m in ['aluminum', 'brass', 'steel', 'copper', 'iron', 'zinc', 'lead-free bronze', 'bronze']:
            return m
        return 'zinc, copper, iron and all other'
    metal_type = normalize_metal_type(row.get('type', ''))
    # Build tariff lookup dict if not already
    
    tariff_lookup = {(c.lower(), m.lower()): v for c, m, v in tariff_data}
    tariff_multiplier = tariff_lookup.get((str(country).strip().lower(), metal_type), 0)
    duty_multiplier = get_duty_multiplier(row['type'], '')
    
    if pd.isna(vol_wapp) or vol_wapp == 0 or pd.isna(supplier_port) or pd.isna(freight_multiplier):
        return "-"

    return round(((vol_wapp * freight_multiplier) + (vol_wapp * tariff_multiplier) + (vol_wapp * duty_multiplier)), 4)

bidsheet_df.insert(
    volume_banded_wapp_freight_idx + 1,
    "Volume-banded WAPP Landed Cost",
    bidsheet_df.apply(calculate_volume_banded_wapp_with_freight ,axis=1)
)

pos = bidsheet_df.columns.get_loc("Extended Cost USD")
bidsheet_df['Volume-banded WAPP Landed Cost'] = pd.to_numeric(bidsheet_df['Volume-banded WAPP Landed Cost'], errors='coerce')
bidsheet_df.insert(pos + 1, "Landed Extended Cost USD", bidsheet_df["Annual Volume (per UOM)"] * bidsheet_df["Volume-banded WAPP Landed Cost"])

supplier_pattern = re.compile(r"^(.*?) - R([12]) - Total Cost Per UOM FOB Port of Origin/Departure \(USD\)$")
supplier_r1_map, supplier_r2_map = {}, {}

for col in bidsheet_df.columns[30:]:
    m = supplier_pattern.match(col)
    if m:
        name, round_num = m.group(1), m.group(2)
        (supplier_r1_map if round_num == '1' else supplier_r2_map)[name] = col

suppliers = sorted(set(supplier_r1_map) & set(supplier_r2_map))
missing_r1 = set(supplier_r2_map) - set(supplier_r1_map)
missing_r2 = set(supplier_r1_map) - set(supplier_r2_map)

# if missing_r1 or missing_r2:
#     print("Warning: Skipped suppliers due to missing R1/R2:")
#     if missing_r1: print("  Missing R1:", missing_r1)
#     if missing_r2: print("  Missing R2:", missing_r2)

supplier_new_cols = {}
supplier_column_order = []

for s in tqdm(suppliers, desc='Supplier Calcs'):
    r1_col = supplier_r1_map[s]
    r2_col = supplier_r2_map[s]
    r1 = pd.to_numeric(bidsheet_df[r1_col], errors='coerce')
    r2 = pd.to_numeric(bidsheet_df[r2_col], errors='coerce')

    # Prepare merge base
    temp_df = bidsheet_df[['ROW ID #', 'Division', 'Part #']].copy()
    temp_df['Supplier'] = s
    temp_df[r2_col] = r2  # Ensure we include numeric R2 for calc
    temp_df[r2_col] = pd.to_numeric(bidsheet_df[r2_col], errors='coerce')

    temp_df['Metal Type'] = bidsheet_df['type']

    # Join freight multipliers
    merged = temp_df.merge(
        freight_lookup_df[freight_lookup_df['Supplier'] == s][['ROW ID #', 'Division', 'Part #', 'Supplier', 'Freight Multiplier', 'Country']],
        on=['ROW ID #', 'Division', 'Part #', 'Supplier'],
        how='left'
    )
    merged = merged.merge(
        tariff_df,
        on=['Country', 'Metal Type'],
        how='left'
    )
    merged['Duty Multiplier'] = merged['Metal Type'].apply(lambda mt: get_duty_multiplier(mt, supplier=s))

    merged['Freight Multiplier'] = merged['Freight Multiplier'].fillna(0)
    merged['Tariff Multiplier'] = merged['Tariff Multiplier'].fillna(0)
    merged['Duty Multiplier'] = merged['Duty Multiplier'].fillna(0)

    # print(f"\n--- Debug for Supplier: {s} ---")
    # print(merged[[r2_col, 'Freight Multiplier', 'Tariff Multiplier', 'Duty Multiplier']].head())
    # print("Non-null R2 count:", merged[r2_col].notna().sum())
    # print("Non-null Freight Multiplier:", merged['Freight Multiplier'].notna().sum())
    # print("Non-null Tariff Multiplier:", merged['Tariff Multiplier'].notna().sum())
    # print("Non-null Duty Multiplier:", merged['Duty Multiplier'].notna().sum())
    
    # print("Unique countries in merged:", merged['Country'].dropna().unique())

    merged['Landed Cost'] = (
        (merged[r2_col] * merged['Freight Multiplier']) + (merged[r2_col] * merged['Tariff Multiplier']) + (merged[r2_col] * merged['Duty Multiplier'])
    ).round(4)

    wapp = bidsheet_df['Volume-banded WAPP']
    wapp_landed = bidsheet_df['Volume-banded WAPP Landed Cost']
    ext_cost = bidsheet_df['Extended Cost USD']
    landed_ext_cost = bidsheet_df['Landed Extended Cost USD']

    delta_pct = ((r1 - r2) / r1).where((r1 != 0) & (r2 != 0)).round(4)
    delta_usd = (delta_pct * ext_cost).round(4)
    final_pct = ((wapp - r2) / wapp).where((r2 !=0) & (wapp != 0)).round(4)
    final_usd = (final_pct * ext_cost).round(4)
    
    supplier_new_cols[f"{s} - Final % savings vs baseline"] = final_pct
    supplier_new_cols[f"{s} - Final USD savings vs baseline"] = final_usd
    supplier_new_cols[f"{s} - R2 - Total landed cost per UOM (USD)"] = merged['Landed Cost']

    merged['Landed Cost'] = pd.to_numeric(merged['Landed Cost'], errors='coerce')
    wapp_landed = pd.to_numeric(wapp_landed, errors='coerce')

    # Fix arithmetic operation
    final_landed_pct = ((wapp_landed - merged['Landed Cost']) / wapp_landed).where(
        (wapp_landed != '-') & (merged['Landed Cost'] != '-') & (merged['Landed Cost'].notna()) & (merged['Landed Cost'] != 0) & (wapp_landed != 0)
    ).round(4)
    final_landed_usd = (final_landed_pct * landed_ext_cost).round(4)

    supplier_new_cols[f"{s} - Final Landed % savings vs baseline"] = final_landed_pct
    supplier_new_cols[f"{s} - Final Landed USD savings vs baseline"] = final_landed_usd

    supplier_column_order.extend([
        r2_col,
        f"{s} - R2 - Total landed cost per UOM (USD)",
        f"{s} - Final % savings vs baseline",
        f"{s} - Final USD savings vs baseline",
        f"{s} - Final Landed % savings vs baseline",
        f"{s} - Final Landed USD savings vs baseline",
    ])

supplier_new_df = pd.DataFrame(supplier_new_cols)
bidsheet_df = pd.concat([bidsheet_df, supplier_new_df], axis=1).copy()

as_is_final_usd_idx = bidsheet_df.columns.get_loc("As Is Final USD")
def calculate_as_is_final_landed(row):
    vol_wapp_landed = row["Volume-banded WAPP Landed Cost"]
    valid_sup = row["Valid Supplier"]
    norm_inc_supplier = row["Normalized incumbent supplier"]

    # Check Volume-banded WAPP and Valid Supplier non-zero and not NaN
    if pd.isna(vol_wapp_landed) or vol_wapp_landed == 0 or pd.isna(valid_sup) or valid_sup == 0:
        return "-"

    if not isinstance(norm_inc_supplier, str) or norm_inc_supplier.strip() == "":
        return "-"

    supplier_col = f"{norm_inc_supplier} - R2 - Total landed cost per UOM (USD)"

    if supplier_col not in bidsheet_df.columns:
        # print(f"Warning: Supplier column '{supplier_col}' not found in DataFrame for row {row.name}.")
        return "-"

    supplier_r2_landed_cost = row[supplier_col]

    # Check supplier cost non-zero and not NaN
    if pd.isna(supplier_r2_landed_cost) or supplier_r2_landed_cost == 0:
        return "-"

    try:
        result = (vol_wapp_landed - supplier_r2_landed_cost) / vol_wapp_landed
        return round(result, 4)  # rounded to 4 decimals, change if needed
    except Exception as e:
        print(e)
        return "-"

# Insert new column next to "As Is R1 %"
bidsheet_df.insert(as_is_final_usd_idx + 1, "As Is Final Landed %", bidsheet_df.apply(calculate_as_is_final_landed, axis=1))


as_is_final_pct_idx = bidsheet_df.columns.get_loc("As Is Final Landed %")

def calculate_as_is_final_landed_usd(row):
    vol_wapp = row["Volume-banded WAPP Landed Cost"]
    valid_sup = row["Valid Supplier"]
    as_is_final_landed_pct = row["As Is Final Landed %"]
    ext_cost_usd = row.get("Landed Extended Cost USD", None)

    # Check required fields
    if pd.isna(vol_wapp) or vol_wapp == 0 or pd.isna(valid_sup) or valid_sup == 0:
        return "-"
    if as_is_final_landed_pct == "-" or pd.isna(as_is_final_landed_pct):
        return "-"
    if ext_cost_usd is None or pd.isna(ext_cost_usd):
        return "-"

    try:
        result = as_is_final_landed_pct * ext_cost_usd
        return round(result, 4)  # round as needed
    except Exception as e:
        print(e)
        return "-"

# Insert new column next to "As Is R1 %"
bidsheet_df.insert(as_is_final_pct_idx + 1, "As Is Final Landed USD", bidsheet_df.apply(calculate_as_is_final_landed_usd, axis=1))



# Build supplier to R1/R2 column mapping
supplier_round_cols = {}
for col in bidsheet_df.columns[33:]:
    if col.endswith("Total landed cost per UOM (USD)"):
        parts = col.split(" - ")
        if len(parts) >= 3:
            supplier = parts[0].strip()
            round_tag = parts[1].strip()
            if supplier not in supplier_round_cols:
                supplier_round_cols[supplier] = {}
            supplier_round_cols[supplier][round_tag] = col

# For each row, build a dict of supplier: value (R2 if present, else R1)
def get_supplier_values(row):
    values = {}
    for supplier, rounds in supplier_round_cols.items():
        val = None
        if 'R2' in rounds:
            v = row[rounds['R2']]
            if not pd.isna(v) and v != '' and float(v) != 0:
                val = float(v)
        if val is None and 'R1' in rounds:
            v = row[rounds['R1']]
            if not pd.isna(v) and v != '' and float(v) != 0:
                val = float(v)
        if val is not None:
            values[supplier] = val
    return values
# === Step 6: Min/2nd Min/Outlier Flag ===
final_landed_min_bids, final_landed_min_bids_supplier, second_landed_min_bids, second_landed_min_suppliers = [], [], [], []
for idx, row in bidsheet_df.iterrows():
    supplier_vals = get_supplier_values(row)
    for supplier, val in supplier_vals.items():
        for round_tag in ['R2']:
            if round_tag in supplier_round_cols[supplier]:
                col = supplier_round_cols[supplier][round_tag]

    if supplier_vals:
        sorted_bids = sorted(supplier_vals.items(), key=lambda x: x[1])
        min_supplier, min_bid = sorted_bids[0]
        second_min_supplier, second_min_bid = sorted_bids[1] if len(sorted_bids) > 1 else ("-", "-")
    else:
        min_bid = min_supplier = second_min_bid = second_min_supplier = "-"
        

    final_landed_min_bids.append(min_bid)
    final_landed_min_bids_supplier.append(min_supplier)
    second_landed_min_bids.append(second_min_bid)
    second_landed_min_suppliers.append(second_min_supplier)


pos = bidsheet_df.columns.get_loc("Final 2nd Lowest Bid Supplier")

bidsheet_df.insert(pos+1, "Final Min Bid Landed", final_landed_min_bids)
bidsheet_df.insert(pos+2, "Final Minimum Bid Landed Supplier", final_landed_min_bids_supplier)
bidsheet_df.insert(pos+3, "2nd Lowest Landed Bid", second_landed_min_bids)
bidsheet_df.insert(pos+4, "2nd Lowest Bid Landed Supplier", second_landed_min_suppliers)

bidsheet_df['Annual Volume (per UOM)'] = pd.to_numeric(bidsheet_df['Annual Volume (per UOM)'], errors='coerce')
bidsheet_df['Volume-banded WAPP Landed Cost'] = pd.to_numeric(bidsheet_df['Volume-banded WAPP Landed Cost'], errors='coerce')


bidsheet_df['Final Min Bid Landed'] = pd.to_numeric(bidsheet_df['Final Min Bid Landed'], errors='coerce')
bidsheet_df['Volume-banded WAPP Landed Cost'] = pd.to_numeric(bidsheet_df['Volume-banded WAPP Landed Cost'], errors='coerce')

bidsheet_df['Cherry Pick Landed Final %'] = ((bidsheet_df['Volume-banded WAPP Landed Cost'] - bidsheet_df['Final Min Bid Landed']) / bidsheet_df['Volume-banded WAPP Landed Cost']).where(
        (bidsheet_df['Volume-banded WAPP Landed Cost'] != '-') & (bidsheet_df['Final Min Bid Landed'] != '-') & (bidsheet_df['Final Min Bid Landed'].notna()) & (bidsheet_df['Final Min Bid Landed'] != 0) & (bidsheet_df['Volume-banded WAPP Landed Cost'] != 0)
    ).round(4)
bidsheet_df.loc[bidsheet_df['Volume-banded WAPP Landed Cost'] == 0, 'Cherry Pick Landed Final %'] = np.nan

bidsheet_df['Landed Extended Cost USD'] = pd.to_numeric(bidsheet_df['Landed Extended Cost USD'], errors='coerce')

bidsheet_df['Cherry Pick Landed Final USD'] = (pd.to_numeric(bidsheet_df['Cherry Pick Landed Final %'], errors='coerce') * bidsheet_df['Landed Extended Cost USD']).round(4)
bidsheet_df.loc[bidsheet_df['Cherry Pick Landed Final %'].isna(), 'Cherry Pick Landed Final USD'] = np.nan

mbs_landed_idx = bidsheet_df.columns.get_loc("Final Minimum Bid Landed Supplier")
bidsheet_df.insert(mbs_landed_idx, "Cherry Pick Landed Final %", bidsheet_df.pop("Cherry Pick Landed Final %"))
bidsheet_df.insert(mbs_landed_idx+1, "Cherry Pick Landed Final USD", bidsheet_df.pop("Cherry Pick Landed Final USD"))

cherry_pick_final_landed_usd_idx = bidsheet_df.columns.get_loc("Cherry Pick Landed Final USD")

def awardable_min_bid_final_landed(row):
    vol_wapp = row["Volume-banded WAPP Landed Cost"]
    valid_sup = row["Valid Supplier"]
    cherry_pick_final_pct = row.get("Cherry Pick Landed Final %", None)

    if pd.isna(vol_wapp) or vol_wapp == 0 or pd.isna(valid_sup) or valid_sup == 0:
        return "No baseline or bid"
    if cherry_pick_final_pct == "-" or pd.isna(cherry_pick_final_pct):
        return "No baseline or bid"

    try:
        return "Yes" if cherry_pick_final_pct > 0 else "No"
    except Exception:
        return "No baseline or bid"

bidsheet_df.insert(
    cherry_pick_final_landed_usd_idx + 1,
    "Awardable Min Bid Final Landed(+0% savings)",
    bidsheet_df.apply(awardable_min_bid_final_landed, axis=1)
)

# Reorder columns to enforce the supplier grouping order
pre_supplier_cols = list(bidsheet_df.columns[:40])
post_supplier_cols = [col for col in bidsheet_df.columns if col not in pre_supplier_cols and col not in supplier_column_order]
bidsheet_df = bidsheet_df[pre_supplier_cols + supplier_column_order + post_supplier_cols]

os.makedirs("new", exist_ok=True)

# Remove all columns containing 'R1 - Total Cost Per UOM FOB Port of Origin/Departure (USD)'
cols_to_remove = [col for col in bidsheet_df.columns if 'R1 - Total Cost Per UOM FOB Port of Origin/Departure (USD)' in str(col)]
bidsheet_df.drop(columns=cols_to_remove, inplace=True)

bidsheet_df.to_excel(output_file, index=False)

wb = load_workbook(output_file)
ws = wb.active
header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

# --- Define Fill Colors ---
fill_purple = PatternFill(start_color="800080", end_color="800080", fill_type="solid")
fill_red    = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
fill_orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
fill_green  = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
fill_subtle_grey = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")  # subtle grey
fill_subtle_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # subtle blue

# --- Float formatting columns ---
float_cols = [14, 15, 16, 17, 18, 19, 22, 23, 24, 25, 26, 27, 28, 32, 34, 35, 36, 39] + list(range(41, len(header)+1))  # 1-based indices
# Predefine substrings and last 5 column indices
target_substrings = [
    "- Final % savings vs baseline", 
    "- Final USD savings vs baseline"
]

# Columns (1-based) matching any of the substrings
special_col_indices = {
    idx + 1 for idx, col in enumerate(header)
    if any(substr in str(col) for substr in target_substrings)
}

# Combine logic: precompute which columns get '-' on empty
dash_fill_cols = special_col_indices.union(set(range(len(header) - 4 + 1, len(header) + 1)))  # 1-based
number_format = '0.0000'
for col_idx in tqdm(float_cols, desc='Float formatting'):
    if col_idx > len(header):
        continue

    is_dash_fill_col = col_idx in dash_fill_cols

    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        value = cell.value

        if isinstance(value, (int, float)):
            if cell.number_format != number_format:
                cell.number_format = number_format
        elif value in [None, '']:
            if is_dash_fill_col:
                cell.value = '-'
            else:
                cell.value = 0
                cell.number_format = number_format
# # Now do formatting
# for col_idx in tqdm(float_cols, desc='Float formatting'):
#     if col_idx > len(header):
#         continue

#     is_dash_fill_col = col_idx in dash_fill_cols

#     for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
#         cell = row[0]  # only one column

#         if isinstance(cell.value, (int, float)):
#             cell.number_format = '0.0000'
#         elif cell.value in [None, '']:
#             if is_dash_fill_col:
#                 cell.value = '-'
#             else:
#                 cell.value = 0
#                 cell.number_format = '0.0000'

# --- Yellow fill for last 5 columns ---
last_5_col_indices = range(len(header)-4, len(header)+1)  # 1-based
for col_idx in tqdm(last_5_col_indices, desc='Yellow fill (last 5 cols)'):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
        for cell in row:
            cell.fill = fill_yellow

# --- Color fill logic for 24+ columns with specific header ---
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

# Predefine fills only once
fill_map = {
    "purple": fill_purple,
    "red": fill_red,
    "orange": fill_orange,
    "green": fill_green
}

# Cache WAPP values to avoid repeatedly accessing cells
wapp_col_idx = header.index("Volume-banded WAPP") + 1  # 1-based
wapp_values = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    row_num = row[0].row
    wapp_cell = row[wapp_col_idx - 1]
    try:
        wapp_values[row_num] = float(wapp_cell.value)
    except (TypeError, ValueError):
        wapp_values[row_num] = None  # Mark invalid

# Pre-filter target columns once
target_col_idxs = [i + 1 for i, h in enumerate(header) if "Total Cost Per UOM FOB Port of Origin/Departure (USD)" in str(h)]

# Iterate once over rows and apply fill efficiently
for row in tqdm(ws.iter_rows(min_row=2, max_row=ws.max_row), desc='Bid color fill logic'):
    row_num = row[0].row
    wapp = wapp_values.get(row_num)
    if wapp in (None, 0):
        continue

    for col_idx in target_col_idxs:
        bid_cell = row[col_idx - 1]
        try:
            bid = float(bid_cell.value)
        except (TypeError, ValueError):
            continue

        if bid == 0:
            continue

        diff_ratio = (wapp - bid) / wapp

        if diff_ratio < -0.40:
            bid_cell.fill = fill_map["purple"]
        elif -0.40 <= diff_ratio <= 0:
            bid_cell.fill = fill_map["red"]
        elif 0 < diff_ratio <= 0.40:
            bid_cell.fill = fill_map["orange"]
        elif diff_ratio > 0.40:
            bid_cell.fill = fill_map["green"]

# --- Optimized Header coloring for R1/R2 columns with tqdm ---
fill_map = {
    "R1 - Total Cost Per UOM FOB Port of Origin/Departure (USD)": fill_subtle_grey,
    "R2 - Total Cost Per UOM FOB Port of Origin/Departure (USD)": fill_subtle_blue
}

for idx, col_header in tqdm(enumerate(header), total=len(header), desc="Coloring Headers"):
    header_str = str(col_header)
    for key, fill in fill_map.items():
        if key in header_str:
            ws.cell(row=1, column=idx+1).fill = fill
            break  # Stop after first match

# --- Green fill for 13th column (M) ---
col_13_letter = get_column_letter(13)  # 'M' for 13th column

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=13, max_col=13):
    cell = row[0]
    cell.fill = fill_green

wb.save(output_file)
print(f"\n✔ Done. Script run time: {time.time() - start_time:.2f} seconds")
