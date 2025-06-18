import pandas as pd
import os 
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from handler.handler import extract_supplier_name
import numpy as np
from scipy import stats

# All Folder location information 
FILES_FOLDER_LOCATION = "./files"
CLEANED_FILES_FOLDER_LOCATION = "./cleaned_files/bidsheet_brass"
CONSOLIDATED_FILE_LOCATION = "./consolidate"
CONSOLIDATE_FILE_NAME = "bidsheet_brass_outlier_consolidate"

# Color for outliers (Blue, in hex: #87CEEB)
OUTLIER_FILL = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')

# Column name related information 
CONSOLIDATED_COMMON_COLUMN_NAMES = ["ROW ID #", "Division", 'Part #', "Item Description "]
SUPPLIER_COLUMNS_NAMES = ["Price per UOM EXW (USD)", 
    "Freight Cost per UOM to Port of Origin/Departure (USD)", 
    "Total Cost Per UOM FOB Port of Origin/Departure (USD)",
    "Additional information (please use this column only if absolutely necessary)"
]
BASE_SUPPLIER_COLUMNS = ["Price per UOM EXW (USD)", 
    "Freight Cost per UOM to Port of Origin/Departure (USD)", 
    "Total Cost Per UOM FOB Port of Origin/Departure (USD)"
]
ADDITIONAL_INFO_COLUMN = "Additional information (please use this column only if absolutely necessary)"

NUMERIC_COLUMNS = [
    "Price per UOM EXW (USD)", 
    "Freight Cost per UOM to Port of Origin/Departure (USD)", 
    "Total Cost Per UOM FOB Port of Origin/Departure (USD)"
]

OUTPUT_FILE = f"{CONSOLIDATED_FILE_LOCATION}/{CONSOLIDATE_FILE_NAME}.xlsx"
OUTLIER_NUMERIC_COLUMNS = ["Total Cost Per UOM FOB Port of Origin/Departure (USD)"]


def format_numeric_value(value):
    if pd.isna(value) or value == "":
        return ""
    try:
        if isinstance(value, str):
            cleaned_value = value.replace('$', '').replace(',', '').strip()
            if cleaned_value == '' or cleaned_value.lower() == 'nan':
                return ""
            numeric_value = float(cleaned_value)
        else:
            numeric_value = float(value)
        formatted = f"{numeric_value:.4f}"
        formatted = formatted.rstrip('0').rstrip('.')
        return formatted
    except (ValueError, TypeError):
        return str(value) if value is not None else ""

def is_numeric_column(column_name):
    return column_name in NUMERIC_COLUMNS

files = os.listdir(FILES_FOLDER_LOCATION)

for item in files: 
    
    print("Processing excel sheet ---------------------------------------")
    print(f"./files/{item}")

    cleaned_csv_file_name = f"{item.split('.')[0]}_cleaned.xlsx"
    os.makedirs(CLEANED_FILES_FOLDER_LOCATION, exist_ok=True)
    df = pd.read_excel(f"./files/{item}", "1. Bidsheet Brass", header=None)
    header_row_index = None
    col_start_index = None
    for index, row in df.iterrows():
        for col_index, value in enumerate(row):
            if isinstance(value, str) and "ROW ID #" in value:
                header_row_index = index
                col_start_index = col_index
                break
        if header_row_index is not None:
            break
    if header_row_index is not None and col_start_index is not None:
        headers = df.iloc[header_row_index, col_start_index:].tolist()
        
        print("Headers information ---------------------------------------")
        print(headers)

        data_rows = df.iloc[header_row_index + 1:, col_start_index:]
        print(data_rows.iloc[0])
        data_rows.columns = headers
        data_rows.reset_index(drop=True, inplace=True)
        data_rows.to_excel(f"{CLEANED_FILES_FOLDER_LOCATION}/{cleaned_csv_file_name}", index=False)
        print(f"Data saved to '{CLEANED_FILES_FOLDER_LOCATION}/{cleaned_csv_file_name}'")
    else:
        print("No row containing 'ROW ID #' was found.")

def check_additional_info_data(all_data):
    suppliers_with_additional_info = set()
    for file_name, df in all_data.items():
        if ADDITIONAL_INFO_COLUMN in df.columns:
            additional_info_series = df[ADDITIONAL_INFO_COLUMN]
            has_data = additional_info_series.notna() & (additional_info_series.astype(str).str.strip() != "") & (additional_info_series.astype(str).str.strip() != "nan")
            if has_data.any():
                suppliers_with_additional_info.add(file_name)
                print(f"Supplier {file_name} has additional information data")
            else:
                print(f"Supplier {file_name} has NO additional information data")
        else:
            print(f"Supplier {file_name} doesn't have additional information column")
    return suppliers_with_additional_info

def get_supplier_columns(file_name, suppliers_with_additional_info):
    if file_name in suppliers_with_additional_info:
        return SUPPLIER_COLUMNS_NAMES
    else:
        return BASE_SUPPLIER_COLUMNS

def is_valid_supplier_value(value):
    if pd.isna(value) or value == "":
        return False
    str_value = str(value).strip()
    if str_value == "" or str_value.lower() == "nan":
        return False
    try:
        float_value = float(str_value.replace('$', '').replace(',', ''))
        return float_value != 0
    except (ValueError, TypeError):
        return True

def count_valid_suppliers(row_data, file_names, suppliers_with_additional_info):
    valid_count = 0
    for file_name in file_names:
        supplier_info = row_data['supplier_data'].get(file_name, {})
        supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
        has_valid_data = False
        for col_name in supplier_columns:
            value = supplier_info.get(col_name, "")
            if is_valid_supplier_value(value):
                has_valid_data = True
                break
        if has_valid_data:
            valid_count += 1
    return valid_count

def create_consolidated_dataset(all_data, suppliers_with_additional_info):
    consolidated_rows = {}
    for file_name, df in all_data.items():
        print(f"Processing data from {file_name}...")
        supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
        for _, row in df.iterrows():
            row_key_parts = []
            common_data = {}
            for col in CONSOLIDATED_COMMON_COLUMN_NAMES:
                if col in df.columns:
                    value = str(row[col]) if pd.notna(row[col]) else ""
                    row_key_parts.append(value)
                    common_data[col] = row[col] if pd.notna(row[col]) else ""
                else:
                    row_key_parts.append("")
                    common_data[col] = ""
            row_key = "|".join(row_key_parts)
            if row_key not in consolidated_rows:
                consolidated_rows[row_key] = {
                    'common_data': common_data,
                    'supplier_data': {}
                }
            supplier_data = {}
            for col in supplier_columns:
                if col in df.columns:
                    value = row[col] if pd.notna(row[col]) else ""
                    if is_numeric_column(col):
                        value = format_numeric_value(value)
                    supplier_data[col] = value
                else:
                    supplier_data[col] = ""
            consolidated_rows[row_key]['supplier_data'][file_name] = supplier_data
    return consolidated_rows

def has_valid_supplier_data(supplier_data):
    for value in supplier_data.values():
        if value != "" and value != 0 and value != "0" and value != "0.0000":
            return True
    return False

# >>>>> MODIFIED FOR ROW-WISE OUTLIER DETECTION <<<<<

def is_valid_nonzero(value):
    """Check if value is valid and non-zero for outlier analysis"""
    try:
        if pd.isna(value):
            return False
        str_val = str(value).strip()
        if str_val == "" or str_val.lower() == "nan":
            return False
        float_val = float(str_val.replace("$","").replace(",",""))
        return float_val != 0  # Exclude 0 values as requested
    except Exception:
        return False

def grubbs_test(values, alpha=0.05):
    """
    Return indices of outliers in the values using Grubbs' test (two-sided).
    Iteratively remove outliers until no more are found (optional).

    :param values: List or numpy array of values
    :param alpha: Significance level for Grubbs' test
    :return: Set of indices that are outliers
    """
    x = np.array(values)
    print(x)
    outlier_indices = set()
    original_indices = list(range(len(x)))

    while len(x) > 2:
        mean = np.mean(x)
        std = np.std(x, ddof=1)
        if std == 0:
            break

        abs_diffs = np.abs(x - mean)
        max_idx = np.argmax(abs_diffs)
        G = abs_diffs[max_idx] / std

        n = len(x)
        # Critical value for Grubbs' test
        t = stats.t.ppf(1 - alpha / (2 * n), n - 2)
        G_crit = ((n - 1) / np.sqrt(n)) * np.sqrt(t**2 / (n - 2 + t**2))

        if G > G_crit:
            # Remove outlier
            outlier_indices.add(original_indices[max_idx])
            # Remove from arrays for potential second outlier detection
            x = np.delete(x, max_idx)
            original_indices.pop(max_idx)
        else:
            break

    return outlier_indices
        
def detect_outliers_rowwise(values, method='auto'):
    """
    Given a list of numbers for a single row, return indices that are outliers.
    Uses Z-score method for small datasets (typical for supplier comparison within a row).
    """
    if len(values) < 0:
        return set()
    
    values = np.array(values, dtype='float')
    n = len(values)
    print(f"Row-wise outlier detection - Values for this row: {values}")
    
    outlier_idxs = set()
    
    # For row-wise analysis, we typically have few suppliers (2-10), so use Z-score method
    mean = np.mean(values)
    std = np.std(values, ddof=0)
    
    if std == 0:  # All values are the same
        return set()
    
    # Use 2 standard deviations as threshold (can be adjusted as needed)
    z_scores = np.abs(values - mean) / std
    outlier_idxs = set(np.where(z_scores > 2)[0])
    
    print(f"Mean: {mean:.4f}, Std: {std:.4f}")
    print(f"Z-scores: {z_scores}")
    print(f"Found {len(outlier_idxs)} outliers in this row: indices {outlier_idxs}")
    
    return outlier_idxs

def build_rowwise_outlier_lookup(consolidated_data, file_names, suppliers_with_additional_info):
    """
    Build a row-wise outlier lookup that maps (row_key, supplier_idx, col_name) to outlier status
    """
    print("Building row-wise outlier lookup...")
    
    outlier_lookup = {}
    
    for row_key, row_data in consolidated_data.items():
        # print(f"\nProcessing row: {row_key[:50]}...")  # Show first 50 chars of row key
        
        for col_name in OUTLIER_NUMERIC_COLUMNS:
            # Collect values for this specific row and column across all suppliers
            row_values = []
            supplier_indices = []
            
            for supplier_idx, file_name in enumerate(file_names):
                supplier_info = row_data['supplier_data'].get(file_name, {})
                supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
                
                if col_name not in supplier_columns:
                    continue
                    
                value = supplier_info.get(col_name, "")
                if is_valid_nonzero(value):
                    try:
                        val = float(str(value).replace("$", "").replace(",", ""))
                        row_values.append(val)
                        supplier_indices.append(supplier_idx)
                    except Exception:
                        continue
            
            # Only perform outlier detection if we have at least 2 valid values
            if len(row_values) >= 2:
                # print(f"Column {col_name}: {len(row_values)} valid values")
                
                # Detect outliers for this row
                outlier_indices = grubbs_test(row_values, alpha=0.05)
                
                # Map outlier indices back to supplier indices
                for outlier_idx in outlier_indices:
                    if outlier_idx < len(supplier_indices):
                        supplier_idx = supplier_indices[outlier_idx]
                        
                        # Initialize nested structure if needed
                        if row_key not in outlier_lookup:
                            outlier_lookup[row_key] = {}
                        if supplier_idx not in outlier_lookup[row_key]:
                            outlier_lookup[row_key][supplier_idx] = {}
                        
                        # Mark as outlier
                        outlier_lookup[row_key][supplier_idx][col_name] = True
                        print(f"Marked outlier: Supplier {supplier_idx}, Value: {row_values[outlier_idx]}")
            else:
                pass
                # print(f"Column {col_name}: Only {len(row_values)} valid values, skipping outlier detection")
    
    return outlier_lookup

# <<<<< END ROW-WISE OUTLIER LOGIC

def create_consolidated_excel(consolidated_data, file_names, suppliers_with_additional_info):
    # Build row-wise outlier lookup before creating Excel
    rowwise_outlier_lookup = build_rowwise_outlier_lookup(consolidated_data, file_names, suppliers_with_additional_info)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated Suppliers"
    light_gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    current_col = 1
    supplier_index = 0
    for col_name in CONSOLIDATED_COMMON_COLUMN_NAMES:
        header_cell = ws.cell(row=1, column=current_col, value="")
        header_cell.alignment = wrap_alignment
        current_col += 1
    valid_supplier_cell = ws.cell(row=1, column=current_col, value="")
    valid_supplier_cell.alignment = wrap_alignment
    valid_supplier_cell.fill = green_fill
    current_col += 1
    for file_name in file_names:
        supplier_user_information = extract_supplier_name(file_name)
        supplier_name = supplier_user_information[1]
        supplier_name = supplier_name.replace("_cleaned", "")
        supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
        background_fill = light_gray_fill if supplier_index % 2 == 0 else white_fill
        for col_name in supplier_columns:
            supplier_cell = ws.cell(row=1, column=current_col, value=supplier_name)
            supplier_cell.alignment = wrap_alignment
            supplier_cell.fill = background_fill
            current_col += 1
        supplier_index += 1
    current_col = 1
    for col_name in CONSOLIDATED_COMMON_COLUMN_NAMES:
        header_cell = ws.cell(row=2, column=current_col, value=col_name)
        header_cell.alignment = wrap_alignment
        current_col += 1
    valid_supplier_header = ws.cell(row=2, column=current_col, value="Supplier bids count")
    valid_supplier_header.alignment = wrap_alignment
    valid_supplier_header.fill = green_fill
    current_col += 1
    supplier_index = 0
    for file_name in file_names:
        supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
        background_fill = light_gray_fill if supplier_index % 2 == 0 else white_fill
        for col_name in supplier_columns:
            header_cell = ws.cell(row=2, column=current_col, value=col_name)
            header_cell.alignment = wrap_alignment
            header_cell.fill = background_fill
            current_col += 1
        supplier_index += 1

    row_num = 3
    for row_key, row_data in consolidated_data.items():
        current_col = 1
        for col_name in CONSOLIDATED_COMMON_COLUMN_NAMES:
            data_cell = ws.cell(row=row_num, column=current_col, value=row_data['common_data'].get(col_name, ""))
            data_cell.alignment = Alignment(wrap_text=True, vertical='top')
            current_col += 1
        valid_supplier_count = 0
        for file_name in file_names:
            supplier_info = row_data['supplier_data'].get(file_name, {})
            if has_valid_supplier_data(supplier_info):
                valid_supplier_count += 1
        valid_supplier_data_cell = ws.cell(row=row_num, column=current_col, value=valid_supplier_count)
        valid_supplier_data_cell.alignment = Alignment(wrap_text=True, vertical='top')
        if valid_supplier_count > 0:
            valid_supplier_data_cell.fill = green_fill
        current_col += 1

        # Fill in supplier data row cells, color blue if identified as outlier row-wise
        sup_idx = 0
        for file_name in file_names:
            supplier_info = row_data['supplier_data'].get(file_name, {})
            supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
            for col_name in supplier_columns:
                cell_value = supplier_info.get(col_name, "")
                data_cell = ws.cell(row=row_num, column=current_col, value=cell_value)
                data_cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # >>>>> ROW-WISE OUTLIER COLORING <<<<<
                if is_numeric_column(col_name) and col_name in OUTLIER_NUMERIC_COLUMNS:
                    # Check if this specific cell is marked as an outlier row-wise
                    is_outlier = (
                        row_key in rowwise_outlier_lookup and
                        sup_idx in rowwise_outlier_lookup[row_key] and
                        col_name in rowwise_outlier_lookup[row_key][sup_idx]
                    )
                    if is_outlier:
                        data_cell.fill = OUTLIER_FILL
                        print(f"Marked outlier: Row {row_num}, Supplier {sup_idx}, Column {col_name}, Value: {cell_value}")
                
                current_col += 1
            sup_idx += 1
        row_num += 1

    total_columns = len(CONSOLIDATED_COMMON_COLUMN_NAMES) + 1
    for file_name in file_names:
        supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
        total_columns += len(supplier_columns)
    for col_num in range(1, total_columns + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    for row_num in range(3, ws.max_row + 1):
        ws.row_dimensions[row_num].height = 30
    wb.save(OUTPUT_FILE)

def process_cleaned_files(): 
    try: 
        cleaned_files = os.listdir(CLEANED_FILES_FOLDER_LOCATION)
        excel_files = [f for f in cleaned_files if f.endswith(('.xlsx', '.xls'))]
        if not excel_files:
            print("No Excel files found in the cleaned files folder!")
            return
        print(f"Found {len(excel_files)} Excel files to process")

        all_data = {}
        all_rows_data = []
        
        for item in excel_files:
            print("Processing cleaned file ---------------------------------------")
            file_path = os.path.join(CLEANED_FILES_FOLDER_LOCATION, item)
            print(f"Processing: {file_path}")
            try:
                df = pd.read_excel(file_path)
                file_name_key = os.path.splitext(item)[0]
                all_data[file_name_key] = df
                print(f"Successfully loaded {len(df)} rows from {item}")
                print(f"Columns in file: {list(df.columns)}")
            except Exception as e:
                print(f"Error processing {item}: {str(e)}")
                continue
        if not all_data:
            print("No valid data found in any files!")
            return
        
        print("\nChecking for additional information data...")
        suppliers_with_additional_info = check_additional_info_data(all_data)
        
        print(f"Suppliers with additional info: {suppliers_with_additional_info}")
        print("\nCreating consolidated dataset...")
        consolidated_data = create_consolidated_dataset(all_data, suppliers_with_additional_info)
        
        print("Creating consolidated Excel file with row-wise outlier detection...")
        create_consolidated_excel(consolidated_data, list(all_data.keys()), suppliers_with_additional_info)
        
        print(f"\nConsolidated file created: {OUTPUT_FILE}")
        print(f"Suppliers including additional info column: {len(suppliers_with_additional_info)}")
        print(f"Suppliers excluding additional info column: {len(all_data) - len(suppliers_with_additional_info)}")
    except Exception as e: 
        print(f"Error in process_cleaned_files: {str(e)}")

process_cleaned_files()
