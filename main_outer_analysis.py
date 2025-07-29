import pandas as pd
import os 
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from handler.handler import extract_supplier_name
import numpy as np
from scipy import stats

# --------- All files related information --------------- # 
# --------- Contains file information ------------------- # 
# 1. Bidsheet Brass 
# 2. Bidsheet Steel
# 3. Bidsheet Other Metals

ALL_FILES = [
    {
        "consolidation_folder": "bidsheet_steel", 
        "consolidate_file_name": "bidsheet_steel_outlier_consolidate", 
        "sheet_name": "2. Bidsheet Steel",
    }, 
    {
        "consolidation_folder": "bidsheet_brass",
        "consolidate_file_name": "bidsheet_brass_outlier_consolidate",  
        "sheet_name": "1. Bidsheet Brass",
    }, 
    {
        "consolidation_folder": "bidsheet_other_metal",
        "consolidate_file_name": "bidsheet_other_metal_outlier_consolidate",
        "sheet_name": "3. Bidsheet Other Metals",
    }
]

for main_files_name_information in ALL_FILES:

    # All Folder location information 
    FILES_FOLDER_LOCATION = "./files"
    CLEANED_FILES_FOLDER_LOCATION = f"./cleaned_files/{main_files_name_information['consolidation_folder']}"
    CONSOLIDATED_FILE_LOCATION = "./consolidate"
    CONSOLIDATE_FILE_NAME = main_files_name_information["consolidate_file_name"]

    # Color for outliers (Blue, in hex: #87CEEB)
    OUTLIER_FILL = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')

    # Column name related information 
    CONSOLIDATED_COMMON_COLUMN_NAMES = ["ROW ID #", "Division", 'Part #', "Item Description ", "Product Group", "Part Family", "Average Order Quantity (per UOM)","Min Order Quantity (per UOM)","Max Order Quantity (per UOM)", "Order frequency",	"Annual Volume (per UOM)" ]

    ROW_KEY_COLUMNS = ["ROW ID #", "Division", 'Part #', "Item Description "]
    SUPPLIER_COLUMNS_NAMES = [
        "Total Cost Per UOM FOB Port of Origin/Departure (USD)",
    ]
    BASE_SUPPLIER_COLUMNS = [
        "Total Cost Per UOM FOB Port of Origin/Departure (USD)"
    ]
    ADDITIONAL_INFO_COLUMN = "DONOTCONSIDER"

    NUMERIC_COLUMNS = [
        "Price per UOM EXW (USD)", 
        "Freight Cost per UOM to Port of Origin/Departure (USD)", 
        "Total Cost Per UOM FOB Port of Origin/Departure (USD)"
    ]

    OUTPUT_FILE = f"{CONSOLIDATED_FILE_LOCATION}/{CONSOLIDATE_FILE_NAME}.xlsx"
    OUTLIER_NUMERIC_COLUMNS = ["Total Cost Per UOM FOB Port of Origin/Departure (USD)"]

    # ------------------- Helper functions --------------------- #
    # Convert numeric value to four decimal places and remove trailing zeros
    
    def format_numeric_value(value):
        """
            Convert a value to four decimal places and remove trailing zeros.
        """
        if pd.isna(value) or value == "":
            return ""
        try:
            if isinstance(value, str):
                cleaned_value = value.replace('$', '').replace(',', '').strip()
                if cleaned_value == '' or cleaned_value.lower() == 'nan':
                    return ""
                numeric_value = float(cleaned_value)
            else:
                numeric_value = float(value)
            formatted = f"{numeric_value:.4f}"
            formatted = formatted.rstrip('0').rstrip('.')
            return formatted
        except (ValueError, TypeError):
            return str(value) if value is not None else ""

    def is_numeric_column(column_name):
        return column_name in NUMERIC_COLUMNS

    # All files fetch from files folder -------------------------------------
    # processes it and moves it to cleaned files.
    files = os.listdir(FILES_FOLDER_LOCATION)

    for item in files: 
        
        print("Processing excel sheet ---------------------------------------")
        print(f"./files/{item}")

        cleaned_csv_file_name = f"{item.split('.')[0]}_cleaned.xlsx"
        os.makedirs(CLEANED_FILES_FOLDER_LOCATION, exist_ok=True)
        df = pd.read_excel(f"./files/{item}", main_files_name_information["sheet_name"], header=None)
        header_row_index = None
        col_start_index = None
        for index, row in df.iterrows():
            for col_index, value in enumerate(row):
                if isinstance(value, str) and "ROW ID #" in value:
                    header_row_index = index
                    col_start_index = col_index
                    break
            if header_row_index is not None:
                break
        if header_row_index is not None and col_start_index is not None:
            headers = df.iloc[header_row_index, col_start_index:].tolist()
            
            data_rows = df.iloc[header_row_index + 1:, col_start_index:]
            data_rows.columns = headers
            data_rows.reset_index(drop=True, inplace=True)
            data_rows.to_excel(f"{CLEANED_FILES_FOLDER_LOCATION}/{cleaned_csv_file_name}", index=False)
            print(f"Data saved to '{CLEANED_FILES_FOLDER_LOCATION}/{cleaned_csv_file_name}'")
        else:
            print("No row containing 'ROW ID #' was found.")

    def check_additional_info_data(all_data):
        suppliers_with_additional_info = set()
        for file_name, df in all_data.items():
            if ADDITIONAL_INFO_COLUMN in df.columns:
                additional_info_series = df[ADDITIONAL_INFO_COLUMN]
                has_data = additional_info_series.notna() & (additional_info_series.astype(str).str.strip() != "") & (additional_info_series.astype(str).str.strip() != "nan")
                if has_data.any():
                    suppliers_with_additional_info.add(file_name)
                    print(f"Supplier {file_name} has additional information data")
                else:
                    print(f"Supplier {file_name} has NO additional information data")
            else:
                print(f"Supplier {file_name} doesn't have additional information column")
        return suppliers_with_additional_info

    def get_supplier_columns(file_name, suppliers_with_additional_info):
        if file_name in suppliers_with_additional_info:
            return SUPPLIER_COLUMNS_NAMES
        else:
            return BASE_SUPPLIER_COLUMNS

    def is_valid_supplier_value(value):
        if pd.isna(value) or value == "":
            return False
        str_value = str(value).strip()
        if str_value == "" or str_value.lower() == "nan":
            return False
        try:
            float_value = float(str_value.replace('$', '').replace(',', ''))
            return float_value != 0
        except (ValueError, TypeError):
            return True

    def count_valid_suppliers(row_data, file_names, suppliers_with_additional_info):
        valid_count = 0
        for file_name in file_names:
            supplier_info = row_data['supplier_data'].get(file_name, {})
            supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
            has_valid_data = False
            for col_name in supplier_columns:
                value = supplier_info.get(col_name, "")
                if is_valid_supplier_value(value):
                    has_valid_data = True
                    break
            if has_valid_data:
                valid_count += 1
        return valid_count

    def create_consolidated_dataset(all_data, suppliers_with_additional_info):
        consolidated_rows = {}
        for file_name, df in all_data.items():
            print(f"Processing data from {file_name}...")
            supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
            for _, row in df.iterrows():
                row_key_parts = []
                common_data = {}
                for col in CONSOLIDATED_COMMON_COLUMN_NAMES:
                    if col in df.columns:
                        value = str(row[col]) if pd.notna(row[col]) else ""
                        row_key_parts.append(value)
                        common_data[col] = str(row[col]) if pd.notna(row[col]) else ""
                    else:
                        row_key_parts.append("")
                        common_data[col] = ""
                row_key = "|".join(row_key_parts)
                if row_key not in consolidated_rows:
                    consolidated_rows[row_key] = {
                        'common_data': common_data,
                        'supplier_data': {}
                    }
                supplier_data = {}
                for col in supplier_columns:
                    if col in df.columns:
                        value = row[col] if pd.notna(row[col]) else ""
                        if is_numeric_column(col):
                            value = format_numeric_value(value)
                        supplier_data[col] = value
                    else:
                        supplier_data[col] = ""
                consolidated_rows[row_key]['supplier_data'][file_name] = supplier_data
        return consolidated_rows

    def has_valid_supplier_data(supplier_data):
        for value in supplier_data.values():
            # If it's an int or float, check directly
            if isinstance(value, (int, float)) and value != 0:
                return True
           
            # If it's a string, try converting to float
            if isinstance(value, str):
                try:
                    numeric_value = float(value)
                    if numeric_value != 0:
                        return True
                except ValueError:
                    # Not a numeric string, skip
                    continue
 
        return False

    # --------------------- Row wise outlier detection ---------------- #

    def is_valid_nonzero(value):
        """Check if value is valid and non-zero for outlier analysis"""
        try:
            if pd.isna(value):
                return False
            str_val = str(value).strip()
            if str_val == "" or str_val.lower() == "nan":
                return False
            float_val = float(str_val.replace("$","").replace(",",""))
            return float_val != 0  # Exclude 0 values as requested
        except Exception:
            return False

    def grubbs_test(values, alpha=0.05):
        """
        Return indices of outliers in the values using Grubbs' test (two-sided).
        Iteratively remove outliers until no more are found (optional).

        :param values: List or numpy array of values
        :param alpha: Significance level for Grubbs' test
        :return: Set of indices that are outliers
        """
        x = np.array(values)
        print(x)
        outlier_indices = set()
        original_indices = list(range(len(x)))

        while len(x) > 2:
            mean = np.mean(x)
            std = np.std(x, ddof=1)
            if std == 0:
                break

            abs_diffs = np.abs(x - mean)
            max_idx = np.argmax(abs_diffs)
            G = abs_diffs[max_idx] / std

            n = len(x)
            # Critical value for Grubbs' test
            t = stats.t.ppf(1 - alpha / (2 * n), n - 2)
            G_crit = ((n - 1) / np.sqrt(n)) * np.sqrt(t**2 / (n - 2 + t**2))

            if G > G_crit:
                # Remove outlier
                outlier_indices.add(original_indices[max_idx])
                # Remove from arrays for potential second outlier detection
                x = np.delete(x, max_idx)
                original_indices.pop(max_idx)
            else:
                break

        return outlier_indices
            
    def detect_outliers_rowwise(values, method='auto'):
        """
        Given a list of numbers for a single row, return indices that are outliers.
        Uses Z-score method for small datasets (typical for supplier comparison within a row).
        """
        if len(values) < 0:
            return set()
        
        values = np.array(values, dtype='float')
        n = len(values)
        print(f"Row-wise outlier detection - Values for this row: {values}")
        
        outlier_idxs = set()
        
        # For row-wise analysis, we typically have few suppliers (2-10), so use Z-score method
        mean = np.mean(values)
        std = np.std(values, ddof=0)
        
        if std == 0:  # All values are the same
            return set()
        
        # Use 2 standard deviations as threshold (can be adjusted as needed)
        z_scores = np.abs(values - mean) / std
        outlier_idxs = set(np.where(z_scores > 2)[0])
        
        print(f"Mean: {mean:.4f}, Std: {std:.4f}")
        print(f"Z-scores: {z_scores}")
        print(f"Found {len(outlier_idxs)} outliers in this row: indices {outlier_idxs}")
        
        return outlier_idxs

    def build_rowwise_outlier_lookup(consolidated_data, file_names, suppliers_with_additional_info):
        """
        Build a row-wise outlier lookup that maps (row_key, supplier_idx, col_name) to outlier status
        """
        print("Building row-wise outlier lookup...")
        
        outlier_lookup = {}
        
        for row_key, row_data in consolidated_data.items():
            # print(f"\nProcessing row: {row_key[:50]}...")  # Show first 50 chars of row key
            
            for col_name in OUTLIER_NUMERIC_COLUMNS:
                # Collect values for this specific row and column across all suppliers
                row_values = []
                supplier_indices = []
                
                for supplier_idx, file_name in enumerate(file_names):
                    supplier_info = row_data['supplier_data'].get(file_name, {})
                    supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
                    
                    if col_name not in supplier_columns:
                        continue
                        
                    value = supplier_info.get(col_name, "")
                    if is_valid_nonzero(value):
                        try:
                            val = float(str(value).replace("$", "").replace(",", ""))
                            row_values.append(val)
                            supplier_indices.append(supplier_idx)
                        except Exception:
                            continue
                
                # Only perform outlier detection if we have at least 2 valid values
                if len(row_values) >= 2:
                    # print(f"Column {col_name}: {len(row_values)} valid values")
                    
                    # Detect outliers for this row
                    outlier_indices = grubbs_test(row_values, alpha=0.05)
                    
                    # Map outlier indices back to supplier indices
                    for outlier_idx in outlier_indices:
                        if outlier_idx < len(supplier_indices):
                            supplier_idx = supplier_indices[outlier_idx]
                            
                            # Initialize nested structure if needed
                            if row_key not in outlier_lookup:
                                outlier_lookup[row_key] = {}
                            if supplier_idx not in outlier_lookup[row_key]:
                                outlier_lookup[row_key][supplier_idx] = {}
                            
                            # Mark as outlier
                            outlier_lookup[row_key][supplier_idx][col_name] = True
                            print(f"Marked outlier: Supplier {supplier_idx}, Value: {row_values[outlier_idx]}")
                else:
                    pass
                    # print(f"Column {col_name}: Only {len(row_values)} valid values, skipping outlier detection")
        
        return outlier_lookup

    #  ------------------ New statistics calculation ------------------- #

    def calculate_row_statistics(row_data, file_names, suppliers_with_additional_info, target_column):
        """
        Calculate mean, variance, and standard deviation for a specific column across all suppliers in a row
        """
        values = []
        
        for file_name in file_names:
            supplier_info = row_data['supplier_data'].get(file_name, {})
            supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
            
            if target_column in supplier_columns:
                value = supplier_info.get(target_column, "")
                if is_valid_nonzero(value):
                    try:
                        val = float(str(value).replace("$", "").replace(",", ""))
                        values.append(val)
                    except Exception:
                        continue
        
        if len(values) == 0:
            return {"mean": "", "variance": "", "std_dev": ""}
        elif len(values) == 1:
            return {"mean": values[0], "variance": 0, "std_dev": 0}
        else:
            mean_val = np.mean(values)
            variance_val = np.var(values, ddof=1)  # Sample variance
            std_dev_val = np.std(values, ddof=1)   # Sample standard deviation
            
            return {
                "mean": round(mean_val, 4),
                "variance": round(variance_val, 4),
                "std_dev": round(std_dev_val, 4)
            }

    def create_consolidated_excel(consolidated_data, file_names, suppliers_with_additional_info):
        # Build row-wise outlier lookup before creating Excel
        rowwise_outlier_lookup = build_rowwise_outlier_lookup(consolidated_data, file_names, suppliers_with_additional_info)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Consolidated Suppliers"
        light_gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        # >>>>> NEW STATISTICAL COLUMN FILLS <<<<<
        yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')  # For statistical columns
        wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        current_col = 1
        supplier_index = 0
        
        # Common columns headers
        for col_name in CONSOLIDATED_COMMON_COLUMN_NAMES:
            header_cell = ws.cell(row=1, column=current_col, value="")
            header_cell.alignment = wrap_alignment
            current_col += 1
        
        # Valid supplier column
        valid_supplier_cell = ws.cell(row=1, column=current_col, value="")
        valid_supplier_cell.alignment = wrap_alignment
        valid_supplier_cell.fill = green_fill
        current_col += 1
        
        # Supplier columns headers (moved before statistical columns)
        for file_name in file_names:
            supplier_user_information = extract_supplier_name(file_name)
            supplier_name = supplier_user_information[1]
            supplier_name = supplier_name.replace("_cleaned", "")
            supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
            background_fill = light_gray_fill if supplier_index % 2 == 0 else white_fill
            for col_name in supplier_columns:
                supplier_cell = ws.cell(row=1, column=current_col, value=supplier_name)
                supplier_cell.alignment = wrap_alignment
                supplier_cell.fill = background_fill
                current_col += 1
            supplier_index += 1
        
        # >>>>> NEW STATISTICAL COLUMNS HEADERS (NOW AT END) <<<<<
        # Statistical columns headers (row 1)
        stats_headers = ["Statistics", "Statistics", "Statistics"]
        for header in stats_headers:
            stats_cell = ws.cell(row=1, column=current_col, value=header)
            stats_cell.alignment = wrap_alignment
            stats_cell.fill = yellow_fill
            current_col += 1
        
        # Second row headers
        current_col = 1
        for col_name in CONSOLIDATED_COMMON_COLUMN_NAMES:
            header_cell = ws.cell(row=2, column=current_col, value=col_name)
            header_cell.alignment = wrap_alignment
            current_col += 1
        
        valid_supplier_header = ws.cell(row=2, column=current_col, value="Valid Supplier")
        valid_supplier_header.alignment = wrap_alignment
        valid_supplier_header.fill = green_fill
        current_col += 1
        
        # Supplier columns sub-headers (moved before statistical columns)
        supplier_index = 0
        for file_name in file_names:
            supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
            background_fill = light_gray_fill if supplier_index % 2 == 0 else white_fill
            for col_name in supplier_columns:
                header_cell = ws.cell(row=2, column=current_col, value=col_name)
                header_cell.alignment = wrap_alignment
                header_cell.fill = background_fill
                current_col += 1
            supplier_index += 1
        
        # >>>>> NEW STATISTICAL COLUMNS SUB-HEADERS (NOW AT END) <<<<<
        # Statistical columns sub-headers (row 2)
        stats_sub_headers = ["Mean (Total Cost)", "Variance (Total Cost)", "Std Dev (Total Cost)"]
        for sub_header in stats_sub_headers:
            stats_sub_cell = ws.cell(row=2, column=current_col, value=sub_header)
            stats_sub_cell.alignment = wrap_alignment
            stats_sub_cell.fill = yellow_fill
            current_col += 1

        row_num = 3
        for row_key, row_data in consolidated_data.items():
            current_col = 1
            
            # Common data columns
            for col_name in CONSOLIDATED_COMMON_COLUMN_NAMES:
                data_cell = ws.cell(row=row_num, column=current_col, value=row_data['common_data'].get(col_name, ""))
                data_cell.alignment = Alignment(wrap_text=True, vertical='top')
                current_col += 1
            
            # Valid supplier count
            valid_supplier_count = 0
            for file_name in file_names:
                supplier_info = row_data['supplier_data'].get(file_name, {})
                if has_valid_supplier_data(supplier_info):
                    valid_supplier_count += 1
            valid_supplier_data_cell = ws.cell(row=row_num, column=current_col, value=valid_supplier_count)
            valid_supplier_data_cell.alignment = Alignment(wrap_text=True, vertical='top')
            if valid_supplier_count > 0:
                valid_supplier_data_cell.fill = green_fill
            current_col += 1
            
            # Fill in supplier data row cells (moved before statistical columns)
            sup_idx = 0
            for file_name in file_names:
                supplier_info = row_data['supplier_data'].get(file_name, {})
                supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
                for col_name in supplier_columns:
                    cell_value = supplier_info.get(col_name, "")
                    data_cell = ws.cell(row=row_num, column=current_col, value=cell_value)
                    data_cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    # >>>>> ROW-WISE OUTLIER COLORING <<<<<
                    if is_numeric_column(col_name) and col_name in OUTLIER_NUMERIC_COLUMNS:
                        # Check if this specific cell is marked as an outlier row-wise
                        is_outlier = (
                            row_key in rowwise_outlier_lookup and
                            sup_idx in rowwise_outlier_lookup[row_key] and
                            col_name in rowwise_outlier_lookup[row_key][sup_idx]
                        )
                        if is_outlier:
                            data_cell.fill = OUTLIER_FILL
                            print(f"Marked outlier: Row {row_num}, Supplier {sup_idx}, Column {col_name}, Value: {cell_value}")
                    
                    current_col += 1
                sup_idx += 1

            # >>>>> NEW STATISTICAL CALCULATIONS (NOW AT END) <<<<<
            # Calculate statistics for "Total Cost Per UOM FOB Port of Origin/Departure (USD)"
            target_column = "Total Cost Per UOM FOB Port of Origin/Departure (USD)"
            stats = calculate_row_statistics(row_data, file_names, suppliers_with_additional_info, target_column)
            
            # Mean column
            mean_cell = ws.cell(row=row_num, column=current_col, value=stats["mean"])
            mean_cell.alignment = Alignment(wrap_text=True, vertical='top')
            mean_cell.fill = yellow_fill
            current_col += 1
            
            # Variance column
            variance_cell = ws.cell(row=row_num, column=current_col, value=stats["variance"])
            variance_cell.alignment = Alignment(wrap_text=True, vertical='top')
            variance_cell.fill = yellow_fill
            current_col += 1
            
            # Standard deviation column
            std_dev_cell = ws.cell(row=row_num, column=current_col, value=stats["std_dev"])
            std_dev_cell.alignment = Alignment(wrap_text=True, vertical='top')
            std_dev_cell.fill = yellow_fill
            current_col += 1
            
            row_num += 1

        # Adjust column widths
        total_columns = len(CONSOLIDATED_COMMON_COLUMN_NAMES) + 1 + 3  # +3 for new statistical columns
        for file_name in file_names:
            supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
            total_columns += len(supplier_columns)
        
        for col_num in range(1, total_columns + 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 20
        
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 30
        for row_num in range(3, ws.max_row + 1):
            ws.row_dimensions[row_num].height = 30
        
        wb.save(OUTPUT_FILE)
        
    def process_cleaned_files(): 
        try: 
            cleaned_files = os.listdir(CLEANED_FILES_FOLDER_LOCATION)
            excel_files = [f for f in cleaned_files if f.endswith(('.xlsx', '.xls'))]
            if not excel_files:
                print("No Excel files found in the cleaned files folder!")
                return
            print(f"Found {len(excel_files)} Excel files to process")

            all_data = {}
            all_rows_data = []
            
            for item in excel_files:
                print("Processing cleaned file ---------------------------------------")
                file_path = os.path.join(CLEANED_FILES_FOLDER_LOCATION, item)
                print(f"Processing: {file_path}")
                try:
                    df = pd.read_excel(file_path)
                    file_name_key = os.path.splitext(item)[0]
                    all_data[file_name_key] = df
                    print(f"Successfully loaded {len(df)} rows from {item}")
                    print(f"Columns in file: {list(df.columns)}")
                except Exception as e:
                    print(f"Error processing {item}: {str(e)}")
                    continue
            if not all_data:
                print("No valid data found in any files!")
                return
            
            print("\nChecking for additional information data...")
            suppliers_with_additional_info = check_additional_info_data(all_data)
            
            print(f"Suppliers with additional info: {suppliers_with_additional_info}")
            print("\nCreating consolidated dataset...")
            consolidated_data = create_consolidated_dataset(all_data, suppliers_with_additional_info)
            
            print("Creating consolidated Excel file with row-wise outlier detection and statistical columns...")
            create_consolidated_excel(consolidated_data, list(all_data.keys()), suppliers_with_additional_info)
            
            print(f"\nConsolidated file created: {OUTPUT_FILE}")
            print(f"Suppliers including additional info column: {len(suppliers_with_additional_info)}")
            print(f"Suppliers excluding additional info column: {len(all_data) - len(suppliers_with_additional_info)}")
            print("Added statistical columns: Mean, Variance, and Standard Deviation for Total Cost Per UOM")
        except Exception as e: 
            print(f"Error in process_cleaned_files: {str(e)}")

    process_cleaned_files()

