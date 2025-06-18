import pandas as pd
import os 
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from handler.handler import extract_supplier_name

# All Folder location information 
FILES_FOLDER_LOCATION = "./files"
CLEANED_FILES_FOLDER_LOCATION = "./cleaned_files/bidsheet_other_metal"
CONSOLIDATED_FILE_LOCATION = "./consolidate"
CONSOLIDATE_FILE_NAME = "bidsheet_other_metal_consolidate"

# Column name related information 
CONSOLIDATED_COMMON_COLUMN_NAMES = ["ROW ID #", "Division", 'Part #', "Item Description "]
SUPPLIER_COLUMNS_NAMES = ["Price per UOM EXW (USD)", 
    "Freight Cost per UOM to Port of Origin/Departure (USD)", 
    "Total Cost Per UOM FOB Port of Origin/Departure (USD)",
    "Additional information (please use this column only if absolutely necessary)"
]

# Base supplier columns (without additional info)
BASE_SUPPLIER_COLUMNS = ["Price per UOM EXW (USD)", 
    "Freight Cost per UOM to Port of Origin/Departure (USD)", 
    "Total Cost Per UOM FOB Port of Origin/Departure (USD)"
]


# Numeric columns that should be formatted with 4 decimal places
NUMERIC_COLUMNS = [
    "Price per UOM EXW (USD)", 
    "Freight Cost per UOM to Port of Origin/Departure (USD)", 
    "Total Cost Per UOM FOB Port of Origin/Departure (USD)"
]


ADDITIONAL_INFO_COLUMN = "Additional information (please use this column only if absolutely necessary)"

# OUTPUT Files related information 
OUTPUT_FILE = f"{CONSOLIDATED_FILE_LOCATION}/{CONSOLIDATE_FILE_NAME}.xlsx"

def format_numeric_value(value):
    """
    Format numeric values to show up to 4 decimal places
    """
    if pd.isna(value) or value == "":
        return ""
    
    try:
        # Convert to float if it's not already
        if isinstance(value, str):
            # Remove any currency symbols or spaces
            cleaned_value = value.replace('$', '').replace(',', '').strip()
            if cleaned_value == '' or cleaned_value.lower() == 'nan':
                return ""
            numeric_value = float(cleaned_value)
        else:
            numeric_value = float(value)
        
        # Format to 4 decimal places, but remove trailing zeros
        formatted = f"{numeric_value:.4f}"
        # Remove trailing zeros and decimal point if not needed
        formatted = formatted.rstrip('0').rstrip('.')
        return formatted
        
    except (ValueError, TypeError):
        # If conversion fails, return the original value
        return str(value) if value is not None else ""

def is_numeric_column(column_name):
    """
    Check if a column contains numeric data that should be formatted
    """
    return column_name in NUMERIC_COLUMNS

def process_excel_files():
    """
    Process all Excel files in the FILES_FOLDER_LOCATION and extract data from the specified sheet
    """
    files = os.listdir(FILES_FOLDER_LOCATION)
    
    for item in files: 
        print("Processing excel sheet ---------------------------------------")
        print(f"./files/{item}")

        # Cleaned csv file name 
        cleaned_csv_file_name = f"{item.split('.')[0]}_cleaned.xlsx"

        # Make sure folder exists or not 
        os.makedirs(CLEANED_FILES_FOLDER_LOCATION, exist_ok=True)

        try:
            # Read Excel file
            df = pd.read_excel(f"./files/{item}", "3. Bidsheet Other Metals", header=None)

            # Find the row containing "ROW ID #"
            header_row_index = None
            col_start_index = None

            for index, row in df.iterrows():
                for col_index, value in enumerate(row):
                    if isinstance(value, str) and "ROW ID #" in value:
                        header_row_index = index
                        col_start_index = col_index
                        break
                if header_row_index is not None:
                    break

            if header_row_index is not None and col_start_index is not None:
                # Set new headers starting from the found column
                headers = df.iloc[header_row_index, col_start_index:].tolist()
                print("Headers information ---------------------------------------")
                print(headers)

                # Extract data rows below the header, also starting from same column
                data_rows = df.iloc[header_row_index + 1:, col_start_index:]
                
                # Set proper headers
                data_rows.columns = headers
                data_rows.reset_index(drop=True, inplace=True)
                
                data_rows.to_excel(f"{CLEANED_FILES_FOLDER_LOCATION}/{cleaned_csv_file_name}", index=False)
                print(f"Data saved to '{CLEANED_FILES_FOLDER_LOCATION}/{cleaned_csv_file_name}'")
            else:
                print("No row containing 'ROW ID #' was found.")
                
        except Exception as e:
            print(f"Error processing {item}: {str(e)}")
            continue

def check_additional_info_data(all_data):
    """
    Check which suppliers have actual data in the additional information column
    Enhanced version with better validation
    """
    suppliers_with_additional_info = set()
    
    for file_name, df in all_data.items():
        print(f"Checking additional info for supplier: {file_name}")
        
        if ADDITIONAL_INFO_COLUMN in df.columns:
            # Check if any row has non-empty additional info
            additional_info_series = df[ADDITIONAL_INFO_COLUMN]
            
            # More robust checking for actual data
            has_data = (
                additional_info_series.notna() & 
                (additional_info_series.astype(str).str.strip() != "") & 
                (additional_info_series.astype(str).str.strip() != "nan") &
                (additional_info_series.astype(str).str.lower().str.strip() != "none")
            )
            
            if has_data.any():
                suppliers_with_additional_info.add(file_name)
                print(f"✓ Supplier {file_name} has additional information data")
                # Show sample of additional info data
                sample_data = additional_info_series[has_data].head(3).tolist()
                print(f"  Sample data: {sample_data}")
            else:
                print(f"✗ Supplier {file_name} has NO additional information data")
        else:
            print(f"✗ Supplier {file_name} doesn't have additional information column")
    
    return suppliers_with_additional_info

def is_valid_supplier_value(value):
    """
    Check if a value is considered valid for supplier validation
    (non-blank and non-zero)
    """
    if pd.isna(value) or value == "":
        return False
    
    # Convert to string and check if it's empty or just whitespace
    str_value = str(value).strip()
    if str_value == "" or str_value.lower() == "nan":
        return False
    
    # Try to convert to float to check if it's zero
    try:
        float_value = float(str_value.replace('$', '').replace(',', ''))
        return float_value != 0
    except (ValueError, TypeError):
        # If it's not a number but has content, consider it valid
        return True

def count_valid_suppliers(row_data, file_names, suppliers_with_additional_info):
    """
    Count how many suppliers have valid (non-blank, non-zero) data for a row
    """
    valid_count = 0
    
    for file_name in file_names:
        supplier_info = row_data['supplier_data'].get(file_name, {})
        supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
        
        # Check if any column for this supplier has valid data
        has_valid_data = False
        for col_name in supplier_columns:
            value = supplier_info.get(col_name, "")
            if is_valid_supplier_value(value):
                has_valid_data = True
                break
        
        if has_valid_data:
            valid_count += 1
    
    return valid_count

def get_supplier_columns(file_name, suppliers_with_additional_info):
    """
    Get the appropriate columns for each supplier based on whether they have additional info data
    """
    if file_name in suppliers_with_additional_info:
        print(f"Including additional info column for {file_name}")
        return SUPPLIER_COLUMNS_NAMES  # Include all columns including additional info
    else:
        print(f"Excluding additional info column for {file_name}")
        return BASE_SUPPLIER_COLUMNS  # Exclude additional info column

def create_consolidated_dataset(all_data, suppliers_with_additional_info):
    """
    Create a consolidated dataset from all files
    """
    # Get all unique rows based on common columns
    consolidated_rows = {}
    
    for file_name, df in all_data.items():
        print(f"Processing data from {file_name}...")
        
        # Get appropriate columns for this supplier
        supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
        
        for _, row in df.iterrows():
            # Create a unique key for each row based on common columns
            # Handle missing columns gracefully
            row_key_parts = []
            common_data = {}
            
            for col in CONSOLIDATED_COMMON_COLUMN_NAMES:
                if col in df.columns:
                    value = str(row[col]) if pd.notna(row[col]) else ""
                    row_key_parts.append(value)
                    common_data[col] = row[col] if pd.notna(row[col]) else ""
                else:
                    row_key_parts.append("")
                    common_data[col] = ""
            
            row_key = "|".join(row_key_parts)
            
            if row_key not in consolidated_rows:
                consolidated_rows[row_key] = {
                    'common_data': common_data,
                    'supplier_data': {}
                }
            
            # Add supplier-specific data using only relevant columns
            supplier_data = {}
            for col in supplier_columns:
                if col in df.columns:
                    value = row[col] if pd.notna(row[col]) else ""
                    # Format numeric columns with 4 decimal places
                    if is_numeric_column(col):
                        value = format_numeric_value(value)
                    supplier_data[col] = value
                else:
                    supplier_data[col] = ""
            
            consolidated_rows[row_key]['supplier_data'][file_name] = supplier_data
    
    return consolidated_rows

def has_valid_supplier_data(supplier_data):
    """
    Check if a supplier has valid (non-blank, non-zero) data in any of their columns
    """
    for value in supplier_data.values():
        if value != "" and value != 0 and value != "0" and value != "0.0000":
            return True
    return False

def create_consolidated_excel(consolidated_data, file_names, suppliers_with_additional_info):
    """
    Create Excel file with two-row header structure - supplier names in row 1, column names in row 2
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated Suppliers"
    
    # Define styles
    light_gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # Light gray
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # White
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
    wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Create header structure - Two rows
    current_col = 1
    supplier_index = 0  # Track supplier index for alternating colors
    
    # Row 1: Supplier names (empty for common columns)
    # Add empty cells for common columns in row 1
    for col_name in CONSOLIDATED_COMMON_COLUMN_NAMES:
        header_cell = ws.cell(row=1, column=current_col, value="")
        header_cell.alignment = wrap_alignment
        current_col += 1
    
    # Add "Valid Supplier" column header in row 1 (empty) with green background
    valid_supplier_cell = ws.cell(row=1, column=current_col, value="")
    valid_supplier_cell.alignment = wrap_alignment
    valid_supplier_cell.fill = green_fill  # Green background for Valid Supplier column
    current_col += 1
    
    # Add supplier names for each of their columns
    for file_name in file_names:
        print("Supplier name related information------------------------------------------")
        supplier_user_information = extract_supplier_name(file_name)
        supplier_name = supplier_user_information[1]
        supplier_name = supplier_name.replace("_cleaned", "")
        print(supplier_name)
        
        # Get appropriate columns for this supplier
        supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
        
        # Determine background color based on supplier index (alternating)
        if supplier_index % 2 == 0:
            background_fill = light_gray_fill  # Even index = light gray
        else:
            background_fill = white_fill  # Odd index = white
        
        # Add supplier name for each of their columns
        for col_name in supplier_columns:
            supplier_cell = ws.cell(row=1, column=current_col, value=supplier_name)
            supplier_cell.alignment = wrap_alignment
            supplier_cell.fill = background_fill  # Apply alternating background color
            current_col += 1
            
        supplier_index += 1  # Increment supplier index for next iteration
    
    # Row 2: Column names
    current_col = 1
    
    # Add common column headers in row 2
    for col_name in CONSOLIDATED_COMMON_COLUMN_NAMES:
        header_cell = ws.cell(row=2, column=current_col, value=col_name)
        header_cell.alignment = wrap_alignment
        current_col += 1
    
    # Add "Valid Supplier" column header in row 2 with green background
    valid_supplier_header = ws.cell(row=2, column=current_col, value="Valid Supplier")
    valid_supplier_header.alignment = wrap_alignment
    valid_supplier_header.fill = green_fill  # Green background for Valid Supplier column
    current_col += 1
    
    # Add supplier column headers in row 2
    supplier_index = 0
    for file_name in file_names:
        supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
        
        # Determine background color based on supplier index (alternating)
        if supplier_index % 2 == 0:
            background_fill = light_gray_fill  # Even index = light gray
        else:
            background_fill = white_fill  # Odd index = white
        
        for col_name in supplier_columns:
            header_cell = ws.cell(row=2, column=current_col, value=col_name)
            header_cell.alignment = wrap_alignment
            header_cell.fill = background_fill  # Apply same background color as supplier name
            current_col += 1
            
        supplier_index += 1
    
    # Add data rows
    row_num = 3  # Start from row 3 since we have two header rows
    for row_key, row_data in consolidated_data.items():
        current_col = 1
        
        # Add common data
        for col_name in CONSOLIDATED_COMMON_COLUMN_NAMES:
            data_cell = ws.cell(row=row_num, column=current_col, 
                               value=row_data['common_data'].get(col_name, ""))
            data_cell.alignment = Alignment(wrap_text=True, vertical='top')
            current_col += 1
        
        # Calculate and add Valid Supplier count with conditional green highlighting
        valid_supplier_count = 0
        for file_name in file_names:
            supplier_info = row_data['supplier_data'].get(file_name, {})
            if has_valid_supplier_data(supplier_info):
                valid_supplier_count += 1
        
        valid_supplier_data_cell = ws.cell(row=row_num, column=current_col, value=valid_supplier_count)
        valid_supplier_data_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Highlight cell with green if valid_supplier_count > 0
        if valid_supplier_count > 0:
            valid_supplier_data_cell.fill = green_fill
        
        current_col += 1
        
        # Add supplier data
        for file_name in file_names:
            supplier_info = row_data['supplier_data'].get(file_name, {})
            supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
            for col_name in supplier_columns:
                cell_value = supplier_info.get(col_name, "")
                data_cell = ws.cell(row=row_num, column=current_col, value=cell_value)
                data_cell.alignment = Alignment(wrap_text=True, vertical='top')
                current_col += 1
        
        row_num += 1
    
    # Calculate total columns dynamically (including the new Valid Supplier column)
    total_columns = len(CONSOLIDATED_COMMON_COLUMN_NAMES) + 1  # +1 for Valid Supplier column
    for file_name in file_names:
        supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
        total_columns += len(supplier_columns)
    
    # Set fixed column widths and wrap text
    for col_num in range(1, total_columns + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20  # Standard width
    
    # Set row heights to accommodate wrapped text
    ws.row_dimensions[1].height = 30  # Supplier name row
    ws.row_dimensions[2].height = 30  # Column name row
    for row_num in range(3, ws.max_row + 1):
        ws.row_dimensions[row_num].height = 30
    
    # Save the workbook
    wb.save(OUTPUT_FILE)
  
def process_cleaned_files(): 
    """
    Function information: 
        Apply for all process cleaned files with enhanced additional info logic
    """
    try: 
        # Reading cleaned files related csv files 
        cleaned_files = os.listdir(CLEANED_FILES_FOLDER_LOCATION)
        excel_files = [f for f in cleaned_files if f.endswith(('.xlsx', '.xls'))]

        if not excel_files:
            print("No Excel files found in the cleaned files folder!")
            return
    
        print(f"Found {len(excel_files)} Excel files to process")

        # Dictionary to store all data from files
        all_data = {}

        # Process each cleaned file
        for item in excel_files:
            print("Processing cleaned file ---------------------------------------")
            file_path = os.path.join(CLEANED_FILES_FOLDER_LOCATION, item)
            print(f"Processing: {file_path}")
            
            try:
                # Read the Excel file
                df = pd.read_excel(file_path)
                
                # Store the dataframe for later use
                file_name_key = os.path.splitext(item)[0]  # Remove extension
                all_data[file_name_key] = df
                
                print(f"Successfully loaded {len(df)} rows from {item}")
                print(f"Columns in file: {list(df.columns)}")
                
            except Exception as e:
                print(f"Error processing {item}: {str(e)}")
                continue
        
        if not all_data:
            print("No valid data found in any files!")
            return
        
        # Check which suppliers have additional information data
        print("\n" + "="*60)
        print("CHECKING FOR ADDITIONAL INFORMATION DATA")
        print("="*60)
        suppliers_with_additional_info = check_additional_info_data(all_data)
        print(f"\nSummary:")
        print(f"Suppliers with additional info: {suppliers_with_additional_info}")
        print(f"Total suppliers with additional info: {len(suppliers_with_additional_info)}")
        print(f"Total suppliers without additional info: {len(all_data) - len(suppliers_with_additional_info)}")
        
        # Create consolidated dataset
        print("\n" + "="*60)
        print("CREATING CONSOLIDATED DATASET")
        print("="*60)
        consolidated_data = create_consolidated_dataset(all_data, suppliers_with_additional_info)
        
        # Create Excel file with proper headers
        print("\n" + "="*60)
        print("CREATING CONSOLIDATED EXCEL FILE")
        print("="*60)
        
        # Ensure consolidate directory exists
        os.makedirs(CONSOLIDATED_FILE_LOCATION, exist_ok=True)
        
        create_consolidated_excel(consolidated_data, list(all_data.keys()), suppliers_with_additional_info)
        
        print(f"\n" + "="*60)
        print("CONSOLIDATION COMPLETE")
        print("="*60)
        print(f"Consolidated file created: {OUTPUT_FILE}")
        print(f"Total unique rows processed: {len(consolidated_data)}")
        print(f"Suppliers including additional info column: {len(suppliers_with_additional_info)}")
        print(f"Suppliers excluding additional info column: {len(all_data) - len(suppliers_with_additional_info)}")
        
        # List suppliers with and without additional info
        if suppliers_with_additional_info:
            print(f"\nSuppliers WITH additional info:")
            for supplier in suppliers_with_additional_info:
                print(f"  - {supplier}")
        
        suppliers_without_info = set(all_data.keys()) - suppliers_with_additional_info
        if suppliers_without_info:
            print(f"\nSuppliers WITHOUT additional info:")
            for supplier in suppliers_without_info:
                print(f"  - {supplier}")
    
    except Exception as e: 
        print(f"Error in process_cleaned_files: {str(e)}")
        import traceback
        traceback.print_exc()

def main():
    """
    Main function to run the entire process
    """
    print("Starting Bidsheet Other Metal Processing...")
    print("="*60)
    
    # Step 1: Process raw Excel files
    print("STEP 1: Processing raw Excel files...")
    process_excel_files()
    
    # Step 2: Process cleaned files and create consolidated output
    print("\nSTEP 2: Processing cleaned files and creating consolidated output...")
    process_cleaned_files()
    
    print("\nProcess completed successfully!")

if __name__ == "__main__":
    main()