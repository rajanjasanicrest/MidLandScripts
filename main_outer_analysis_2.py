import pandas as pd
import os 
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from handler.handler import extract_supplier_name
import numpy as np
from scipy import stats

# --------- All files related information --------------- # 
# --------- Contains file information ------------------- # 
# 1. Bidsheet Brass 
# 2. Bidsheet Steel
# 3. Bidsheet Other Metals

ALL_FILES = [
    {
        "consolidation_folder": "bidsheet_other_metal",
        "consolidate_file_name": "bidsheet_other_metal_outlier_consolidate",
        "sheet_name": "3. Bidsheet Other Metals",
    },
    {
        "consolidation_folder": "bidsheet_steel", 
        "consolidate_file_name": "bidsheet_steel_outlier_consolidate", 
        "sheet_name": "2. Bidsheet Steel",
    }, 
    {
        "consolidation_folder": "bidsheet_brass",
        "consolidate_file_name": "bidsheet_brass_outlier_consolidate",  
        "sheet_name": "1. Bidsheet Brass",
    }, 
]

for main_files_name_information in ALL_FILES:

    # All Folder location information 
    FILES_FOLDER_LOCATION = "./files"
    R2_FILES_FOLDER_LOCATION = "./files round 2"
    CLEANED_FILES_FOLDER_LOCATION = f"./cleaned_files/{main_files_name_information['consolidation_folder']}"
    CONSOLIDATED_FILE_LOCATION = "./consolidate"
    CONSOLIDATE_FILE_NAME = main_files_name_information["consolidate_file_name"]

    # Color for outliers (Blue, in hex: #87CEEB)
    OUTLIER_FILL = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')

    # Column name related information 
    CONSOLIDATED_COMMON_COLUMN_NAMES = ["ROW ID #", "Division", 'Part #', "Item Description", "Product Group", "Part Family", "Average Order Quantity (per UOM)","Min Order Quantity (per UOM)","Max Order Quantity (per UOM)", "Order frequency",	"Annual Volume (per UOM)" ]

    ROW_KEY_COLUMNS = ["ROW ID #", "Division", 'Part #', "Item Description"]

    SUPPLIER_COLUMNS_NAMES = [
        "Total Cost Per UOM FOB Port of Origin/Departure (USD)",
    ]
    BASE_SUPPLIER_COLUMNS = [
        "Total Cost Per UOM FOB Port of Origin/Departure (USD)"
    ]
    ADDITIONAL_INFO_COLUMN = "DONOTCONSIDER"

    NUMERIC_COLUMNS = [
        "Price per UOM EXW (USD)", 
        "Freight Cost per UOM to Port of Origin/Departure (USD)", 
        "Total Cost Per UOM FOB Port of Origin/Departure (USD)"
    ]

    OUTPUT_FILE = f"{CONSOLIDATED_FILE_LOCATION}/{CONSOLIDATE_FILE_NAME}.xlsx"
    OUTLIER_NUMERIC_COLUMNS = ["Total Cost Per UOM FOB Port of Origin/Departure (USD)"]

    # ------------------- Helper functions --------------------- #
    # Convert numeric value to four decimal places and remove trailing zeros
    
    def format_numeric_value(value):
        """
            Convert a value to four decimal places and remove trailing zeros.
        """
        if pd.isna(value) or value == "":
            return ""
        try:
            if isinstance(value, str):
                cleaned_value = value.replace('$', '').replace(',', '').strip()
                if cleaned_value == '' or cleaned_value.lower() == 'nan':
                    return ""
                numeric_value = float(cleaned_value)
            else:
                numeric_value = float(value)
            formatted = f"{numeric_value:.4f}"
            formatted = formatted.rstrip('0').rstrip('.')
            return formatted
        except (ValueError, TypeError):
            return str(value) if value is not None else ""

    def is_numeric_column(column_name):
        return column_name in NUMERIC_COLUMNS

    # files = os.listdir(FILES_FOLDER_LOCATION)
    # for item in files: 
        
    #     print("Processing excel sheet ---------------------------------------")
    #     print(f"./files/{item}")

    #     cleaned_csv_file_name = f"{item.split('.')[0]}_cleaned.xlsx"
    #     os.makedirs(CLEANED_FILES_FOLDER_LOCATION, exist_ok=True)
    #     df = pd.read_excel(f"./files/{item}", main_files_name_information["sheet_name"], header=None)
    #     header_row_index = None
    #     col_start_index = None
    #     for index, row in df.iterrows():
    #         for col_index, value in enumerate(row):
    #             if isinstance(value, str) and "ROW ID #" in value:
    #                 header_row_index = index
    #                 col_start_index = col_index
    #                 break
    #         if header_row_index is not None:
    #             break
    #     if header_row_index is not None and col_start_index is not None:
    #         headers = df.iloc[header_row_index, col_start_index:].tolist()
            
    #         data_rows = df.iloc[header_row_index + 1:, col_start_index:]
    #         data_rows.columns = headers
    #         data_rows.reset_index(drop=True, inplace=True)
    #         data_rows.to_excel(f"{CLEANED_FILES_FOLDER_LOCATION}/{cleaned_csv_file_name}", index=False)
    #         print(f"Data saved to '{CLEANED_FILES_FOLDER_LOCATION}/{cleaned_csv_file_name}'")
    #     else:
    #         print("No row containing 'ROW ID #' was found.")

    # round2files = os.listdir(R2_FILES_FOLDER_LOCATION)
    # for item in round2files:
    #     print("Processing Round 2 excel sheets ---------------------------------------")
    #     print(f"./files/{item}")

    #     cleaned_csv_file_name = f"{item.split('.')[0]}_r2_cleaned.xlsx"
    #     os.makedirs(CLEANED_FILES_FOLDER_LOCATION, exist_ok=True)
    #     df = pd.read_excel(f"./files round 2/{item}", main_files_name_information["sheet_name"], header=None)
    #     header_row_index = None
    #     col_start_index = None
    #     for index, row in df.iterrows():
    #         for col_index, value in enumerate(row):
    #             if isinstance(value, str) and "ROW ID #" in value:
    #                 header_row_index = index
    #                 col_start_index = col_index
    #                 break
    #         if header_row_index is not None:
    #             break
    #     if header_row_index is not None and col_start_index is not None:
    #         headers = df.iloc[header_row_index, col_start_index:].tolist()
            
    #         data_rows = df.iloc[header_row_index + 1:, col_start_index:]
    #         data_rows.columns = headers
    #         data_rows.reset_index(drop=True, inplace=True)
    #         data_rows.to_excel(f"{CLEANED_FILES_FOLDER_LOCATION}/{cleaned_csv_file_name}", index=False)
    #         print(f"Data saved to '{CLEANED_FILES_FOLDER_LOCATION}/{cleaned_csv_file_name}'")
    #     else:
    #         print("No row containing 'ROW ID #' was found.")

    def check_additional_info_data(all_data):
        suppliers_with_additional_info = set()
        for file_name, df in all_data.items():
            if ADDITIONAL_INFO_COLUMN in df.columns:
                additional_info_series = df[ADDITIONAL_INFO_COLUMN]
                has_data = additional_info_series.notna() & (additional_info_series.astype(str).str.strip() != "") & (additional_info_series.astype(str).str.strip() != "nan")
                if has_data.any():
                    suppliers_with_additional_info.add(file_name)
                    print(f"Supplier {file_name} has additional information data")
                else:
                    print(f"Supplier {file_name} has NO additional information data")
            else:
                print(f"Supplier {file_name} doesn't have additional information column")
        return suppliers_with_additional_info

    def get_supplier_columns(file_name, suppliers_with_additional_info):
        if file_name in suppliers_with_additional_info:
            return SUPPLIER_COLUMNS_NAMES
        else:
            return BASE_SUPPLIER_COLUMNS

    def is_valid_supplier_value(value):
        if pd.isna(value) or value == "":
            return False
        str_value = str(value).strip()
        if str_value == "" or str_value.lower() == "nan":
            return False
        try:
            float_value = float(str_value.replace('$', '').replace(',', ''))
            return float_value != 0
        except (ValueError, TypeError):
            return True

    def count_valid_suppliers(row_data, file_names, suppliers_with_additional_info):
        valid_count = 0
        for file_name in file_names:
            supplier_info = row_data['supplier_data'].get(file_name, {})
            supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
            supplier_columns = [s for s in supplier_columns if 'R2' in s]
            has_valid_data = False
            for col_name in supplier_columns:
                value = supplier_info.get(col_name, "")
                if is_valid_supplier_value(value):
                    has_valid_data = True
                    break
            if has_valid_data:
                valid_count += 1
        return valid_count

    from collections import defaultdict

    def create_consolidated_dataset(all_data, suppliers_with_additional_info, display_names):
        consolidated_rows = {}

        # Track which supplier has R1 and/or R2
        supplier_rounds_map = defaultdict(set)
        supplier_columns_map = {}  # base_name -> column list from either R1 or R2

        for file_key, df in all_data.items():
            print(f"Processing data from {file_key}...")

            df.columns = [str(col).strip() for col in df.columns]

            is_r2 = "_r2" in file_key.lower()
            round_label = "R2" if is_r2 else "R1"

            base_name = (
                file_key.split('--')[-1]
                .replace('R2_r2_cleaned', '')
                .replace('_cleaned', '')
                .strip()
            )

            pretty_supplier_key = f"{base_name} - {round_label}"

            supplier_rounds_map[base_name].add(round_label)

            supplier_columns = get_supplier_columns(file_key, suppliers_with_additional_info)
            supplier_columns_map[base_name] = supplier_columns  # latest seen is fine

            for _, row in df.iterrows():
                row_key_parts = []
                common_data = {}
                for col in ROW_KEY_COLUMNS:
                    if col in df.columns:
                        value = row[col]
                        # Standardize: convert to string, strip whitespace, handle NaN/missing
                        if pd.isna(value) or str(value).strip().lower() in ["", "nan", "none"]:
                            clean_value = ""
                        else:
                            clean_value = str(value).strip()
                        row_key_parts.append(clean_value)
                        common_data[col] = clean_value
                    else:
                        row_key_parts.append("")
                        common_data[col] = ""
                # Add the rest of CONSOLIDATED_COMMON_COLUMN_NAMES to common_data for Excel output
                for col in CONSOLIDATED_COMMON_COLUMN_NAMES:
                    if col not in ROW_KEY_COLUMNS:
                        if col in df.columns:
                            value = row[col]
                            if pd.isna(value) or str(value).strip().lower() in ["", "nan", "none"]:
                                clean_value = ""
                            else:
                                clean_value = str(value).strip()
                            common_data[col] = clean_value
                        else:
                            common_data[col] = ""
                row_key = "|".join(row_key_parts)

                if row_key not in consolidated_rows:
                    consolidated_rows[row_key] = {
                        'common_data': common_data,
                        'supplier_data': {}
                    }

                supplier_data = {}
                for col in supplier_columns:
                    full_col_name = f"{pretty_supplier_key} - {col}"
                    if col in df.columns:
                        if pd.notna(row[col]):
                            val = row[col]
                            value = format_numeric_value(val) if is_numeric_column(col) else str(val)
                        else:
                            value = ""
                        supplier_data[full_col_name] = value
                    else:
                        supplier_data[full_col_name] = 0.0

                consolidated_rows[row_key]['supplier_data'][pretty_supplier_key] = supplier_data

        # 🔧 Post-process: Add empty R1 or R2 if missing
        for row in consolidated_rows.values():
            for base_name, rounds in supplier_rounds_map.items():
                all_rounds = {"R1", "R2"}
                missing_rounds = all_rounds - rounds

                for missing_round in missing_rounds:
                    pretty_supplier_key = f"{base_name} - {missing_round}"
                    if pretty_supplier_key not in row['supplier_data']:
                        empty_data = {}
                        for col in supplier_columns_map.get(base_name, []):
                            full_col_name = f"{pretty_supplier_key} - {col}"
                            empty_data[full_col_name] = ""
                        row['supplier_data'][pretty_supplier_key] = empty_data

        return consolidated_rows


    def has_valid_supplier_data(supplier_data):
        for value in supplier_data.values():
            # If it's an int or float, check directly
            if isinstance(value, (int, float)) and value != 0:
                return True
           
            # If it's a string, try converting to float
            if isinstance(value, str):
                try:
                    numeric_value = float(value)
                    if numeric_value != 0:
                        return True
                except ValueError:
                    # Not a numeric string, skip
                    continue
 
        return False

    # --------------------- Row wise outlier detection ---------------- #

    def is_valid_nonzero(value):
        """Check if value is valid and non-zero for outlier analysis"""
        try:
            if pd.isna(value):
                return False
            str_val = str(value).strip()
            if str_val == "" or str_val.lower() == "nan":
                return False
            float_val = float(str_val.replace("$","").replace(",",""))
            return float_val != 0  # Exclude 0 values as requested
        except Exception:
            return False

    def grubbs_test(values, alpha=0.05):
        """
        Return indices of outliers in the values using Grubbs' test (two-sided).
        Iteratively remove outliers until no more are found (optional).

        :param values: List or numpy array of values
        :param alpha: Significance level for Grubbs' test
        :return: Set of indices that are outliers
        """
        x = np.array(values)
        print(x)
        outlier_indices = set()
        original_indices = list(range(len(x)))

        while len(x) > 2:
            mean = np.mean(x)
            std = np.std(x, ddof=1)
            if std == 0:
                break

            abs_diffs = np.abs(x - mean)
            max_idx = np.argmax(abs_diffs)
            G = abs_diffs[max_idx] / std

            n = len(x)
            # Critical value for Grubbs' test
            t = stats.t.ppf(1 - alpha / (2 * n), n - 2)
            G_crit = ((n - 1) / np.sqrt(n)) * np.sqrt(t**2 / (n - 2 + t**2))

            if G > G_crit:
                # Remove outlier
                outlier_indices.add(original_indices[max_idx])
                # Remove from arrays for potential second outlier detection
                x = np.delete(x, max_idx)
                original_indices.pop(max_idx)
            else:
                break

        return outlier_indices
            
    def detect_outliers_rowwise(values, method='auto'):
        """
        Given a list of numbers for a single row, return indices that are outliers.
        Uses Z-score method for small datasets (typical for supplier comparison within a row).
        """
        if len(values) < 0:
            return set()
        
        values = np.array(values, dtype='float')
        n = len(values)
        print(f"Row-wise outlier detection - Values for this row: {values}")
        
        outlier_idxs = set()
        
        # For row-wise analysis, we typically have few suppliers (2-10), so use Z-score method
        mean = np.mean(values)
        std = np.std(values, ddof=0)
        
        if std == 0:  # All values are the same
            return set()
        
        # Use 2 standard deviations as threshold (can be adjusted as needed)
        z_scores = np.abs(values - mean) / std
        outlier_idxs = set(np.where(z_scores > 2)[0])
        
        print(f"Mean: {mean:.4f}, Std: {std:.4f}")
        print(f"Z-scores: {z_scores}")
        print(f"Found {len(outlier_idxs)} outliers in this row: indices {outlier_idxs}")
        
        return outlier_idxs

    def build_rowwise_outlier_lookup(consolidated_data, file_names, suppliers_with_additional_info):
        """
        Build a row-wise outlier lookup that maps (row_key, supplier_idx, col_name) to outlier status
        """
        print("Building row-wise outlier lookup...")
        
        outlier_lookup = {}
        
        for row_key, row_data in consolidated_data.items():
            # print(f"\nProcessing row: {row_key[:50]}...")  # Show first 50 chars of row key
            
            for col_name in OUTLIER_NUMERIC_COLUMNS:
                # Collect values for this specific row and column across all suppliers
                row_values = []
                supplier_indices = []
                
                for supplier_idx, file_name in enumerate(file_names):
                    supplier_info = row_data['supplier_data'].get(file_name, {})
                    supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
                    
                    if col_name not in supplier_columns:
                        continue
                        
                    value = supplier_info.get(col_name, "")
                    if is_valid_nonzero(value):
                        try:
                            val = float(str(value).replace("$", "").replace(",", ""))
                            row_values.append(val)
                            supplier_indices.append(supplier_idx)
                        except Exception:
                            continue
                
                # Only perform outlier detection if we have at least 2 valid values
                if len(row_values) >= 2:
                    # print(f"Column {col_name}: {len(row_values)} valid values")
                    
                    # Detect outliers for this row
                    outlier_indices = grubbs_test(row_values, alpha=0.05)
                    
                    # Map outlier indices back to supplier indices
                    for outlier_idx in outlier_indices:
                        if outlier_idx < len(supplier_indices):
                            supplier_idx = supplier_indices[outlier_idx]
                            
                            # Initialize nested structure if needed
                            if row_key not in outlier_lookup:
                                outlier_lookup[row_key] = {}
                            if supplier_idx not in outlier_lookup[row_key]:
                                outlier_lookup[row_key][supplier_idx] = {}
                            
                            # Mark as outlier
                            outlier_lookup[row_key][supplier_idx][col_name] = True
                            print(f"Marked outlier: Supplier {supplier_idx}, Value: {row_values[outlier_idx]}")
                else:
                    pass
                    # print(f"Column {col_name}: Only {len(row_values)} valid values, skipping outlier detection")
        
        return outlier_lookup

    #  ------------------ New statistics calculation ------------------- #

    def calculate_row_statistics(row_data, file_names, suppliers_with_additional_info, target_column):
        """
        Calculate mean, variance, and standard deviation for a specific column across all suppliers in a row
        """
        values = []
        
        for file_name in file_names:
            supplier_info = row_data['supplier_data'].get(file_name, {})
            supplier_columns = get_supplier_columns(file_name, suppliers_with_additional_info)
            
            if target_column in supplier_columns:
                value = supplier_info.get(target_column, "")
                if is_valid_nonzero(value):
                    try:
                        val = float(str(value).replace("$", "").replace(",", ""))
                        values.append(val)
                    except Exception:
                        continue
        
        if len(values) == 0:
            return {"mean": "", "variance": "", "std_dev": ""}
        elif len(values) == 1:
            return {"mean": values[0], "variance": 0, "std_dev": 0}
        else:
            mean_val = np.mean(values)
            variance_val = np.var(values, ddof=1)  # Sample variance
            std_dev_val = np.std(values, ddof=1)   # Sample standard deviation
            
            return {
                "mean": round(mean_val, 4),
                "variance": round(variance_val, 4),
                "std_dev": round(std_dev_val, 4)
            }

    def create_consolidated_excel(consolidated_data, file_keys, suppliers_with_additional_info):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Consolidated Suppliers"

        wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        light_gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

        # 🧱 Step 1: Gather all supplier full column names
        all_supplier_columns = []

        for row_data in consolidated_data.values():
            for supplier_key, supplier_data in row_data['supplier_data'].items():
                for col in supplier_data:
                    if col not in all_supplier_columns:
                        all_supplier_columns.append(col)
            break  # Just check one row for keys (they're consistent)

        # 🧱 Step 2: Write headers (2 rows)
        current_col = 1

        # Row 1: Common column headers
        for col_name in CONSOLIDATED_COMMON_COLUMN_NAMES:
            ws.cell(row=1, column=current_col, value="").alignment = wrap_alignment
            ws.cell(row=2, column=current_col, value=col_name).alignment = wrap_alignment
            current_col += 1

        # Valid supplier count column
        ws.cell(row=1, column=current_col, value="").alignment = wrap_alignment
        ws.cell(row=2, column=current_col, value="Valid Supplier").alignment = wrap_alignment
        ws.cell(row=2, column=current_col).fill = green_fill
        current_col += 1

        # Row 1 & 2: Supplier-specific headers
        supplier_grouped = {}

        from collections import defaultdict
        # Group by base supplier, then split into R1 and R2
        supplier_r1_r2_map = defaultdict(lambda: {"R1": [], "R2": []})

        for col in all_supplier_columns:
            parts = col.split(" - ")
            if len(parts) >= 3:
                base_supplier = parts[0].strip()
                round_label = parts[1].strip()
                if round_label in ["R1", "R2"]:
                    supplier_r1_r2_map[base_supplier][round_label].append(col)

        # Now sort supplier names alphabetically and flatten with R1 + R2 adjacent
        supplier_grouped = {}
        sorted_supplier_names = sorted(supplier_r1_r2_map.keys(), key=lambda x: x.lower())

        for idx, supplier in enumerate(sorted_supplier_names):
            r1_cols = supplier_r1_r2_map[supplier].get("R1", [])
            r2_cols = supplier_r1_r2_map[supplier].get("R2", [])
            if r1_cols:
                supplier_grouped[f"{supplier} - R1"] = r1_cols
            if r2_cols:
                supplier_grouped[f"{supplier} - R2"] = r2_cols


        for idx, (supplier_key, cols) in enumerate(supplier_grouped.items()):
            background_fill = light_gray_fill if idx % 2 == 0 else white_fill
            for col in cols:
                ws.cell(row=1, column=current_col, value=supplier_key).alignment = wrap_alignment
                ws.cell(row=1, column=current_col).fill = background_fill

                ws.cell(row=2, column=current_col, value=" - ".join(col.split(" - ")[2:])).alignment = wrap_alignment
                ws.cell(row=2, column=current_col).fill = background_fill

                current_col += 1

        # 🧱 Step 3: Fill data rows
        row_num = 3
        for row_key, row_data in consolidated_data.items():
            current_col = 1

            # Write common columns
            for col_name in CONSOLIDATED_COMMON_COLUMN_NAMES:
                ws.cell(row=row_num, column=current_col, value=row_data['common_data'].get(col_name, "")).alignment = Alignment(wrap_text=True, vertical='top')
                current_col += 1

            # Count valid suppliers
            valid_count = 0
            from collections import defaultdict

            # Create a mapping for supplier base names to their R1/R2 data
            supplier_round_map = defaultdict(lambda: {"R1": {}, "R2": {}})

            # Organize supplier data by base name and round
            for supplier_key, supplier_data in row_data['supplier_data'].items():
                if " - R1" in supplier_key:
                    base = supplier_key.replace(" - R1", "")
                    supplier_round_map[base]["R1"] = supplier_data
                elif " - R2" in supplier_key:
                    base = supplier_key.replace(" - R2", "")
                    supplier_round_map[base]["R2"] = supplier_data

            # Evaluate validity (prefer R2 if it has valid values, else R1)
            for supplier, rounds in supplier_round_map.items():
                r2_data = rounds["R2"]
                r1_data = rounds["R1"]

                if any(is_valid_supplier_value(v) for v in r2_data.values()):
                    valid_count += 1
                elif any(is_valid_supplier_value(v) for v in r1_data.values()):
                    valid_count += 0

            ws.cell(row=row_num, column=current_col, value=valid_count).alignment = Alignment(wrap_text=True, vertical='top')
            if valid_count > 0:
                ws.cell(row=row_num, column=current_col).fill = green_fill
            current_col += 1

            # Supplier data
            for supplier_key, cols in supplier_grouped.items():
                data = row_data['supplier_data'].get(supplier_key, {})
                for full_col_name in cols:
                    value = data.get(full_col_name, "")
                    ws.cell(row=row_num, column=current_col, value=value).alignment = Alignment(wrap_text=True, vertical='top')
                    current_col += 1

            row_num += 1

        # 🧱 Step 4: Auto-set column widths
        for col_num in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 20

        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 30
        for r in range(3, ws.max_row + 1):
            ws.row_dimensions[r].height = 30

        wb.save(OUTPUT_FILE)
        print(f"[SUCCESS] Saved Excel file to: {OUTPUT_FILE}")

        
    def process_cleaned_files(): 
        try: 
            cleaned_files = os.listdir(CLEANED_FILES_FOLDER_LOCATION)
            r1_excel_files = [f for f in cleaned_files if f.endswith(('.xlsx', '.xls')) and '_r2' not in f.lower() and 'NTC' not in f]
            r2_excel_files = [f for f in cleaned_files if f.endswith(('.xlsx', '.xls')) and '_r2' in f.lower() and 'NTC' not in f]

            if not r1_excel_files and not r2_excel_files:
                print("No Excel files found in the cleaned files folder!")
                return
            print(f"Found {len(r1_excel_files)} Round 1 and {len(r2_excel_files)} Round 2 Excel files")

            all_data = {}       # {'giraffe_stainless_r1': df}
            display_names = {}  # {'giraffe_stainless_r1': 'Original filename'}

            # Load R1 files
            for item in r1_excel_files:
                file_path = os.path.join(CLEANED_FILES_FOLDER_LOCATION, item)
                try:
                    df = pd.read_excel(file_path, engine='openpyxl')
                    # df.columns = [str(col).strip() for col in df.columns]  # Normalize columns
                    file_name_key = os.path.splitext(item)[0]
                    all_data[file_name_key] = df
                    print(f"[R1] Loaded {len(df)} rows from {item}")
                except Exception as e:
                    print(f"[R1] Error processing {item}: {str(e)}")

            # Load R2 files
            for item in r2_excel_files:
                file_path = os.path.join(CLEANED_FILES_FOLDER_LOCATION, item)
                try:
                    df = pd.read_excel(file_path, engine='openpyxl')
                    # df.columns = [str(col).strip() for col in df.columns]  # Normalize columns
                    file_name_key = os.path.splitext(item)[0]
                    all_data[file_name_key] = df
                    print(f"[R2] Loaded {len(df)} rows from {item}")
                except Exception as e:
                    print(f"[R2] Error processing {item}: {str(e)}")

            if not all_data:
                print("No valid data found in any files!")
                return

            print("\nChecking for additional information data...")
            suppliers_with_additional_info = check_additional_info_data(all_data)

            print(f"Suppliers with additional info: {suppliers_with_additional_info}")
            print("\nCreating consolidated dataset...")
            consolidated_data = create_consolidated_dataset(all_data, suppliers_with_additional_info, display_names)

            print("Creating consolidated Excel file with row-wise outlier detection and statistical columns...")
            create_consolidated_excel(consolidated_data, list(all_data.keys()), suppliers_with_additional_info)

            print(f"\nConsolidated file created: {OUTPUT_FILE}")
            print(f"Suppliers including additional info column: {len(suppliers_with_additional_info)}")
            print(f"Suppliers excluding additional info column: {len(all_data) - len(suppliers_with_additional_info)}")
            print("Added statistical columns: Mean, Variance, and Standard Deviation for Total Cost Per UOM")

        except Exception as e: 
            print(f"Error in process_cleaned_files: {str(e)}")

    process_cleaned_files()

