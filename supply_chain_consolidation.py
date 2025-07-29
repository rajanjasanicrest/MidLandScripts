import pandas as pd
import os 
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from handler.handler import extract_supplier_name

# All Folder location information 
FILES_FOLDER_LOCATION = "./files round 2"
CLEANED_FILES_FOLDER_LOCATION = "./cleaned_files/supply_chain"
CONSOLIDATED_FILE_LOCATION = "./new"
CONSOLIDATED_FILE_NAME = "supplier_chain_improve_bidsheet_consolidate"

# Sheet names that we're processing
SHEET_NAMES = ["lead_estimation", "retails_packaging", "order_quantity"]

# -------------------------- Lead estimation related columns ------------------------ #
LEAD_COMMON_COLUMN_NAMES = ["L#"] 
LEAD_SUPPLIER_COLUMN_NAMES = [
    "Lead time - time frame", 
    "Lead time in days from PO receipt to shipment Port of Origin/Departure", 
    "% improvement versus the first quarter of year 1"
]

# -------------------------- Retails related columns -------------------------------- #
RETAILS_COMMON_COLUMN_NAMES = ["S#"] 
RETAILS_SUPPLIER_COLUMN_NAMES = [
    "Product secondary packaging", 
    "Secondary packaging label", 
    "Product primary packaging",
    "Primary packaging label",
    "Individual part tagging",
    "Surcharge % on EXW price"
]

# ------------------------- Order related columns ---------------------------------- # 
ORDER_COMMON_COLUMN_NAMES = ["O#"]
ORDER_SUPPLIER_COLUMN_NAMES = [
    "Part family description e.g. SAE 45° flare brass fittings - female adapters, one-way brass ball valve FGH x MGH", 
    "Order quantity range in whole number", 
    "Discount % on EXW price"    
]

def format_numeric_value(value):
    """
    Format numeric values to 4 decimal places if they are integers or floats
    Returns the formatted value or original value if not numeric
    """
    try:
        # Check if value is numeric (int, float, or numeric string)
        if pd.isna(value) or value == "":
            return value
        
        # Convert to float if it's a string representation of a number
        if isinstance(value, str):
            try:
                numeric_value = float(value)
            except ValueError:
                return value  # Return original if not convertible to number
        elif isinstance(value, (int, float)):
            numeric_value = float(value)
        else:
            return value  # Return original if not a numeric type
        
        # Format to 4 decimal places
        return f"{numeric_value:.4f}"
    
    except (ValueError, TypeError):
        return value  # Return original value if any error occurs

def cleanup_files():
    """
    Process raw Excel files and create cleaned versions with separate sheets
    """
    files = os.listdir(FILES_FOLDER_LOCATION)
    files2 = os.listdir("./files")

    all_files = files + files2
    
    for item in all_files: 
        print("Processing excel sheet ---------------------------------------")
        print(f"./files round 2/{item}")

        # Cleaned csv file name 
        cleaned_csv_file_name = f"{item.split('.')[0]}_cleaned.xlsx"

        # Make sure folder exists or not 
        os.makedirs(CLEANED_FILES_FOLDER_LOCATION, exist_ok=True)

        try:
            if item in files:
                df = pd.read_excel(f"{FILES_FOLDER_LOCATION}/{item}", sheet_name="4. Supply Chain Improv bidsheet")
            else:
                df = pd.read_excel(f"./files/{item}", sheet_name="4. Supply Chain Improv bidsheet")

        except Exception as e:
            print(f"Failed to read sheet from {item}: {e}")
            continue

        # Find the row containing "L #"
        lead_header_start_row_index = None
        lead_header_end_row_index = None
        lead_col_start_index = None

        # Check Lead estimation row related values ----------------------------------------------
        for index, row in df.iterrows():
            for col_index, value in enumerate(row):
                if isinstance(value, str) and "L#" in value:
                    lead_header_start_row_index = index
                    lead_col_start_index = col_index
                    
                    for order_check_index, order_row_check in df.iterrows(): 
                        for order_col_index, order_value in enumerate(order_row_check):
                            if isinstance(order_value, str) and "Order quantity improvement" in order_value:
                                lead_header_end_row_index = order_check_index

            if lead_header_start_row_index is not None:
                break

        if (
            lead_header_start_row_index is not None and 
            lead_header_end_row_index is not None and 
            lead_header_end_row_index > lead_header_start_row_index
        ):
            # Extract data from 2nd to 5th columns (i.e., col 1 to 4)
            lead_df = df.iloc[lead_header_start_row_index:lead_header_end_row_index, 1:5]
            print(f"Lead data extracted: {len(lead_df)} rows")
        else:
            lead_df = None
            print("Lead header start/end indexes not found properly.")

        # Retails packaging related row related values -----------------------------------------------
        retails_header_start_row_index = None
        retails_header_end_row_index = None
        retails_col_start_index = None

        for index, row in df.iterrows():
            for col_index, value in enumerate(row):
                if isinstance(value, str) and "S#" in value:
                    retails_header_start_row_index = index
                    retails_col_start_index = col_index
                    
                    for retails_check_index, retails_row_check in df.iterrows(): 
                        for retails_col_index, retails_order_value in enumerate(retails_row_check):
                            if isinstance(retails_order_value, str) and "Lead time improvement" in retails_order_value:
                                retails_header_end_row_index = retails_check_index

        if (
            retails_header_start_row_index is not None and 
            retails_header_end_row_index is not None and 
            retails_header_end_row_index > retails_header_start_row_index
        ):
            # Extract header row
            header_row = df.iloc[retails_header_start_row_index:retails_header_start_row_index + 1, 1:8]

            # Skip the row after header, then take rest
            data_rows = df.iloc[retails_header_start_row_index + 2:retails_header_end_row_index, 1:8]

            # Concatenate header + data
            retails_df = pd.concat([header_row, data_rows], ignore_index=True)
            print(f"Retails data extracted: {len(retails_df)} rows")
        else:
            retails_df = None
            print("Retails header start/end indexes not found properly.")

        # Order quantity related data information ---------------------------------------------------
        order_quantity_start_row_index = None
        order_quantity_col_start_index = None

        # Find the start row for "O#"
        for index, row in df.iterrows():
            for col_index, value in enumerate(row):
                if isinstance(value, str) and "O#" in value:
                    order_quantity_start_row_index = index
                    order_quantity_col_start_index = col_index
                    break
            if order_quantity_start_row_index is not None:
                break

        # Extract order-related data from "O#" to the end of the sheet
        if order_quantity_start_row_index is not None:
            order_quantity_df = df.iloc[order_quantity_start_row_index:, 1:5]  # columns 2 to 5
            print(f"Order quantity data extracted: {len(order_quantity_df)} rows")
        else:
            order_quantity_df = None
            print("Order quantity section ('O#') not found.")

        # Save all three dataframes to the same Excel file
        if any([lead_df is not None, retails_df is not None, order_quantity_df is not None]):
            with pd.ExcelWriter(f"{CLEANED_FILES_FOLDER_LOCATION}/{cleaned_csv_file_name}", engine='openpyxl') as writer:
                if lead_df is not None:
                    lead_df.to_excel(writer, index=False, header=False, sheet_name="lead_estimation")
                if retails_df is not None:
                    retails_df.to_excel(writer, index=False, header=False, sheet_name="retails_packaging")
                if order_quantity_df is not None:
                    order_quantity_df.to_excel(writer, index=False, header=False, sheet_name="order_quantity")
            print(f"✅ Cleaned file saved: {cleaned_csv_file_name}")
        else:
            print(f"⚠️ Skipping saving {cleaned_csv_file_name} - no valid sheets found")

def get_sheet_column_config(sheet_name):
    """
    Get the appropriate column configuration for each sheet type
    """
    if sheet_name == "lead_estimation":
        return LEAD_COMMON_COLUMN_NAMES, LEAD_SUPPLIER_COLUMN_NAMES
    elif sheet_name == "retails_packaging":
        return RETAILS_COMMON_COLUMN_NAMES, RETAILS_SUPPLIER_COLUMN_NAMES
    elif sheet_name == "order_quantity":
        return ORDER_COMMON_COLUMN_NAMES, ORDER_SUPPLIER_COLUMN_NAMES
    else:
        return [], []

def create_consolidated_dataset_rowwise(all_data, common_columns, supplier_columns):
    """
    Create a consolidated dataset in row-wise format
    Each row will contain: Supplier Name, Common Columns, Supplier Columns
    """
    consolidated_rows = []
    
    for file_name, df in all_data.items():
        print(f"Processing data from {file_name}...")
        
        if df is None or df.empty:
            print(f"No data found in {file_name}, skipping...")
            continue
        
        # Extract supplier name
        try:
            supplier_name = file_name.split('--')[-1].strip()
            supplier_name = supplier_name.replace("_cleaned", "").replace(" R2", "").strip()
        except:
            supplier_name = file_name.replace("_cleaned", "")
        
        for _, row in df.iterrows():
            # Create a single row with supplier name, common data, and supplier data
            consolidated_row = {}
            
            # Add supplier name as first column
            consolidated_row['Supplier Name'] = supplier_name
            
            # Add common columns data with decimal formatting
            for i, col in enumerate(common_columns):
                # Use column index if column name not found
                if col in df.columns:
                    value = row[col] if pd.notna(row[col]) else ""
                elif i < len(df.columns):
                    value = row.iloc[i] if pd.notna(row.iloc[i]) else ""
                else:
                    value = ""
                
                # Apply decimal formatting to numeric values
                consolidated_row[col] = format_numeric_value(value)
            
            # Add supplier-specific columns data with decimal formatting
            for i, col in enumerate(supplier_columns):
                # Use column index offset by number of common columns
                col_index = i + len(common_columns)
                if col in df.columns:
                    value = row[col] if pd.notna(row[col]) else ""
                elif col_index < len(df.columns):
                    value = row.iloc[col_index] if pd.notna(row.iloc[col_index]) else ""
                else:
                    value = ""
                
                # Apply decimal formatting to numeric values
                consolidated_row[col] = format_numeric_value(value)
            
            consolidated_rows.append(consolidated_row)
    
    return consolidated_rows

def create_consolidated_excel_rowwise(consolidated_data, common_columns, supplier_columns, output_file):
    """
    Create Excel file with row-wise structure
    Format: Supplier Name | Common Columns | Supplier Columns
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated Suppliers"
    
    # Define styles - All headers will have green background
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
    wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Create a number format style for 4 decimal places
    decimal_style = NamedStyle(name="decimal_4", number_format="0.0000")
    
    # Create headers
    headers = ['Supplier Name'] + common_columns + supplier_columns
    
    # Add headers to first row - All with green background
    for col_num, header in enumerate(headers, 1):
        header_cell = ws.cell(row=1, column=col_num, value=header)
        header_cell.alignment = wrap_alignment
        header_cell.fill = green_fill  # All headers get green background
    
    # Add data rows
    for row_num, row_data in enumerate(consolidated_data, 2):  # Start from row 2
        for col_num, header in enumerate(headers, 1):
            cell_value = row_data.get(header, "")
            data_cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            data_cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Apply decimal formatting to numeric cells (already formatted in data)
            # The formatting is already applied in the data, so no need to reformat here
    
    # Set column widths and row heights
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20  # Fixed width of 20
    
    # Set row heights to accommodate wrapped text
    for row_num in range(1, ws.max_row + 1):
        ws.row_dimensions[row_num].height = 30
    
    # Save the workbook
    wb.save(output_file)
    print(f"Consolidated file saved: {output_file}")

def create_combined_consolidated_excel_rowwise(all_sheet_data, output_file):
    """
    Create a single Excel file with multiple sheets for all sheet types in row-wise format
    """
    wb = Workbook()
    
    # Remove the default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    header_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
    new_supplier_fill = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')  # Light bluish
    wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Create a number format style for 4 decimal places
    decimal_style = NamedStyle(name="decimal_4", number_format="0.0000")
    if "decimal_4" not in wb.named_styles:
        wb.add_named_style(decimal_style)
    
    for sheet_name, sheet_info in all_sheet_data.items():
        if not sheet_info['consolidated_data']:
            print(f"No data for sheet: {sheet_name}, skipping...")
            continue
            
        print(f"Creating sheet: {sheet_name}")
        
        # Create new worksheet
        ws = wb.create_sheet(title=sheet_name)
        
        consolidated_data = sheet_info['consolidated_data']
        common_columns = sheet_info['common_columns']
        supplier_columns = sheet_info['supplier_columns']
        
        # Create headers
        headers = ['Supplier Name'] + common_columns + supplier_columns
        
        # Add headers to first row - All with green background
        for col_num, header in enumerate(headers, 1):
            header_cell = ws.cell(row=1, column=col_num, value=header)
            header_cell.alignment = wrap_alignment
            header_cell.fill = header_fill  # All headers get green background
        
        previous_supplier = None 

        # Add data rows
        for row_num, row_data in enumerate(consolidated_data, 2):  # Start from row 2
            current_supplier = row_data.get("Supplier Name", "")
            is_new_supplier = current_supplier != previous_supplier

            for col_num, header in enumerate(headers, 1):
                cell_value = row_data.get(header, "")
                data_cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                data_cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Apply background fill for new suppliers
                if is_new_supplier:
                    data_cell.fill = new_supplier_fill
                
                # The numeric formatting is already applied in the data processing step
                # Values are already formatted to 4 decimal places where applicable
        
            previous_supplier = current_supplier
        
        # Set column widths and row heights
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 20  # Fixed width of 20
        
        # Set row heights to accommodate wrapped text
        for row_num in range(1, ws.max_row + 1):
            ws.row_dimensions[row_num].height = 30
    
    # Save the workbook
    wb.save(output_file)
    print(f"Combined consolidated file saved: {output_file}")

def process_cleaned_files(): 
    """
    Process all cleaned files and create a single consolidated Excel file with multiple sheets in row-wise format
    """
    try:
        # Make sure consolidated folder exists
        os.makedirs(CONSOLIDATED_FILE_LOCATION, exist_ok=True)
        
        cleaned_files = os.listdir(CLEANED_FILES_FOLDER_LOCATION)
        # Reading cleaned files related csv files 
        from collections import defaultdict

        cleaned_files = os.listdir(CLEANED_FILES_FOLDER_LOCATION)

        # Group files by supplier name (normalized)
        supplier_files = defaultdict(dict)
        used_r1 = []
        used_r2 = []

        for f in cleaned_files:
            if not f.endswith(('.xlsx', '.xls')):
                continue

            base = f.replace("_cleaned.xlsx", "")
            is_r2 = "R2" in base

            # Normalize supplier key (remove " R2" if present)
            supplier_key = base.replace(" R2", "")

            if is_r2:
                supplier_files[supplier_key]['R2'] = f
            else:
                supplier_files[supplier_key]['R1'] = f

        # Prefer R2 if available
        excel_files = []
        for supplier, files in supplier_files.items():
            if files.get('R2'):
                excel_files.append(files['R2'])
                used_r2.append(supplier)
            elif files.get('R1'):
                excel_files.append(files['R1'])
                used_r1.append(supplier)

        print(f"\n✔️ Total unique suppliers considered: {len(supplier_files)}")
        print(f"   - From R2: {len(used_r2)} suppliers")
        print(f"   - From R1 (fallback): {len(used_r1)} suppliers")
        missing_suppliers = set(supplier_files.keys()) - set(used_r1) - set(used_r2)
        if missing_suppliers:
            print(f"⚠️ Warning: {len(missing_suppliers)} suppliers had neither R1 nor R2: {missing_suppliers}")

        # Dictionary to store all sheet data
        all_sheet_data = {}

        # Process each sheet type separately
        for sheet_name in SHEET_NAMES:
            print(f"\n{'='*60}")
            print(f"Processing sheet type: {sheet_name}")
            print(f"{'='*60}")
            
            # Get column configuration for this sheet type
            common_columns, supplier_columns = get_sheet_column_config(sheet_name)
            
            if not common_columns and not supplier_columns:
                print(f"No column configuration found for sheet: {sheet_name}")
                continue
            
            # Dictionary to store all data from files for this sheet type
            all_data = {}

            for item in excel_files: 
                print(f"Processing file: {item} - Sheet: {sheet_name}")
                file_path = os.path.join(CLEANED_FILES_FOLDER_LOCATION, item)

                try: 
                    # Read the specific sheet from the Excel file 
                    df = pd.read_excel(file_path, sheet_name=sheet_name)

                    # Store the dataframe for later use
                    file_name_key = os.path.splitext(item)[0]  # Remove extension
                    all_data[file_name_key] = df

                    print(f"Successfully loaded {len(df)} rows from {item} - {sheet_name}")
                    if not df.empty:
                        print(f"Columns in file: {list(df.columns)}")
                
                except Exception as e: 
                    print(f"Error processing {item} - {sheet_name}: {str(e)}")
                    all_data[os.path.splitext(item)[0]] = None
                    continue
            
            if not any(df is not None and not df.empty for df in all_data.values()):
                print(f"No valid data found for sheet type: {sheet_name}")
                all_sheet_data[sheet_name] = {
                    'consolidated_data': [],
                    'common_columns': common_columns,
                    'supplier_columns': supplier_columns
                }
                continue
            
            # Create consolidated dataset for this sheet type in row-wise format
            print(f"\nCreating consolidated dataset for {sheet_name} in row-wise format...")
            consolidated_data = create_consolidated_dataset_rowwise(all_data, common_columns, supplier_columns)
            
            # Store sheet data for combined Excel creation
            all_sheet_data[sheet_name] = {
                'consolidated_data': consolidated_data,
                'common_columns': common_columns,
                'supplier_columns': supplier_columns
            }
            
            print(f"Total rows created for {sheet_name}: {len(consolidated_data)}")
        
        # Create single combined Excel file with all sheets in row-wise format
        output_file = f"{CONSOLIDATED_FILE_LOCATION}/{CONSOLIDATED_FILE_NAME}.xlsx"
        print(f"\n{'='*60}")
        print(f"Creating combined consolidated Excel file: {output_file}")
        print(f"{'='*60}")
        
        create_combined_consolidated_excel_rowwise(all_sheet_data, output_file)
        
        print(f"\n{'='*60}")
        print("Combined consolidated file created successfully!")
        print(f"File saved: {output_file}")
        print(f"Sheets included: {list(all_sheet_data.keys())}")
        
        # Print summary of rows per sheet
        for sheet_name, sheet_info in all_sheet_data.items():
            row_count = len(sheet_info['consolidated_data'])
            print(f"  - {sheet_name}: {row_count} rows")
        
        print(f"{'='*60}")
        
    except Exception as e: 
        print(f"Error in process_cleaned_files: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    # First clean up the raw files
    print("Starting file cleanup process...")
    # cleanup_files()
    
    print("\nStarting consolidation process...")
    process_cleaned_files()