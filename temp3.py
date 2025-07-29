# === Master Consolidation File ====


import re
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment


# === Step 1: Merge Files ===
file_info = [
    ("consolidate/bidsheet_steel_outlier_consolidate.xlsx", "steel"),
    ("consolidate/bidsheet_brass_outlier_consolidate.xlsx", "brass"),
    ("consolidate/bidsheet_other_metal_outlier_consolidate.xlsx", "other metal")
]

dfs = []

for file_path, metal_type in file_info:
    df = pd.read_excel(file_path, header=[0, 1])
    new_cols = []
    counter = {}

    for col in df.columns:
        lvl1, lvl2 = col
        if lvl1.startswith("Unnamed") or pd.isna(lvl1):
            new_cols.append(("", lvl2))
        else:
            base = lvl1.strip()
            counter[base] = counter.get(base, 0) + 1
            unique_lvl1 = f"{base}_{counter[base]}"
            new_cols.append((unique_lvl1, lvl2))

    df.columns = pd.MultiIndex.from_tuples(new_cols)
    df.insert(5, ('', 'type'), metal_type)
    dfs.append(df)

combined_df = pd.concat(dfs, axis=0, ignore_index=True)

first_6 = dfs[0].columns[:5].tolist() + [('', 'type')]  
last_3 = dfs[0].columns[-3:].tolist()
middle = [col for col in combined_df.columns if col not in first_6 and col not in last_3]
middle_sorted = sorted(middle, key=lambda x: (re.sub(r"_\d+$", "", x[0])))

combined_df = combined_df.loc[:, first_6 + middle_sorted + last_3]

cleaned_cols = []
for lvl1, lvl2 in combined_df.columns:
    if lvl1 == "":
        cleaned_cols.append(("", lvl2))
    else:
        cleaned_lvl1 = re.sub(r"_\d+$", "", lvl1)
        cleaned_cols.append((cleaned_lvl1, lvl2))

combined_df.columns = pd.MultiIndex.from_tuples(cleaned_cols)

columns_to_remove = []
for col in combined_df.columns:
    lvl1, lvl2 = col
    if ("additional information" in str(lvl2).lower() and
        "please use this column only if absolutely necessary" in str(lvl2).lower()):
        has_data = any(pd.notna(v) and str(v).strip() != "" for v in combined_df[col])
        if not has_data:
            columns_to_remove.append(col)

if columns_to_remove:
    combined_df = combined_df.drop(columns=columns_to_remove)

def format_value(value):
    if pd.isna(value):
        return ""
    elif isinstance(value, (int, float, np.number)):
        return "0" if value == 0 else round(float(value), 4)
    return value

for col_idx, col in enumerate(combined_df.columns):
    if col_idx >= 4:
        combined_df[col] = combined_df[col].apply(format_value)

# === Step 2: Save merged sheet temporarily for IQR processing ===
merged_file = "combined_bidsheet_outlier_2.xlsx"
flattened_df = pd.DataFrame(columns=["{}-{}".format(a, b) if a else b for a, b in combined_df.columns])
for col_idx, col in enumerate(combined_df.columns):
    flattened_df[flattened_df.columns[col_idx]] = combined_df[col]

flattened_df.to_excel(merged_file, index=False)

# === Step 3: IQR Outlier Detection ===
input_file = merged_file
df = pd.read_excel(input_file)

target_cols = [col for col in df.columns if "-Total Cost" in str(col)]
outlier_mask = pd.DataFrame(False, index=df.index, columns=target_cols)
iqr_values = []

for idx, row in df.iterrows():
    values = row[target_cols].dropna()
    values = [float(v) for v in values if v != 0]
    values = pd.Series(values)
    if len(values) <= 1:
        iqr_values.append(np.nan)
        continue
    Q1 = values.quantile(0.25)
    Q3 = values.quantile(0.75)
    IQR = Q3 - Q1
    iqr_values.append(IQR)
    lower = Q1 - 1.5 * IQR
    upper = Q3 + 1.5 * IQR
    for col in target_cols:
        val = row[col]
        if pd.notnull(val):
            val_float = float(val)
            if val_float < lower or val_float > upper:
                outlier_mask.at[idx, col] = True

df["IQR_Value"] = iqr_values
df["IQR_Value"] = df["IQR_Value"]

mean_without_high_outliers = []

for idx, row in df.iterrows():
    bids = []
    IQR = df.at[idx, "IQR_Value"]
    values = row[target_cols].dropna()

    for col in target_cols:
        val = row[col]
        if pd.isna(val) or float(val) == 0:
            continue

        val_float = float(val)

        # Include if NOT a high outlier (i.e., val <= Q3 + 1.5*IQR)
        if not outlier_mask.at[idx, col] or val_float <= values.quantile(0.75) + 1.5 * IQR:
            bids.append(val_float)

    if bids:
        mean_val = round(sum(bids) / len(bids), 4)
    else:
        mean_val = ""

    mean_without_high_outliers.append(mean_val)

df["Arithmetic Average "] = mean_without_high_outliers

output_file = "output_with_outliers.xlsx"


# === Add Min Bid, Min Supplier, Outlier Flag, Second Minimum Bid ===
def extract_supplier(col_name):
    return str(col_name).split("-")[0].strip()

min_bids = []
min_suppliers = []
second_min_bids = []
second_min_suppliers = []
has_outlier_flags = []

for idx, row in df.iterrows():
    bids = {}
    outlier_bids = {}
    valid_for_avg = []

    IQR = row.get("IQR_Value", None)

    for col in target_cols:
        val = row[col]
        if pd.isna(val) or val == "" or float(val) == 0:
            continue

        supplier = extract_supplier(col)
        val_float = float(val)
        bids[supplier] = val_float

        # Save all outliers
        if outlier_mask.at[idx, col]:
            outlier_bids[supplier] = val_float

        # Valid for average if:
        # - Not a high outlier (i.e., value <= Q3 + 1.5*IQR)
        # - Not 0 or NaN (already handled)
        if IQR is not None:
            Q3 = row[target_cols].dropna().astype(float).quantile(0.75)
            upper = Q3 + 1.5 * IQR
            if val_float <= upper:
                valid_for_avg.append(val_float)
        else:
            valid_for_avg.append(val_float)  # fallback

    # --- Min and Second Min Logic ---
    if bids:
        sorted_bids = sorted(bids.items(), key=lambda x: x[1])
        min_supplier, min_bid = sorted_bids[0]
        if len(sorted_bids) > 1:
            second_min_supplier, second_min_bid = sorted_bids[1]
        else:
            second_min_supplier = ""
            second_min_bid = ""
    else:
        min_bid = ""
        min_supplier = ""
        second_min_bid = ""
        second_min_supplier = ""

    # --- Has lower-end outlier logic ---
    has_lower_outlier = any(
        supplier in outlier_bids and outlier_bids[supplier] == min_bid
        for supplier in outlier_bids
    )

    # Append everything
    min_bids.append(min_bid)
    min_suppliers.append(min_supplier)
    second_min_bids.append(second_min_bid)
    second_min_suppliers.append(second_min_supplier)
    has_outlier_flags.append("Yes" if has_lower_outlier else "No")

# Assign to dataframe
df["Min Bid"] = min_bids
df["Min Supplier"] = min_suppliers
df["Has Outlier"] = has_outlier_flags
df["2nd Lowest Minimum Bid"] = second_min_bids
df["2nd Lowest Minimum Bid Supplier"] = second_min_suppliers

cols = df.columns.tolist()
try:
    idx = cols.index("Valid Supplier")
    new_cols = cols[:idx+1] + ["Min Bid", "Min Supplier", "Has Outlier", "2nd Lowest Minimum Bid", "2nd Lowest Minimum Bid Supplier"] + [c for c in cols if c not in cols[:idx+1] + ["Min Bid", "Min Supplier", "Has Outlier", "2nd Lowest Minimum Bid", "2nd Lowest Minimum Bid Supplier"]]
    df = df[new_cols]
except ValueError:
    pass


df = df.replace({np.nan: ""})
df.to_excel(output_file, index=False)

# === Step 4: Styling ===
wb = load_workbook(output_file)
ws = wb.active

orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
light_gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

headers = [cell.value for cell in ws[1]]

# Highlight outliers in orange
for i, col in enumerate(df.columns):
    if col in target_cols:
        col_letter = get_column_letter(i + 1)
        for j in range(len(df)):
            if outlier_mask.iloc[j, target_cols.index(col)] and df.iloc[j, i] != 0:
                ws[f"{col_letter}{j + 2}"].fill = orange_fill

# Last 5 columns → Yellow
last_5_col_indices = range(len(headers) - 5, len(headers))
for col_idx in last_5_col_indices:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx + 1, max_col=col_idx + 1):
        for cell in row:
            cell.fill = yellow_fill

# 5th column → Light Green
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=13, max_col=13):
    for cell in row:
        cell.fill = light_green_fill

# From 19th column → alternate header colors by company
fills = [light_gray_fill, white_fill]

for idx, col_idx in enumerate(range(19, len(headers))):
    fill = fills[idx % 2]
    col_letter = get_column_letter(col_idx + 1)
    ws[f"{col_letter}1"].fill = fill


# # Right-align from 7th column onward (excluding "Additional information")
# for col_idx, header in enumerate(headers[6:], start=6):
#     if "Additional information" not in str(header):
#         for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx + 1, max_col=col_idx + 1):
#             for cell in row:
#                 cell.alignment = Alignment(horizontal="right")


# === Clean any lingering NaN/None cells in the worksheet ===
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        if cell.value is None or (isinstance(cell.value, float) and pd.isna(cell.value)) or cell.value=='nan':
            cell.value = ""


# === Apply number format from column 13 (G) onwards, excluding "Additional information" ===
# Loop from column 13 (G) to total_cols
total_cols = ws.max_column

for col_idx, header in enumerate(headers[20:], start=20):
    is_last_five = col_idx >= total_cols - 5  # Check if current col is one of the last 4

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx + 1, max_col=col_idx + 1):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.0000'
            elif cell.value is None or cell.value == '':
                if is_last_five:
                    cell.value = '-'
                else:
                    cell.value = 0.0000
                    cell.number_format = '0.0000'

# === Apply text format '@' to 'Part #' column to prevent Excel date conversion ===
try:
    part_col_idx = headers.index('Part #') + 1  # openpyxl is 1-based
except ValueError:
    part_col_idx = None

if part_col_idx:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=part_col_idx, max_col=part_col_idx):
        for cell in row:
            cell.number_format = '@'  # Excel text format
            if cell.value is not None:
                cell.value = str(cell.value)  # force string to avoid auto-conversion

# === Identify relevant column indices ===
sellonsky_col_idx = next((i + 1 for i, h in enumerate(headers) if 'Sellonsky-Safehome-' in h), None) + 1
valid_supplier_col_idx = next((i + 1 for i, h in enumerate(headers) if h.strip() == 'Valid Supplier'), None)

if sellonsky_col_idx and valid_supplier_col_idx:
    incoterm_col_idx = valid_supplier_col_idx + 1

    # Insert new column to make space for 'Incoterm'
    ws.insert_cols(incoterm_col_idx)

    # Update headers list (since we added a column)
    ws.cell(row=1, column=incoterm_col_idx, value='Incoterm')

    # Fill in Incoterm values
    for row in range(2, ws.max_row + 1):
        sellonsky_value = ws.cell(row=row, column=sellonsky_col_idx).value
        if isinstance(sellonsky_value, (int, float)) and sellonsky_value > 0:
            ws.cell(row=row, column=incoterm_col_idx).value = 'DDP'

wb.save(output_file)
print(f"✅ All processing complete. Final output saved to: {output_file}")