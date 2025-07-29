import pandas as pd
import numpy as np
import re
import os
import time
from datetime import datetime
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

start_time = time.time()

bidsheet_file = "new/bidsheet_master_consolidate.xlsx"
wapp_file = "wapp2.xlsx"
p21_file = "P21 supplier bid supplier norm 070725v3.xlsx"
output_file = "new/bidsheet_updated.xlsx"
wierd_list = ["7000-04-06","7000-08-04","7000-08-16","7000-10-06","7000-12-16","7002-06-04","7002-06-06","7002-06-08","7002-08-06","7002-08-08","7002-08-12","7002-12-08","7003-04-04","7003-06-06","7003-06-08","7003-08-08","7003-12-12","7004-04-02","7004-04-04","7004-04-06","7004-04-08","7004-06-04","7004-06-06","7004-06-08","7004-08-06","7004-08-08","7004-08-10","7004-08-12","7004-10-08","7004-10-10","7004-10-12","7004-12-12","7004-12-16","7012-05-04","7022-06-08","7022-10-06","7022-10-10","7022-12-08","7022-12-16","7032-02-02","7032-02-04","7032-02-06","7032-04-02","7032-04-04","7032-04-06","7032-04-08","7032-04-12","7032-06-02","7032-06-04","7032-06-06","7032-06-08","7032-06-12","7032-06-16","7032-08-04","7032-08-06","7032-08-08","7032-08-12","7032-08-16","7032-12-04","7032-12-06","7032-12-08","7032-12-12","7032-12-16","7033-02-04","7033-04-02","7033-04-04","7033-04-06","7033-06-04","7033-06-06","7033-06-08","7033-08-04","7033-08-06","7033-08-08","7033-08-12","7033-12-04","7033-12-06","7033-12-08","7033-12-12","7034-04-02","7034-04-04","7034-06-06","7034-08-06","7034-08-08","7034-12-12","7040-04-04","7040-08-12","7040-12-12","7042-02-04","7042-04-06","7042-04-08","7042-06-04","7042-06-08","7042-08-06","7042-12-08","7042-12-16","7062-04-04","7062-06-04","7062-06-06","7062-06-08","7062-08-06","7062-08-10","7062-10-08","7062-10-12","7062-12-06","7062-12-08","7062-12-12","7062-12-16","7202-04-06","7202-04-08","7202-05-04","7202-06-02","7202-08-04","7202-10-06","7202-10-12","7204-04-04","7204-06-06","7204-12-12","7802-04-06","7802-06-04","7802-06-06","7802-06-08","7802-08-06","9000-02-02","9000-04-02","9000-04-04","9000-06-04","9000-06-06","9000-08-04","9000-08-06","9000-08-08","9000-10-10","9000-12-08","9000-12-12","9001-06-02","9001-06-04","9001-08-04","9001-08-06","9001-12-06","9001-12-08","9020-04-02","9020-04-04","9020-04-06","9020-04-08","9020-06-04","9020-06-06","9020-06-08","9020-06-12","9020-06-16","9020-08-04","9020-08-06","9020-08-08","9020-08-12","9020-10-08","9020-10-12","9020-12-04","9020-12-06","9020-12-08","9020-12-12","9020-12-16","9022-04-02","9022-04-04","9022-06-04","9022-06-06","9022-08-04","9022-08-06","9022-08-08","9022-10-08","9022-10-10","9022-12-04","9022-12-06","9022-12-08","9022-12-10","9022-12-12","9023-02-04","9023-02-08","9023-04-06","9023-04-08","9023-04-12","9023-06-02","9023-06-04","9023-06-08","9023-06-16","9024-02-04","9024-02-06","9024-04-02","9024-04-04","9024-04-06","9024-04-08","9024-04-12","9024-06-02","9024-06-04","9024-06-06","9024-06-08","9024-06-12","9024-08-04","9024-08-06","9024-08-08","9024-08-10","9024-08-12","9024-08-16","9024-10-08","9024-10-10","9024-10-12","9024-12-06","9024-12-08","9024-12-12","9024-12-16","9025-04-12","9025-06-16","9025-08-18","9025-08-20","9025-10-22","9025-12-26","9033-02-02","9033-02-06","9033-04-02","9033-04-04","9033-06-04","9033-06-06","9033-08-06","9033-08-08","9033-10-08","9033-10-10","9033-12-08","9033-12-12","9044-04-04","9044-06-06","9044-08-08","9222-04-04","9222-06-04","9222-06-06","9222-08-06","9222-08-08","9222-12-08","9222-12-10","9222-12-12","2501-08-02","2502-03-02","2503-04-08","2701-03-03","2702-05-05","5000-12-04","6400-08-05","6400-10-04","6400-10-14","6404-08-12","6500-06-04","6500-08-12","6505-10-06","7005-08-16","7005-08-18","7005-08-22","7062-08-12","7205-12-26","2403-04-02","2403-05-04","2404-04-12","2404-05-08","2404-06-16","2404-08-02","2404-12-24","2405-05-06","2406-04-04","2406-04-05","2406-05-04","2406-08-05","2406-08-08","7001-04-10","7001-04-12","7001-04-14","7001-04-16","7001-05-14","7001-06-12","7001-06-14","7001-06-16","7001-06-18","7001-08-14","7001-08-16","7001-08-18","7001-08-22","7001-10-18","7001-10-20","7001-10-22","7001-12-22","7001-12-26","7001-12-27","7005-04-18","7005-06-12","7005-06-14","7005-06-16","7005-06-18","7005-06-20","7005-06-22","7005-08-14","7005-08-20","7005-10-16","7005-10-26","7045-02-18","7045-04-16","7045-12-30"]

bidsheet_df = pd.read_excel(bidsheet_file)
wapp_df = pd.read_excel(wapp_file)
p21_df = pd.read_excel(p21_file, skiprows=2)
wapp_df['Norm Item ID'] = wapp_df['Norm Item ID'].astype(str).str.strip().str.upper()

def date_to_excel_serial(date_str):
    base_date = datetime(1899, 12, 30)  # Excel's day 0
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    delta = date_obj - base_date
    return delta.days

band_columns = {
    (1, 50): "1-50",
    (51, 200): "51-200",
    (201, 500): "201-500",
    (501, 1000): "501-1000",
    (1001, 2000): "1001-2000",
    (2001, 5000): "2001-5000",
    (5001, 10000): "5001-10000",
    (10001, 25000): "10001-25000",
    (25001, 100000): "25001-100000",
    (100001, 250000): "100001-250000",
    (250001, float('inf')): "250001+"
}

def get_band_column(aoq):
    for (low, high), col in band_columns.items():
        if low <= aoq <= high:
            return col
    return None

# --- Refined logic for fetching Volume-banded WAPP and Most common supplier ---
volume_wapp, mcs_list = [], []
norm_part_ids = bidsheet_df['Part #'].astype(str).str.strip().str.upper()
wapp_df['Norm Item ID'] = wapp_df['Norm Item ID'].astype(str).str.strip().str.upper()
bidsheet_df['Average Order Quantity (per UOM)'] = pd.to_numeric(bidsheet_df['Average Order Quantity (per UOM)'], errors='coerce')

# Find the actual column name for Most Common Supplier in wapp_df (case-insensitive, stripped)
def find_mcs_column(wapp_df):
    for col in wapp_df.columns:
        if str(col).strip().lower() == 'most common supplier':
            return col
    # Try partial match if exact not found
    for col in wapp_df.columns:
        if 'most common supplier' in str(col).strip().lower():
            return col
    return None

mcs_col_name = find_mcs_column(wapp_df)

for i, row in tqdm(bidsheet_df.iterrows(), total=len(bidsheet_df), desc='Volume WAPP'):
    part_id = norm_part_ids[i]
    aoq = row['Average Order Quantity (per UOM)']

    # Handle weird part numbers
    if part_id in [x.upper() for x in wierd_list]:
        try:
            excel_serial = str(date_to_excel_serial(row['Part #']))
        except Exception:
            excel_serial = None

        wapp_rows = wapp_df[wapp_df['Norm Item ID'] == excel_serial]
        if wapp_rows.empty:
            wapp_rows = wapp_df[wapp_df['Norm Item ID'] == part_id]
    else:
        wapp_rows = wapp_df[wapp_df['Norm Item ID'] == part_id]

    if wapp_rows.empty or pd.isna(aoq):
        volume_wapp.append('-')
        mcs_list.append('')
        continue

    wapp_row = wapp_rows.iloc[0]
    band_col = get_band_column(aoq)
    val = wapp_row.get(band_col, None)
    if val is None or (isinstance(val, float) and np.isnan(val)):
        val = wapp_row.get('Raw WAPP', '-')

    volume_wapp.append(round(val, 4) if isinstance(val, (int, float, np.floating)) and not pd.isna(val) else '-')

    # Fetch Most Common Supplier robustly
    mcs_val = wapp_row.get(mcs_col_name, '') if mcs_col_name else ''
    mcs_list.append(mcs_val if pd.notna(mcs_val) else '')

min_bid_idx = bidsheet_df.columns.get_loc("Final Min Bid")
bidsheet_df.insert(min_bid_idx, "Volume-banded WAPP", volume_wapp)
bidsheet_df.insert(min_bid_idx+1, "Most common supplier", mcs_list)

p21_df['p21_supplier_lower'] = p21_df['P21 supplier'].astype(str).str.lower().str.strip()
bidsheet_df['most_common_supplier_lower'] = bidsheet_df['Most common supplier'].astype(str).str.lower().str.strip()
mapping_dict = dict(zip(p21_df['p21_supplier_lower'], p21_df['Normalized to match bid supplier ']))
def get_normalized_supplier(mcs):
    if mcs in mapping_dict:
        return mapping_dict[mcs].strip() if isinstance(mapping_dict[mcs], str) else mapping_dict[mcs]
    else:
        return "-"
    
# Apply the mapping
normalized_incumbent_supplier = bidsheet_df['most_common_supplier_lower'].apply(get_normalized_supplier)

# Insert the new column next to "Most common supplier"
mcs_col_idx = bidsheet_df.columns.get_loc("Most common supplier")
bidsheet_df.insert(mcs_col_idx + 1, "Normalized incumbent supplier", normalized_incumbent_supplier)

# Optionally, drop the helper lowercase column if not needed anymore
bidsheet_df.drop(columns=['most_common_supplier_lower'], inplace=True)

bidsheet_df['Annual Volume (per UOM)'] = pd.to_numeric(bidsheet_df['Annual Volume (per UOM)'], errors='coerce')
bidsheet_df['Volume-banded WAPP'] = pd.to_numeric(bidsheet_df['Volume-banded WAPP'], errors='coerce')

# Calculate Extended Cost USD
bidsheet_df['Extended Cost USD'] = (bidsheet_df['Annual Volume (per UOM)'] * bidsheet_df['Volume-banded WAPP']).round(4)

# Move Extended Cost USD next to Volume-banded WAPP
wapp_idx = bidsheet_df.columns.get_loc("Volume-banded WAPP")
ext_cost = bidsheet_df.pop("Extended Cost USD")
bidsheet_df.insert(wapp_idx + 1, "Extended Cost USD", ext_cost)

# point 6 & 7 from the mail.
valid_supplier_idx = bidsheet_df.columns.get_loc("Valid Supplier")

def calculate_as_is_r1(row):
    vol_wapp = row["Volume-banded WAPP"]
    valid_sup = row["Valid Supplier"]
    norm_inc_supplier = row["Normalized incumbent supplier"]

    # Check Volume-banded WAPP and Valid Supplier non-zero and not NaN
    if pd.isna(vol_wapp) or vol_wapp == 0 or pd.isna(valid_sup) or valid_sup == 0:
        return "-"

    if not isinstance(norm_inc_supplier, str) or norm_inc_supplier.strip() == "":
        return "-"

    supplier_col = f"{norm_inc_supplier} - R1 - Total Cost Per UOM FOB Port of Origin/Departure (USD)"

    if supplier_col not in bidsheet_df.columns:
        print(f"Warning: Supplier column '{supplier_col}' not found in DataFrame for row {row.name}.")
        return "-"

    supplier_r1_cost = row[supplier_col]

    # Check supplier cost non-zero and not NaN
    if pd.isna(supplier_r1_cost) or supplier_r1_cost == 0:
        return "-"

    try:
        result = (vol_wapp - supplier_r1_cost) / vol_wapp
        return round(result, 4)  # rounded to 6 decimals, change if needed
    except Exception as e:
        print(e)
        return "-"
    
bidsheet_df.insert(valid_supplier_idx + 1, "As Is R1 %", bidsheet_df.apply(calculate_as_is_r1, axis=1))
# Get the index of "As Is R1 %" column (which was just inserted)
as_is_r1_pct_idx = bidsheet_df.columns.get_loc("As Is R1 %")

def calculate_as_is_r1_usd(row):
    vol_wapp = row["Volume-banded WAPP"]
    valid_sup = row["Valid Supplier"]
    as_is_r1_pct = row["As Is R1 %"]
    ext_cost_usd = row.get("Extended Cost USD", None)

    # Check required fields
    if pd.isna(vol_wapp) or vol_wapp == 0 or pd.isna(valid_sup) or valid_sup == 0:
        return "-"
    if as_is_r1_pct == "-" or pd.isna(as_is_r1_pct):
        return "-"
    if ext_cost_usd is None or pd.isna(ext_cost_usd):
        return "-"

    try:
        result = as_is_r1_pct * ext_cost_usd
        return round(result, 4)  # round as needed
    except Exception as e:
        print(e)
        return "-"

# Insert new column next to "As Is R1 %"
bidsheet_df.insert(as_is_r1_pct_idx + 1, "As Is R1 USD", bidsheet_df.apply(calculate_as_is_r1_usd, axis=1))
# AS IS USING R2
normalized_incumbent_supplier_idx = bidsheet_df.columns.get_loc("Normalized incumbent supplier")

def calculate_as_is_final(row):
    vol_wapp = row["Volume-banded WAPP"]
    valid_sup = row["Valid Supplier"]
    norm_inc_supplier = row["Normalized incumbent supplier"]

    # Check Volume-banded WAPP and Valid Supplier non-zero and not NaN
    if pd.isna(vol_wapp) or vol_wapp == 0 or pd.isna(valid_sup) or valid_sup == 0:
        return "-"

    if not isinstance(norm_inc_supplier, str) or norm_inc_supplier.strip() == "":
        return "-"

    supplier_col = f"{norm_inc_supplier} - R2 - Total Cost Per UOM FOB Port of Origin/Departure (USD)"

    if supplier_col not in bidsheet_df.columns:
        print(f"Warning: Supplier column '{supplier_col}' not found in DataFrame for row {row.name}.")
        return "-"

    supplier_r2_cost = row[supplier_col]

    # Check supplier cost non-zero and not NaN
    if pd.isna(supplier_r2_cost) or supplier_r2_cost == 0:
        return "-"

    try:
        result = (vol_wapp - supplier_r2_cost) / vol_wapp
        return round(result, 4)  # rounded to 6 decimals, change if needed
    except Exception as e:
        print(e)
        return "-"
    
bidsheet_df.insert(normalized_incumbent_supplier_idx + 1, "As Is Final %", bidsheet_df.apply(calculate_as_is_final, axis=1))
# Get the index of "As Is Final %" column (which was just inserted)
as_is_final_pct_idx = bidsheet_df.columns.get_loc("As Is Final %")

def calculate_as_is_final_usd(row):
    vol_wapp = row["Volume-banded WAPP"]
    valid_sup = row["Valid Supplier"]
    as_is_final_pct = row["As Is Final %"]
    ext_cost_usd = row.get("Extended Cost USD", None)

    # Check required fields
    if pd.isna(vol_wapp) or vol_wapp == 0 or pd.isna(valid_sup) or valid_sup == 0:
        return "-"
    if as_is_final_pct == "-" or pd.isna(as_is_final_pct):
        return "-"
    if ext_cost_usd is None or pd.isna(ext_cost_usd):
        return "-"

    try:
        result = as_is_final_pct * ext_cost_usd
        return round(result, 4)  # round as needed
    except Exception as e:
        print(e)
        return "-"

# Insert new column next to "As Is R1 %"
bidsheet_df.insert(as_is_final_pct_idx + 1, "As Is Final USD", bidsheet_df.apply(calculate_as_is_final_usd, axis=1))

bidsheet_df['Final Min Bid'] = pd.to_numeric(bidsheet_df['Final Min Bid'], errors='coerce')
bidsheet_df['Cherry Pick min Final %'] = ((bidsheet_df['Volume-banded WAPP'] - bidsheet_df['Final Min Bid']) / bidsheet_df['Volume-banded WAPP']).round(4)
bidsheet_df.loc[bidsheet_df['Volume-banded WAPP'] == 0, 'Cherry Pick min Final %'] = '-'
bidsheet_df['Cherry Pick min Final USD'] = (pd.to_numeric(bidsheet_df['Cherry Pick min Final %'], errors='coerce') * bidsheet_df['Extended Cost USD']).round(4)
bidsheet_df.loc[bidsheet_df['Cherry Pick min Final %'] == '-', 'Cherry Pick min Final USD'] = '-'

mbs_idx = bidsheet_df.columns.get_loc("Final Minimum Bid Supplier")
bidsheet_df.insert(mbs_idx, "Cherry Pick min Final %", bidsheet_df.pop("Cherry Pick min Final %"))
bidsheet_df.insert(mbs_idx+1, "Cherry Pick min Final USD", bidsheet_df.pop("Cherry Pick min Final USD"))

# cherry pick for min R1
min_bid_r1_idx = bidsheet_df.columns.get_loc("Min Bid R1")
bidsheet_df.insert(
    min_bid_r1_idx + 1,
    "Cherry Pick min R1 %",
    bidsheet_df.apply(
        lambda row: (
            (row['Volume-banded WAPP'] - row['Min Bid R1']) / row['Volume-banded WAPP']
        ) if (
            pd.notna(row['Volume-banded WAPP']) and row['Volume-banded WAPP'] != 0
            and pd.notna(row['Valid Supplier']) and row['Valid Supplier'] != 0
            and pd.notna(row['Min Bid R1'])
        ) else '-',
        axis=1
    )
)
# Find index of "Cherry Pick min R1 %" column
cherry_pick_min_r1_pct_idx = bidsheet_df.columns.get_loc("Cherry Pick min R1 %")

bidsheet_df.insert(
    cherry_pick_min_r1_pct_idx + 1,
    "Cherry Pick min R1 USD",
    bidsheet_df.apply(
        lambda row: (
            row["Cherry Pick min R1 %"] * row["Extended Cost USD"]
        ) if (
            pd.notna(row["Volume-banded WAPP"]) and row["Volume-banded WAPP"] != 0
            and pd.notna(row["Valid Supplier"]) and row["Valid Supplier"] != 0
            and pd.notna(row["Cherry Pick min R1 %"]) and row["Cherry Pick min R1 %"] != "-"
            and pd.notna(row["Extended Cost USD"])
        ) else "-",
        axis=1
    )
)

# Find index of "Cherry Pick min R1 USD" column
awardable_col_idx = bidsheet_df.columns.get_loc("Cherry Pick min R1 USD")

# Function to determine if a bid is awardable based on Cherry Pick min R1 %
def awardable_min_bid_r1(row):
    vol_wapp = row["Volume-banded WAPP"]
    valid_sup = row["Valid Supplier"]
    cherry_pick_pct = row["Cherry Pick min R1 %"]

    if pd.isna(vol_wapp) or vol_wapp == 0 or pd.isna(valid_sup) or valid_sup == 0:
        return "No baseline or bid"

    if cherry_pick_pct == "-" or pd.isna(cherry_pick_pct):
        return "No baseline or bid"

    try:
        if cherry_pick_pct > 0:
            return "Yes"
        else:
            return "No"
    except Exception:
        return "No baseline or bid"

# Insert new column next to "Cherry Pick min R1 USD" called "Awardable Min Bid R1 (+0% savings)"
bidsheet_df.insert(
    awardable_col_idx + 1,
    "Awardable Min Bid R1 (+0% savings)",
    bidsheet_df.apply(awardable_min_bid_r1, axis=1)
)

'''
add a column "Min improved R2 vs R1" right next to "Final Min Bid" with this logic:
If "Volume-banded WAPP" is missing or zero OR "Valid Supplier" is zero or missing → "No baseline or bid"
Else if "Final Min Bid" < "Min Bid R1" → "Yes"
Else if "Final Min Bid" ≥ "Min Bid R1" → "No"
'''
final_min_bid_idx = bidsheet_df.columns.get_loc("Final Min Bid")

def min_improved_r2_vs_r1(row):
    vol_wapp = row["Volume-banded WAPP"]
    valid_sup = row["Valid Supplier"]
    final_min_bid = row["Final Min Bid"]
    min_bid_r1 = row["Min Bid R1"]

    if pd.isna(vol_wapp) or vol_wapp == 0 or pd.isna(valid_sup) or valid_sup == 0:
        return "No baseline or bid"
    if pd.isna(final_min_bid) or pd.isna(min_bid_r1):
        return "No baseline or bid"

    try:
        if final_min_bid < min_bid_r1:
            return "Yes"
        else:
            return "No"
    except Exception:
        return "No baseline or bid"

bidsheet_df.insert(
    final_min_bid_idx + 1,
    "Min improved R2 vs R1",
    bidsheet_df.apply(min_improved_r2_vs_r1, axis=1)
)

'''
Then next to the Cherry Pick min Final USD can we add a column called "Awardable Min Bid Final (+0% savings)" 
and in it have the value be Yes if Cherry Pick min Final %  > 0%, No if <= 0% and "No baseline or bid" if there is no Volume Banded WAPP and/or Valid Supplier is 0

'''

cherry_pick_final_usd_idx = bidsheet_df.columns.get_loc("Cherry Pick min Final USD")

def awardable_min_bid_final(row):
    vol_wapp = row["Volume-banded WAPP"]
    valid_sup = row["Valid Supplier"]
    cherry_pick_final_pct = row.get("Cherry Pick min Final %", None)

    if pd.isna(vol_wapp) or vol_wapp == 0 or pd.isna(valid_sup) or valid_sup == 0:
        return "No baseline or bid"
    if cherry_pick_final_pct == "-" or pd.isna(cherry_pick_final_pct):
        return "No baseline or bid"

    try:
        return "Yes" if cherry_pick_final_pct > 0 else "No"
    except Exception:
        return "No baseline or bid"

bidsheet_df.insert(
    cherry_pick_final_usd_idx + 1,
    "Awardable Min Bid Final (+0% savings)",
    bidsheet_df.apply(awardable_min_bid_final, axis=1)
)


supplier_pattern = re.compile(r"^(.*?) - R([12]) - Total Cost Per UOM FOB Port of Origin/Departure \(USD\)$")
supplier_r1_map, supplier_r2_map = {}, {}

for col in bidsheet_df.columns[33:]:
    m = supplier_pattern.match(col)
    if m:
        name, round_num = m.group(1), m.group(2)
        (supplier_r1_map if round_num == '1' else supplier_r2_map)[name] = col

suppliers = sorted(set(supplier_r1_map) & set(supplier_r2_map))
missing_r1 = set(supplier_r2_map) - set(supplier_r1_map)
missing_r2 = set(supplier_r1_map) - set(supplier_r2_map)

if missing_r1 or missing_r2:
    print("Warning: Skipped suppliers due to missing R1/R2:")
    if missing_r1: print("  Missing R1:", missing_r1)
    if missing_r2: print("  Missing R2:", missing_r2)

supplier_new_cols = {}
supplier_column_order = []

for s in tqdm(suppliers, desc='Supplier Calcs'):
    r1_col = supplier_r1_map[s]
    r2_col = supplier_r2_map[s]
    r1 = pd.to_numeric(bidsheet_df[r1_col], errors='coerce')
    r2 = pd.to_numeric(bidsheet_df[r2_col], errors='coerce')
    
    wapp = bidsheet_df['Volume-banded WAPP']
    ext_cost = bidsheet_df['Extended Cost USD']

    r1_pct = ((wapp - r1) / wapp).where((r1 != 0) & wapp.ne(0)).round(4)
    r1_usd = (r1_pct * ext_cost).round(4)
    delta_pct = ((r1 - r2) / r1).where((r1 != 0) & (r2 != 0)).round(4)
    delta_usd = (delta_pct * ext_cost).round(4)
    final_pct = ((wapp - r2) / wapp).where((r2 !=0) & (wapp != 0)).round(4)
    final_usd = (final_pct * ext_cost).round(4)

    supplier_new_cols[f"{s} R1 - Savings vs baseline %"] = r1_pct
    supplier_new_cols[f"{s} R1 - Savings vs baseline USD"] = r1_usd
    supplier_new_cols[f"{s} - Delta R1 vs R2 %"] = delta_pct
    supplier_new_cols[f"{s} - Delta R1 vs R2 USD"] = delta_usd
    supplier_new_cols[f"{s} - Final % savings vs baseline"] = final_pct
    supplier_new_cols[f"{s} - Final USD savings vs baseline"] = final_usd

    supplier_column_order.extend([
        r1_col,
        f"{s} R1 - Savings vs baseline %",
        f"{s} R1 - Savings vs baseline USD",
        r2_col,
        f"{s} - Delta R1 vs R2 %",
        f"{s} - Delta R1 vs R2 USD",
        f"{s} - Final % savings vs baseline",
        f"{s} - Final USD savings vs baseline"
    ])

supplier_new_df = pd.DataFrame(supplier_new_cols)
bidsheet_df = pd.concat([bidsheet_df, supplier_new_df], axis=1).copy()

for col in bidsheet_df.columns:
    if bidsheet_df[col].dtype in [np.float64, np.float32, float]:
        bidsheet_df[col] = bidsheet_df[col].round(4)

# Reorder columns to enforce the supplier grouping order
pre_supplier_cols = list(bidsheet_df.columns[:35])
post_supplier_cols = [col for col in bidsheet_df.columns if col not in pre_supplier_cols and col not in supplier_column_order]
bidsheet_df = bidsheet_df[pre_supplier_cols + supplier_column_order + post_supplier_cols]

os.makedirs("new", exist_ok=True)
bidsheet_df.to_excel(output_file, index=False)

# wb = load_workbook(output_file)
# ws = wb.active
# header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

# # --- Define Fill Colors ---
# fill_purple = PatternFill(start_color="800080", end_color="800080", fill_type="solid")
# fill_red    = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
# fill_orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
# fill_green  = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
# fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
# fill_subtle_grey = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")  # subtle grey
# fill_subtle_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # subtle blue

# # --- Float formatting columns ---
# float_cols = [14, 15, 16, 17, 18, 21, 25, 26, 27, 29,30, 34] + list(range(36, len(header)+1))  # 1-based indices
# # Predefine substrings and last 5 column indices
# target_substrings = [
#     "R1 - Savings vs baseline %", 
#     "R1 - Savings vs baseline USD",
#     "- Delta R1 vs R2 %", 
#     "- Delta R1 vs R2 USD", 
#     "- Final % savings vs baseline", 
#     "- Final USD savings vs baseline"
# ]

# # Columns (1-based) matching any of the substrings
# special_col_indices = {
#     idx + 1 for idx, col in enumerate(header)
#     if any(substr in str(col) for substr in target_substrings)
# }

# # Combine logic: precompute which columns get '-' on empty
# dash_fill_cols = special_col_indices.union(set(range(len(header) - 4 + 1, len(header) + 1)))  # 1-based

# # Now do formatting
# for col_idx in tqdm(float_cols, desc='Float formatting'):
#     if col_idx > len(header):
#         continue

#     is_dash_fill_col = col_idx in dash_fill_cols

#     for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
#         cell = row[0]  # only one column

#         if isinstance(cell.value, (int, float)):
#             cell.number_format = '0.0000'
#         elif cell.value in [None, '']:
#             if is_dash_fill_col:
#                 cell.value = '-'
#             else:
#                 cell.value = 0
#                 cell.number_format = '0.0000'


# # --- Yellow fill for last 5 columns ---
# last_5_col_indices = range(len(header)-4, len(header)+1)  # 1-based
# for col_idx in tqdm(last_5_col_indices, desc='Yellow fill (last 5 cols)'):
#     for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
#         for cell in row:
#             cell.fill = fill_yellow

# # --- Color fill logic for 24+ columns with specific header ---
# from openpyxl.utils import get_column_letter
# from openpyxl.styles import PatternFill

# # Predefine fills only once
# fill_map = {
#     "purple": fill_purple,
#     "red": fill_red,
#     "orange": fill_orange,
#     "green": fill_green
# }

# # Cache WAPP values to avoid repeatedly accessing cells
# wapp_col_idx = header.index("Volume-banded WAPP") + 1  # 1-based
# wapp_values = {}
# for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
#     row_num = row[0].row
#     wapp_cell = row[wapp_col_idx - 1]
#     try:
#         wapp_values[row_num] = float(wapp_cell.value)
#     except (TypeError, ValueError):
#         wapp_values[row_num] = None  # Mark invalid

# # Pre-filter target columns once
# target_col_idxs = [i + 1 for i, h in enumerate(header) if "Total Cost Per UOM FOB Port of Origin/Departure (USD)" in str(h)]

# # Iterate once over rows and apply fill efficiently
# for row in tqdm(ws.iter_rows(min_row=2, max_row=ws.max_row), desc='Bid color fill logic'):
#     row_num = row[0].row
#     wapp = wapp_values.get(row_num)
#     if wapp in (None, 0):
#         continue

#     for col_idx in target_col_idxs:
#         bid_cell = row[col_idx - 1]
#         try:
#             bid = float(bid_cell.value)
#         except (TypeError, ValueError):
#             continue

#         if bid == 0:
#             continue

#         diff_ratio = (wapp - bid) / wapp

#         if diff_ratio < -0.40:
#             bid_cell.fill = fill_map["purple"]
#         elif -0.40 <= diff_ratio <= 0:
#             bid_cell.fill = fill_map["red"]
#         elif 0 < diff_ratio <= 0.40:
#             bid_cell.fill = fill_map["orange"]
#         elif diff_ratio > 0.40:
#             bid_cell.fill = fill_map["green"]

# # --- Optimized Header coloring for R1/R2 columns with tqdm ---
# fill_map = {
#     "R1 - Total Cost Per UOM FOB Port of Origin/Departure (USD)": fill_subtle_grey,
#     "R2 - Total Cost Per UOM FOB Port of Origin/Departure (USD)": fill_subtle_blue
# }

# for idx, col_header in tqdm(enumerate(header), total=len(header), desc="Coloring Headers"):
#     header_str = str(col_header)
#     for key, fill in fill_map.items():
#         if key in header_str:
#             ws.cell(row=1, column=idx+1).fill = fill
#             break  # Stop after first match

# # --- Green fill for 13th column (M) ---
# col_13_letter = get_column_letter(13)  # 'M' for 13th column

# for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=13, max_col=13):
#     cell = row[0]
#     cell.fill = fill_green

# wb.save(output_file)
print(f"\n✔ Done. Script run time: {time.time() - start_time:.2f} seconds")
