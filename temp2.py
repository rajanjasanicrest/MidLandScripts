import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# === Step 1: Read Excel ===
input_file = "combined_bidsheet_outlier_2.xlsx"
df = pd.read_excel(input_file)

# === Round non-zero numerical values to 4 decimal places ===
def round_number(val):
    try:
        if isinstance(val, (int, float)) and val != 0:
            return round(val, 4)
        return val
    except:
        return val

def format_number(val):
    try:
        if isinstance(val, (int, float)) and val != 0:
            return f"{val:.4f}"  # as string: '0.2700'
        return val
    except:
        return val

for col in df.select_dtypes(include=["number"]).columns:
    df[col] = df[col].map(format_number)

# === Step 2: Identify columns containing the keyword in header ===
target_cols = [col for col in df.columns if "-Total Cost" in str(col)]

# === Step 3: Compute IQR for each row and mark outliers ===
outlier_mask = pd.DataFrame(False, index=df.index, columns=target_cols)
iqr_values = []

for idx, row in df.iterrows():
    values = row[target_cols].dropna().astype(float)
    values = [value for value in values if value != 0]
    values = pd.Series(values)

    if len(values) <= 1:
        iqr_values.append(np.nan)
        continue
        
    Q1 = values.quantile(0.25)
    Q3 = values.quantile(0.75)
    IQR = Q3 - Q1
    iqr_values.append(IQR)

    lower = Q1 - 1.5 * IQR
    upper = Q3 + 1.5 * IQR

    for col in target_cols:
        val = row[col]
        if pd.notnull(float(val)) and (float(val) < lower or float(val) > upper):
            outlier_mask.at[idx, col] = True

# Add IQR to last column
df["IQR_Value"] = iqr_values

# === Step 4: Save with highlights ===
output_file = "output_with_outliers.xlsx"
df = df.fillna("")  # Replace all NaNs with empty string
df.to_excel(output_file, index=False)

# === Step 5: Apply Styling ===
wb = load_workbook(output_file)
ws = wb.active

# Define fill colors
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")        # Outliers
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")        # Last 4 columns
light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")   # 5th column
light_gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")    # Alternate headers
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")         # Alternate headers

headers = [cell.value for cell in ws[1]]

# Highlight outliers in orange
for i, col in enumerate(df.columns):
    if col in target_cols:
        col_letter = get_column_letter(i + 1)
        for j in range(len(df)):
            if outlier_mask.iloc[j, target_cols.index(col)] and df.iloc[j, i] != 0:
                ws[f"{col_letter}{j + 2}"].fill = orange_fill

# === Apply column styling ===

# 1. Last 4 columns → Yellow
last_4_col_indices = range(len(headers) - 4, len(headers))
for col_idx in last_4_col_indices:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx + 1, max_col=col_idx + 1):
        for cell in row:
            cell.fill = yellow_fill

# 2. 5th column (index 4) → Light Green
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=5, max_col=5):
    for cell in row:
        cell.fill = light_green_fill

# 3. From 7th column onward → alternate header colors by company
company_to_columns = {}
for i, col_name in enumerate(headers[6:], start=6):
    company = str(col_name).split("-")[0].strip()
    company_to_columns.setdefault(company, []).append(i)

# Apply alternating fill to grouped headers
fills = [light_gray_fill, white_fill]
for idx, (company, cols) in enumerate(company_to_columns.items()):
    fill = fills[idx % 2]
    for col_idx in cols:
        col_letter = get_column_letter(col_idx + 1)
        ws[f"{col_letter}1"].fill = fill

# 4. Right-align data from 7th column onward, excluding "Additional Information" columns
for col_idx, header in enumerate(headers[6:], start=6):
    if "Additional information" not in str(header):
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx + 1, max_col=col_idx + 1):
            for cell in row:
                cell.alignment = Alignment(horizontal="right")

# === Save file ===
wb.save(output_file)
print(f"✅ Done! Output saved with outliers and styling applied: {output_file}")
