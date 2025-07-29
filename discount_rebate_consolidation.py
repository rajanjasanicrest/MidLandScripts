import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from tqdm import tqdm
from collections import defaultdict
from handler.handler import extract_supplier_name

# All Folder location information 
FILES_FOLDER_LOCATION = "./files round 2"
CLEANED_FILES_FOLDER_LOCATION = "./cleaned_files/discount_rebate"
CONSOLIDATED_FILE_LOCATION = "./new"
CONSOLIDATE_FILE_NAME = "discount_rebate_consolidate"

# OUTPUT Files related information 
OUTPUT_FILE = f"{CONSOLIDATED_FILE_LOCATION}/{CONSOLIDATE_FILE_NAME}.xlsx"

# Define sheet names that will be processed
SHEET_NAMES = ["payment_terms", "discount", "bonus_rebate"]

# --------------------------------- Payment terms related columns ---------------------------------------------  #
PAYMENT_TERMS_COMMON_COLUMNS = ["Payment Terms"]
PAYMENT_TERMS_SUPPLIER_COLUMNS = ["% Discount off invoice amount"]

# ----------------------------- Discount related columns -------------------------------------------- # 
DISCOUNT_COMMON_COLUMNS = ["Annual Revenue  Requirement in 1,000 USD"]
DISCOUNT_SUPPLIER_COLUMNS = ["% Discount off EXW Price"]

#  ---------------------- BONUS related columns ------------------------------------- #
BONUS_COMMON_COLUMNS = ["Midland long-term commitment"]
BONUS_SUPPLIER_COLUMNS = ["Bonus/rebate as % of first year spend"]

def format_decimal_value(value):
    """
    Format numeric values to have maximum 4 decimal places
    """
    if pd.isna(value):
        return value
    
    try:
        # Convert to float if it's a number
        if isinstance(value, (int, float)):
            num_value = float(value)
            # Check if it's a whole number
            if num_value == int(num_value):
                return int(num_value)
            else:
                # Round to 4 decimal places and remove trailing zeros
                formatted = round(num_value, 4)
                return formatted
        
        # If it's a string that represents a number
        elif isinstance(value, str):
            # Try to convert string to number
            try:
                num_value = float(value)
                # Check if it's a whole number
                if num_value == int(num_value):
                    return str(int(num_value))
                else:
                    # Round to 4 decimal places
                    formatted = round(num_value, 4)
                    return str(formatted)
            except ValueError:
                # If conversion fails, return original string
                return value
        
        return value
    except:
        return value

def format_dataframe_decimals(df):
    """
    Apply decimal formatting to all numeric columns in a dataframe
    """
    if df is None or df.empty:
        return df
    
    df_copy = df.copy()
    
    # Apply formatting to each cell
    for col in df_copy.columns:
        df_copy[col] = df_copy[col].apply(format_decimal_value)
    
    return df_copy

def cleanup_files():
    """
    Process raw files and extract specific sections, then save them as separate sheets in cleaned Excel files
    """
    # Read all files in the directory
    files = os.listdir(FILES_FOLDER_LOCATION)
    files2 = os.listdir("./files")

    all_files = files + files2
    
    for item in all_files:
        # if not item.endswith(('.xlsx', '.xls')):
        #     continue
            
        print("Processing excel sheet --------------------------------------------")
        print(f"./files round 2/{item}")

        cleaned_excel_file_name = f"{item.split('.')[0]}_cleaned.xlsx"
        cleaned_file_path = os.path.join(CLEANED_FILES_FOLDER_LOCATION, cleaned_excel_file_name)
        os.makedirs(CLEANED_FILES_FOLDER_LOCATION, exist_ok=True)

        try:
            if item in files:
                df = pd.read_excel(f"{FILES_FOLDER_LOCATION}/{item}", sheet_name="5. Disc, cond & rebate bidsheet")
            else:
                df = pd.read_excel(f"./files/{item}", sheet_name="5. Disc, cond & rebate bidsheet")

        except Exception as e:
            print(f"Failed to read sheet from {item}: {e}")
            continue

        # Create a new workbook for the cleaned file
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Payment terms related data extraction -----------------------------------------------------------------------
        payment_terms_df = extract_payment_terms_data(df)
        if payment_terms_df is not None and not payment_terms_df.empty:
            # Apply decimal formatting
            payment_terms_df = format_dataframe_decimals(payment_terms_df)
            ws_payment = wb.create_sheet(title="payment_terms")
            for r in dataframe_to_rows(payment_terms_df, index=False, header=True):
                ws_payment.append(r)
            print(f"✅ Payment terms data saved: {len(payment_terms_df)} rows")
        else:
            print("❌ No payment terms data found")

        # Discount related data extraction ----------------------------------------------------------------------
        discount_df = extract_discount_data(df)
        if discount_df is not None and not discount_df.empty:
            # Apply decimal formatting
            discount_df = format_dataframe_decimals(discount_df)
            ws_discount = wb.create_sheet(title="discount")
            for r in dataframe_to_rows(discount_df, index=False, header=True):
                ws_discount.append(r)
            print(f"✅ Discount data saved: {len(discount_df)} rows")
        else:
            print("❌ No discount data found")

        # Bonus/Rebate related data extraction --------------------------------------------------------------------------
        bonus_df = extract_bonus_rebate_data(df)
        if bonus_df is not None and not bonus_df.empty:
            # Apply decimal formatting
            bonus_df = format_dataframe_decimals(bonus_df)
            ws_bonus = wb.create_sheet(title="bonus_rebate")
            for r in dataframe_to_rows(bonus_df, index=False, header=True):
                ws_bonus.append(r)
            print(f"✅ Bonus/rebate data saved: {len(bonus_df)} rows")
        else:
            print("❌ No bonus/rebate data found")

        # Save the cleaned Excel file only if we have at least one sheet with data
        if len(wb.sheetnames) > 0:
            wb.save(cleaned_file_path)
            print(f"✅ Cleaned file saved: {cleaned_file_path}")
        else:
            print(f"❌ No data extracted from {item}, skipping file creation")

def extract_payment_terms_data(df):
    """Extract payment terms data from the dataframe"""
    try:
        lead_header_start_row_index = None
        lead_header_end_row_index = None

        # Find Payment Terms section
        for index, row in df.iterrows():
            for col_index, value in enumerate(row):
                if isinstance(value, str) and "Payment Terms" in value:
                    lead_header_start_row_index = index
                    break
            if lead_header_start_row_index is not None:
                break

        # Find end of Payment Terms section (where Discounts section starts)
        if lead_header_start_row_index is not None:
            for index, row in df.iterrows():
                if index > lead_header_start_row_index:
                    for col_index, value in enumerate(row):
                        if isinstance(value, str) and "Discounts" in value:
                            lead_header_end_row_index = index
                            break
                if lead_header_end_row_index is not None:
                    break

        if lead_header_start_row_index is not None and lead_header_end_row_index is not None:
            # Extract data from columns 1 to 4 (0-indexed)
            payment_terms_df = df.iloc[lead_header_start_row_index:lead_header_end_row_index, 1:5].copy()
            
            # Clean up the dataframe
            payment_terms_df = payment_terms_df.dropna(how='all')  # Remove completely empty rows
            payment_terms_df.columns = range(len(payment_terms_df.columns))  # Reset column names to numbers
            
            return payment_terms_df
        else:
            return None
            
    except Exception as e:
        print(f"Error extracting payment terms data: {e}")
        return None

def extract_discount_data(df):
    """Extract discount data from the dataframe"""
    try:
        discount_start_row_index = None

        # Find Discount section
        for index, row in df.iterrows():
            for col_index, value in enumerate(row):
                if isinstance(value, str) and "Annual Revenue  Requirement in 1,000 USD" in value:
                    discount_start_row_index = index
                    break
            if discount_start_row_index is not None:
                break

        # Find end of Discount section (where next section starts or end of data)
        discount_end_row_index = None
        if discount_start_row_index is not None:
            for index, row in df.iterrows():
                if index > discount_start_row_index + 10:  # Look ahead reasonable number of rows
                    for col_index, value in enumerate(row):
                        if isinstance(value, str) and "Midland long-term commitment" in value:
                            discount_end_row_index = index
                            break
                if discount_end_row_index is not None:
                    break
            
            # If no end found, take reasonable number of rows
            if discount_end_row_index is None:
                discount_end_row_index = min(discount_start_row_index + 25, len(df))

        if discount_start_row_index is not None:
            # Extract data from columns 1 to 4 (0-indexed)
            discount_df = df.iloc[discount_start_row_index:discount_end_row_index, 1:5].copy()
            
            # Clean up the dataframe
            discount_df = discount_df.dropna(how='all')  # Remove completely empty rows
            discount_df.columns = range(len(discount_df.columns))  # Reset column names to numbers
            
            return discount_df
        else:
            return None
            
    except Exception as e:
        print(f"Error extracting discount data: {e}")
        return None

def extract_bonus_rebate_data(df):
    """Extract bonus/rebate data from the dataframe"""
    try:
        bonus_start_row_index = None

        # Find Bonus/Rebate section
        for index, row in df.iterrows():
            for col_index, value in enumerate(row):
                if isinstance(value, str) and "Midland long-term commitment" in value:
                    bonus_start_row_index = index
                    break
            if bonus_start_row_index is not None:
                break

        if bonus_start_row_index is not None:
            # Extract 4 rows (1 header + 3 data rows) starting from the found row
            bonus_end_row_index = bonus_start_row_index + 4
            
            # Get all columns from the start column to the end
            start_col_index = 0
            for col_index, value in enumerate(df.iloc[bonus_start_row_index]):
                if isinstance(value, str) and "Midland long-term commitment" in value:
                    start_col_index = col_index
                    break
            
            bonus_df = df.iloc[bonus_start_row_index:bonus_end_row_index, start_col_index:].copy()
            
            # Clean up the dataframe
            bonus_df = bonus_df.dropna(how='all', axis=1)  # Remove completely empty columns
            bonus_df = bonus_df.dropna(how='all')  # Remove completely empty rows
            bonus_df.columns = range(len(bonus_df.columns))  # Reset column names to numbers
            
            return bonus_df
        else:
            return None
            
    except Exception as e:
        print(f"Error extracting bonus/rebate data: {e}")
        return None

def get_sheet_column_config(sheet_name):
    """
    Get the appropriate column configuration for each sheet type
    """
    if sheet_name == "payment_terms":
        return PAYMENT_TERMS_COMMON_COLUMNS, PAYMENT_TERMS_SUPPLIER_COLUMNS
    elif sheet_name == "discount":
        return DISCOUNT_COMMON_COLUMNS, DISCOUNT_SUPPLIER_COLUMNS
    elif sheet_name == "bonus_rebate":
        return BONUS_COMMON_COLUMNS, BONUS_SUPPLIER_COLUMNS
    else:
        return [], []

def is_header_row(row, common_columns, supplier_columns):
    """
    Check if a row is a header row by comparing its values with expected column names
    """
    row_values = [str(val).strip().lower() if pd.notna(val) else "" for val in row]
    expected_headers = [col.strip().lower() for col in common_columns + supplier_columns]
    
    # Check if any of the expected headers match any cell in the row
    for header in expected_headers:
        if header and any(header in cell_val for cell_val in row_values if cell_val):
            return True
    
    # Additional check for common header patterns
    header_patterns = ["payment terms", "discount", "rebate", "annual revenue", "midland long-term"]
    for pattern in header_patterns:
        if any(pattern in cell_val for cell_val in row_values if cell_val):
            return True
    
    return False

def create_consolidated_dataset_rowwise(all_data, common_columns, supplier_columns):
    """
    Create a consolidated dataset in row-wise format
    Each row will contain: Supplier Name, Common Columns, Supplier Columns
    """
    consolidated_rows = []
    
    for file_name, df in all_data.items():
        print(f"Processing data from {file_name}...")
        
        if df is None or df.empty:
            print(f"No data found in {file_name}, skipping...")
            continue
        
        # Extract supplier name
        try:
            supplier_name = file_name.split('--')[-1].strip()
            supplier_name = supplier_name.replace("_cleaned", "").replace(" R2", "").strip()
        except:
            supplier_name = file_name.replace("_cleaned", "")
        
        # Filter out header rows
        data_rows = []
        for idx, row in df.iterrows():
            if not is_header_row(row, common_columns, supplier_columns):
                # Also check if the row has meaningful data (not all empty or NaN)
                row_values = [val for val in row if pd.notna(val) and str(val).strip()]
                if row_values:  # Only include rows with some data
                    data_rows.append(row)
        
        print(f"Found {len(data_rows)} data rows (after filtering {len(df) - len(data_rows)} header/empty rows)")
        
        for row in data_rows:
            # Create a single row with supplier name, common data, and supplier data
            consolidated_row = {}
            
            # Add supplier name as first column
            consolidated_row['Supplier Name'] = supplier_name
            
            # Add common columns data
            for i, col in enumerate(common_columns):
                # Use column index if column name not found
                if col in df.columns:
                    value = row[col] if pd.notna(row[col]) else ""
                elif i < len(df.columns):
                    value = row.iloc[i] if pd.notna(row.iloc[i]) else ""
                else:
                    value = ""
                
                # Apply decimal formatting
                consolidated_row[col] = format_decimal_value(value)
            
            # Add supplier-specific columns data
            for i, col in enumerate(supplier_columns):
                # Use column index offset by number of common columns
                col_index = i + len(common_columns)
                if col in df.columns:
                    value = row[col] if pd.notna(row[col]) else ""
                elif col_index < len(df.columns):
                    value = row.iloc[col_index] if pd.notna(row.iloc[col_index]) else ""
                else:
                    value = ""
                
                # Apply decimal formatting
                consolidated_row[col] = format_decimal_value(value)
            
            consolidated_rows.append(consolidated_row)
    
    return consolidated_rows

def create_combined_consolidated_excel_rowwise(all_sheet_data, output_file):
    """
    Create a single Excel file with multiple sheets for all sheet types in row-wise format
    """
    wb = Workbook()
    
    # Remove the default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Define styles
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
    blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # Light blue
    wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    header_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
    new_supplier_fill = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')  # Light bluish
    wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    
    for sheet_name, sheet_info in all_sheet_data.items():
        if not sheet_info['consolidated_data']:
            print(f"No data for sheet: {sheet_name}, skipping...")
            continue
            
        print(f"Creating sheet: {sheet_name}")
        
        # Create new worksheet
        ws = wb.create_sheet(title=sheet_name)
        
        consolidated_data = sheet_info['consolidated_data']
        common_columns = sheet_info['common_columns']
        supplier_columns = sheet_info['supplier_columns']
        
        # Create headers
        headers = ['Supplier Name'] + common_columns + supplier_columns
        
        # Add headers to first row
        for col_num, header in enumerate(headers, 1):
            header_cell = ws.cell(row=1, column=col_num, value=header)
            header_cell.alignment = wrap_alignment
            header_cell.fill = header_fill

        previous_supplier = None
        
        # Add data rows
        for row_num, row_data in enumerate(consolidated_data, 2):  # Start from row 2
            current_supplier = row_data.get("Supplier Name", "")
            is_new_supplier = current_supplier != previous_supplier

            for col_num, header in enumerate(headers, 1):
                cell_value = row_data.get(header, "")
                # Apply decimal formatting to cell value
                formatted_value = format_decimal_value(cell_value)
                
                data_cell = ws.cell(row=row_num, column=col_num, value=formatted_value)
                data_cell.alignment = Alignment(wrap_text=True, vertical='top')

                if is_new_supplier: 
                    data_cell.fill = new_supplier_fill 

            previous_supplier = current_supplier 

        # Set column widths and row heights
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 20  # Fixed width of 20
        
        # Set row heights to accommodate wrapped text
        for row_num in range(1, ws.max_row + 1):
            ws.row_dimensions[row_num].height = 30
    
    # Save the workbook
    wb.save(output_file)
    print(f"Combined consolidated file saved: {output_file}")

def process_cleaned_files(): 
    """
    Process all cleaned files and create a single consolidated Excel file with multiple sheets in row-wise format
    """
    try:
        # Make sure consolidated folder exists
        os.makedirs(CONSOLIDATED_FILE_LOCATION, exist_ok=True)

        if not os.path.exists(CLEANED_FILES_FOLDER_LOCATION):
            print(f"Cleaned files folder does not exist: {CLEANED_FILES_FOLDER_LOCATION}")
            return

        cleaned_files = os.listdir(CLEANED_FILES_FOLDER_LOCATION)

        # Group files by supplier name (normalized)
        supplier_files = defaultdict(dict)
        used_r1 = []
        used_r2 = []

        for f in cleaned_files:
            if not f.endswith(('.xlsx', '.xls')):
                continue

            base = f.replace("_cleaned.xlsx", "")
            is_r2 = "R2" in base

            # Normalize supplier key (remove " R2" if present)
            supplier_key = base.replace(" R2", "")

            if is_r2:
                supplier_files[supplier_key]['R2'] = f
            else:
                supplier_files[supplier_key]['R1'] = f

        # Prefer R2 if available
        excel_files = []
        for supplier, files in supplier_files.items():
            if files.get('R2'):
                excel_files.append(files['R2'])
                used_r2.append(supplier)
            elif files.get('R1'):
                excel_files.append(files['R1'])
                used_r1.append(supplier)

        print(f"\n✔️ Total unique suppliers considered: {len(supplier_files)}")
        print(f"   - From R2: {len(used_r2)} suppliers")
        print(f"   - From R1 (fallback): {len(used_r1)} suppliers")
        missing_suppliers = set(supplier_files.keys()) - set(used_r1) - set(used_r2)
        if missing_suppliers:
            print(f"⚠️ Warning: {len(missing_suppliers)} suppliers had neither R1 nor R2: {missing_suppliers}")

        # Dictionary to store all sheet data
        all_sheet_data = {}

        # Process each sheet type separately
        for sheet_name in SHEET_NAMES:
            print(f"\n{'='*60}")
            print(f"Processing sheet type: {sheet_name}")
            print(f"{'='*60}")

            # Get column configuration for this sheet type
            common_columns, supplier_columns = get_sheet_column_config(sheet_name)

            if not common_columns and not supplier_columns:
                print(f"No column configuration found for sheet: {sheet_name}")
                continue

            # Dictionary to store all data from files for this sheet type
            all_data = {}

            for item in tqdm(excel_files, desc=f"Reading {sheet_name}"):
                file_path = os.path.join(CLEANED_FILES_FOLDER_LOCATION, item)

                try: 
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    df = format_dataframe_decimals(df)
                    file_name_key = os.path.splitext(item)[0]  # Remove extension
                    all_data[file_name_key] = df
                except Exception as e: 
                    print(f"Error processing {item} - {sheet_name}: {str(e)}")
                    all_data[os.path.splitext(item)[0]] = None
                    continue

            if not any(df is not None and not df.empty for df in all_data.values()):
                print(f"No valid data found for sheet type: {sheet_name}")
                all_sheet_data[sheet_name] = {
                    'consolidated_data': [],
                    'common_columns': common_columns,
                    'supplier_columns': supplier_columns
                }
                continue

            print(f"\nCreating consolidated dataset for {sheet_name} in row-wise format...")
            consolidated_data = create_consolidated_dataset_rowwise(all_data, common_columns, supplier_columns)

            all_sheet_data[sheet_name] = {
                'consolidated_data': consolidated_data,
                'common_columns': common_columns,
                'supplier_columns': supplier_columns
            }

            print(f"Total rows created for {sheet_name}: {len(consolidated_data)}")

        output_file = f"{CONSOLIDATED_FILE_LOCATION}/{CONSOLIDATE_FILE_NAME}.xlsx"
        print(f"\n{'='*60}")
        print(f"Creating combined consolidated Excel file: {output_file}")
        print(f"{'='*60}")

        create_combined_consolidated_excel_rowwise(all_sheet_data, output_file)

        print(f"\n{'='*60}")
        print("Combined consolidated file created successfully!")
        print(f"File saved: {output_file}")
        print(f"Sheets included: {list(all_sheet_data.keys())}")
        for sheet_name, sheet_info in all_sheet_data.items():
            row_count = len(sheet_info['consolidated_data'])
            print(f"  - {sheet_name}: {row_count} rows")
        print(f"{'='*60}")

    except Exception as e: 
        print(f"Error in process_cleaned_files: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    print("Starting file cleanup process...")
    # cleanup_files()

    print("\nStarting consolidation process...")
    process_cleaned_files()
