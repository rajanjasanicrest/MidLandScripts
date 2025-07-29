import pandas as pd
import os 
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from handler.handler import extract_supplier_name

# All Folder location information 
FILES_FOLDER_LOCATION = "./files round 2"
CLEANED_FILES_FOLDER_LOCATION = "./cleaned_files/new_product_intro"
CONSOLIDATED_FILE_LOCATION = "./new"
CONSOLIDATE_FILE_NAME = "new_product_intro_consolidate"

# Column name related information 
CONSOLIDATED_COMMON_COLUMN_NAMES = ["N#"]
SUPPLIER_COLUMNS_NAMES = ["Product type", "Range/series description", 
    "Price range offer i.e. lowest priced part in range through highest priced part (USD/each)", 
]

# OUTPUT Files related information 
OUTPUT_FILE = f"{CONSOLIDATED_FILE_LOCATION}/{CONSOLIDATE_FILE_NAME}.xlsx"

def format_decimal_value(value):
    """
    Format numeric values to have maximum 4 decimal places
    """
    if pd.isna(value):
        return value
    
    try:
        # Convert to float if it's a number
        if isinstance(value, (int, float)):
            num_value = float(value)
            # Check if it's a whole number
            if num_value == int(num_value):
                return int(num_value)
            else:
                # Round to 4 decimal places and remove trailing zeros
                formatted = round(num_value, 4)
                return formatted
        
        # If it's a string that represents a number
        elif isinstance(value, str):
            # Try to convert string to number
            try:
                num_value = float(value)
                # Check if it's a whole number
                if num_value == int(num_value):
                    return str(int(num_value))
                else:
                    # Round to 4 decimal places
                    formatted = round(num_value, 4)
                    return str(formatted)
            except ValueError:
                # If conversion fails, return original string
                return value
        
        return value
    except:
        return value

def format_dataframe_decimals(df):
    """
    Apply decimal formatting to all numeric columns in a dataframe
    """
    if df is None or df.empty:
        return df
    
    df_copy = df.copy()
    
    # Apply formatting to each cell
    for col in df_copy.columns:
        df_copy[col] = df_copy[col].apply(format_decimal_value)
    
    return df_copy

def cleanup_files():
    """
    Process raw files and extract specific sections, then save them as cleaned Excel files
    """
    # Read all files in the directory
    files = os.listdir(FILES_FOLDER_LOCATION)
    files2 = os.listdir("./files")

    all_files = files + files2
    
    for item in all_files: 
        # if not item.endswith(('.xlsx', '.xls')):
        #     continue

        print("Processing excel sheet ---------------------------------------")
        print(f"./files round 2/{item}")

        # Cleaned csv file name 
        cleaned_csv_file_name = f"{item.split('.')[0]}_cleaned.xlsx"

        # Make sure folder exists or not 
        os.makedirs(CLEANED_FILES_FOLDER_LOCATION, exist_ok=True)

        try:
            if item in files:
                df = pd.read_excel(f"{FILES_FOLDER_LOCATION}/{item}", sheet_name="6. New Product Intro Sharing")
            else:
                df = pd.read_excel(f"./files/{item}", sheet_name="6. New Product Intro Sharing")

        except Exception as e:
            print(f"Failed to read sheet from {item}: {e}")
            continue

        # Find the row containing "N#"
        header_row_index = None
        col_start_index = None

        for index, row in df.iterrows():
            for col_index, value in enumerate(row):
                if isinstance(value, str) and "N#" in value:
                    header_row_index = index
                    col_start_index = col_index
                    break
            if header_row_index is not None:
                break

        if header_row_index is not None and col_start_index is not None:
            # Set new headers starting from the found column
            headers = df.iloc[header_row_index, col_start_index:].tolist()
            
            # Skip one row after header
            data_rows = df.iloc[header_row_index + 2:, col_start_index:]
            
            # Set proper headers
            data_rows.columns = headers
            data_rows.reset_index(drop=True, inplace=True)
            
            # Apply decimal formatting
            data_rows = format_dataframe_decimals(data_rows)
            
            # Save to Excel
            data_rows.to_excel(f"{CLEANED_FILES_FOLDER_LOCATION}/{cleaned_csv_file_name}", index=False)
            print(f"Data saved to '{CLEANED_FILES_FOLDER_LOCATION}/{cleaned_csv_file_name}'")
        else:
            print("No row containing 'N#' was found.")

def create_consolidated_dataset_rowwise(all_data):
    """
    Create a consolidated dataset in row-wise format
    Each row will contain: Supplier Name, Common Columns, Supplier Columns
    """
    consolidated_rows = []
    
    for file_name, df in all_data.items():
        print(f"Processing data from {file_name}...")
        
        if df is None or df.empty:
            print(f"No data found in {file_name}, skipping...")
            continue
        
        # Extract supplier name
        try:
            supplier_name = file_name.split('--')[-1].strip()
            supplier_name = supplier_name.replace("_cleaned", "").replace(" R2", "").strip()
        except:
            supplier_name = file_name.replace("_cleaned", "")
        
        for _, row in df.iterrows():
            # Create a single row with supplier name, common data, and supplier data
            consolidated_row = {}
            
            # Add supplier name as first column
            consolidated_row['Supplier Name'] = supplier_name
            
            # Add common columns data
            for col in CONSOLIDATED_COMMON_COLUMN_NAMES:
                if col in df.columns:
                    value = row[col] if pd.notna(row[col]) else ""
                else:
                    value = ""
                
                # Apply decimal formatting
                consolidated_row[col] = format_decimal_value(value)
            
            # Add supplier-specific columns data
            for col in SUPPLIER_COLUMNS_NAMES:
                if col in df.columns:
                    value = row[col] if pd.notna(row[col]) else ""
                else:
                    value = ""
                
                # Apply decimal formatting
                consolidated_row[col] = format_decimal_value(value)
            
            consolidated_rows.append(consolidated_row)
    
    return consolidated_rows

def create_consolidated_excel_rowwise(consolidated_data):
    """
    Create Excel file with row-wise structure.
    Highlight new supplier rows with a light background color.
    """
    # Make sure consolidated folder exists
    os.makedirs(CONSOLIDATED_FILE_LOCATION, exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "New Product Intro"
    
    # Styles
    header_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
    new_supplier_fill = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')  # Light blueish
    wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    headers = ['Supplier Name'] + CONSOLIDATED_COMMON_COLUMN_NAMES + SUPPLIER_COLUMNS_NAMES
    
    # Add headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.alignment = wrap_alignment
        cell.fill = header_fill

    previous_supplier = None

    # Add rows with background highlight on new supplier
    for row_num, row_data in enumerate(consolidated_data, 2):  # Start from second row
        current_supplier = row_data.get("Supplier Name", "")
        is_new_supplier = current_supplier != previous_supplier

        for col_num, header in enumerate(headers, 1):
            cell_value = row_data.get(header, "")
            # Apply decimal formatting to cell value
            formatted_value = format_decimal_value(cell_value)
            
            cell = ws.cell(row=row_num, column=col_num, value=formatted_value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if is_new_supplier:
                cell.fill = new_supplier_fill

        previous_supplier = current_supplier
    
    # Set column widths and row heights
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 30

    wb.save(OUTPUT_FILE)

def process_cleaned_files(): 
    """
    Function information: 
        Apply for all process cleaned files in row-wise format
    """
    try: 
        if not os.path.exists(CLEANED_FILES_FOLDER_LOCATION):
            print(f"Cleaned files folder does not exist: {CLEANED_FILES_FOLDER_LOCATION}")
            return
        
        # Reading cleaned files related csv files 
        from collections import defaultdict

        cleaned_files = os.listdir(CLEANED_FILES_FOLDER_LOCATION)

        # Group files by supplier name (normalized)
        supplier_files = defaultdict(dict)
        used_r1 = []
        used_r2 = []

        for f in cleaned_files:
            if not f.endswith(('.xlsx', '.xls')):
                continue

            base = f.replace("_cleaned.xlsx", "")
            is_r2 = "R2" in base

            # Normalize supplier key (remove " R2" if present)
            supplier_key = base.replace(" R2", "")

            if is_r2:
                supplier_files[supplier_key]['R2'] = f
            else:
                supplier_files[supplier_key]['R1'] = f

        # Prefer R2 if available
        excel_files = []
        for supplier, files in supplier_files.items():
            if files.get('R2'):
                excel_files.append(files['R2'])
                used_r2.append(supplier)
            elif files.get('R1'):
                excel_files.append(files['R1'])
                used_r1.append(supplier)

        print(f"\n✔️ Total unique suppliers considered: {len(supplier_files)}")
        print(f"   - From R2: {len(used_r2)} suppliers")
        print(f"   - From R1 (fallback): {len(used_r1)} suppliers")
        missing_suppliers = set(supplier_files.keys()) - set(used_r1) - set(used_r2)
        if missing_suppliers:
            print(f"⚠️ Warning: {len(missing_suppliers)} suppliers had neither R1 nor R2: {missing_suppliers}")


        # Dictionary to store all data from files
        all_data = {}

        # Process each cleaned file
        for item in excel_files:

            print("Processing cleaned file ---------------------------------------")
            file_path = os.path.join(CLEANED_FILES_FOLDER_LOCATION, item)
            print(f"Processing: {file_path}")
            
            try:
                # Read the Excel file
                df = pd.read_excel(file_path)
                
                # Apply decimal formatting to the dataframe
                df = format_dataframe_decimals(df)
                
                # Store the dataframe for later use
                file_name_key = os.path.splitext(item)[0]  # Remove extension
                all_data[file_name_key] = df
                
                print(f"Successfully loaded {len(df)} rows from {item}")
                print(f"Columns in file: {list(df.columns)}")
                
            except Exception as e:
                print(f"Error processing {item}: {str(e)}")
                continue
        
        if not all_data:
            print("No valid data found in any files!")
            return
        
        # Create consolidated dataset in row-wise format
        print("\nCreating consolidated dataset in row-wise format...")
        consolidated_data = create_consolidated_dataset_rowwise(all_data)
        
        # Create Excel file with row-wise structure
        print("Creating consolidated Excel file...")
        create_consolidated_excel_rowwise(consolidated_data)
        
        print(f"\nConsolidated file created: {OUTPUT_FILE}")
        print(f"Total rows in consolidated file: {len(consolidated_data)}")
        
    except Exception as e: 
        print(f"Error in main process: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    # First clean up the raw files
    print("Starting file cleanup process...")
    # cleanup_files()
    
    print("\nStarting consolidation process...")
    process_cleaned_files()