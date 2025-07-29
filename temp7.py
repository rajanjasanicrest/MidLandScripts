# Updated script based on the user request
import re
import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment

# === Config ===
CONSOLIDATED_COMMON_COLUMN_NAMES = [
    "ROW ID #", "Division", 'Part #', "Item Description", "Product Group",
    "Part Family", "Average Order Quantity (per UOM)", "Min Order Quantity (per UOM)",
    "Max Order Quantity (per UOM)", "Order frequency", "Annual Volume (per UOM)"
]

# === Step 1: Read and combine files with multi-header (2 rows) ===
file_info = [
    ("consolidate/bidsheet_steel_outlier_consolidate.xlsx", "steel"),
    ("consolidate/bidsheet_brass_outlier_consolidate.xlsx", "brass"),
    ("consolidate/bidsheet_other_metal_outlier_consolidate.xlsx", "other metal")
]

dfs = []

for file_path, metal_type in file_info:
    df = pd.read_excel(file_path, header=[0, 1])
    new_cols = []
    seen = {}
    for col in df.columns:
        lvl1, lvl2 = col
        if lvl1.startswith("Unnamed") or pd.isna(lvl1):
            new_cols.append(("", lvl2))
        else:
            base = lvl1.strip()
            # Only add suffix if duplicate
            if base in seen:
                seen[base] += 1
                unique_lvl1 = f"{base}_{seen[base]}"
            else:
                seen[base] = 1
                unique_lvl1 = base
            new_cols.append((unique_lvl1, lvl2))
    df.columns = pd.MultiIndex.from_tuples(new_cols)
    df.insert(5, ('', 'type'), metal_type)
    dfs.append(df)

combined_df = pd.concat(dfs, axis=0, ignore_index=True)

first_6 = dfs[0].columns[:5].tolist() + [('', 'type')]

# === Step 2: Flatten MultiIndex columns ===
flat_columns = []
for lvl1, lvl2 in combined_df.columns:
    if not lvl1 or str(lvl1).startswith("Unnamed") or pd.isna(lvl1):
        flat_columns.append(str(lvl2).strip())
    else:
        flat_columns.append(f"{str(lvl1).strip()} - {str(lvl2).strip()}")

combined_df.columns = flat_columns  # Replace MultiIndex with flat columns

# === Step 3: Sort columns from the 12th column onward ===
first_12_cols = combined_df.columns[:12].tolist()
rest_cols = sorted(combined_df.columns[12:])  # sort alphabetically

combined_df = combined_df[first_12_cols + rest_cols]

# # === Step 3: Format numeric values in supplier cols ===
# for col in combined_df.columns[12:]:
#     combined_df[col] = combined_df[col].apply(
#         lambda v: "" if pd.isna(v) else ("0" if v == 0 else round(float(v), 4)) if isinstance(v, (int, float, np.number)) else v
#     )

# === Step 3.5: Fix empty cells & enforce float type in cost columns ===
# for col in combined_df.columns[12:]:
#     if "Total Cost Per UOM FOB Port of Origin/Departure" in col:
#         combined_df[col] = pd.to_numeric(combined_df[col], errors='coerce').fillna(0).astype(float)

# combined_df = combined_df.copy()



# === Step 4: Calculate statistics ===
# Build supplier to R1/R2 column mapping
supplier_round_cols = {}
for col in combined_df.columns[12:]:
    if col.endswith("Total Cost Per UOM FOB Port of Origin/Departure (USD)"):
        parts = col.split(" - ")
        if len(parts) >= 3:
            supplier = parts[0].strip()
            round_tag = parts[1].strip()
            if supplier not in supplier_round_cols:
                supplier_round_cols[supplier] = {}
            supplier_round_cols[supplier][round_tag] = col

# For each row, build a dict of supplier: value (R2 if present, else R1)
def get_supplier_values(row):
    values = {}
    for supplier, rounds in supplier_round_cols.items():
        val = None
        if 'R2' in rounds:
            v = row[rounds['R2']]
            if not pd.isna(v) and v != '' and float(v) != 0:
                val = float(v)
        if val is None and 'R1' in rounds:
            v = row[rounds['R1']]
            if not pd.isna(v) and v != '' and float(v) != 0:
                val = float(v)
        if val is not None:
            values[supplier] = val
    return values

outlier_mask = pd.DataFrame(False, index=combined_df.index, columns=combined_df.columns[12:])
iqr_values = [np.nan] * len(combined_df)
means = [""] * len(combined_df)
variances = [""] * len(combined_df)
std_devs = [""] * len(combined_df)

for idx, row in combined_df.iterrows():
    supplier_vals = get_supplier_values(row)
    vals = list(supplier_vals.values())
    if len(vals) == 1:
        values = pd.Series(vals)
        iqr_values[idx] = 0
        means[idx] = round(values.mean(), 4)
        variances[idx] = 0
        std_devs[idx] = 0
        # Outlier mask for each supplier col (no outlier possible with one value)
        continue
    if len(vals) == 0:
        continue
    values = pd.Series(vals)
    Q1 = values.quantile(0.25)
    Q3 = values.quantile(0.75)
    IQR = Q3 - Q1
    lower = Q1 - 1.5 * IQR
    upper = Q3 + 1.5 * IQR
    iqr_values[idx] = IQR
    means[idx] = round(values.mean(), 4)
    variances[idx] = round(values.var(ddof=0), 4)
    std_devs[idx] = round(values.std(ddof=0), 4)
    # Outlier mask for each supplier col
    for supplier, val in supplier_vals.items():
        for round_tag in ['R2', 'R1']:
            if round_tag in supplier_round_cols[supplier]:
                col = supplier_round_cols[supplier][round_tag]
                if val < lower or val > upper:
                    outlier_mask.at[idx, col] = True

# Now safely assign
combined_df["Statistics-Mean (Total Cost)"] = means
combined_df["Statistics-Variance (Total Cost)"] = variances
combined_df["Statistics-Std Dev (Total Cost)"] = std_devs
combined_df["IQR_Value"] = iqr_values

# === Step 5: Mean without outliers ===
mean_without_high_outliers = []
for idx, row in combined_df.iterrows():
    IQR = combined_df.at[idx, "IQR_Value"]
    supplier_vals = get_supplier_values(row)
    values = pd.Series(list(supplier_vals.values()))
    if IQR is not None and not pd.isna(IQR):
        Q3 = values.quantile(0.75)
        upper = Q3 + 1.5 * IQR
        values = values[values <= upper]
    mean_val = round(values.mean(), 4) if not values.empty else ""
    mean_without_high_outliers.append(mean_val)
combined_df["Arithmetic Average"] = mean_without_high_outliers

# === Step 6: Min/2nd Min/Outlier Flag ===
r1_min_bids, r1_min_bids_supplier, min_bids, min_suppliers, second_min_bids, second_min_suppliers, has_outlier_flags = [], [], [], [], [], [], []
for idx, row in combined_df.iterrows():
    supplier_vals = get_supplier_values(row)
    outlier_bids = {}
    for supplier, val in supplier_vals.items():
        for round_tag in ['R2']:
            if round_tag in supplier_round_cols[supplier]:
                col = supplier_round_cols[supplier][round_tag]
                if outlier_mask.at[idx, col]:
                    outlier_bids[supplier] = val

    if supplier_vals:
        sorted_bids = sorted(supplier_vals.items(), key=lambda x: x[1])
        min_supplier, min_bid = sorted_bids[0]
        second_min_supplier, second_min_bid = sorted_bids[1] if len(sorted_bids) > 1 else ("", "")
    else:
        min_bid = min_supplier = second_min_bid = second_min_supplier = ""

    
        

    has_lower_outlier = any(supplier in outlier_bids and outlier_bids[supplier] == min_bid for supplier in outlier_bids)
    min_bids.append(min_bid)
    min_suppliers.append(min_supplier)
    second_min_bids.append(second_min_bid)
    second_min_suppliers.append(second_min_supplier)
    has_outlier_flags.append("Yes" if has_lower_outlier else "No")
    

# combined_df["Min Bid R1"] = r1_min_bids
# combined_df["Minimum Bid Supplier R1"] = r1_min_bids_supplier
combined_df["Final Min Bid"] = min_bids
combined_df["Has Outlier"] = has_outlier_flags
combined_df["Final 2nd Lowest Bid"] = second_min_bids
combined_df["Final Minimum Bid Supplier"] = min_suppliers
combined_df["Final 2nd Lowest Bid Supplier"] = second_min_suppliers

# === Step 7: Reorder columns ===
# Priority columns at start
start_cols = [col for col in CONSOLIDATED_COMMON_COLUMN_NAMES if col in combined_df.columns]

# Then 'type' + 'Valid Supplier' + min bid details
next_cols = ['type', 'Valid Supplier', 'Min Bid R1', 'Minimum Bid Supplier R1', 'Final Min Bid', 'Final Minimum Bid Supplier', 'Has Outlier', 'Final 2nd Lowest Bid', 'Final 2nd Lowest Bid Supplier']
next_cols = [col for col in next_cols if col in combined_df.columns]

# Then supplier columns (excluding the ones already included)
supplier_cols = [col for col in combined_df.columns if col not in start_cols + next_cols and not col.startswith("Statistics") and col != "Arithmetic Average"]

# Stats at end
stat_cols = [
    "Statistics-Mean (Total Cost)",
    "Statistics-Variance (Total Cost)",
    "Statistics-Std Dev (Total Cost)",
    "IQR_Value",
    "Arithmetic Average"
]
# Only include stat_cols that are not already in start_cols + next_cols + supplier_cols
stat_cols = [col for col in stat_cols if col in combined_df.columns and col not in start_cols + next_cols + supplier_cols]

# Final order
final_cols = start_cols + next_cols + supplier_cols + stat_cols
combined_df = combined_df[final_cols]


# Step 8


output_file = "new/bidsheet_master_consolidate.xlsx"
combined_df = combined_df.replace({np.nan: ""})
combined_df.to_excel(output_file, index=False)

print(f"✅ All processing complete. Final output saved to: {output_file}")